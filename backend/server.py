from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form, Request, Body, Response
from fastapi.responses import RedirectResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.middleware.trustedhost import TrustedHostMiddleware
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
import os
import logging
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr, validator
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
from jose import JWTError, jwt
import base64
from emergentintegrations.llm.chat import LlmChat, UserMessage
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
import pandas as pd
from ofxparse import OfxParser
from PyPDF2 import PdfReader
from google.oauth2 import service_account
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request as GoogleRequest
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseUpload, MediaIoBaseDownload
import json
import tempfile
import httpx
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from security_utils import (
    validate_password_strength,
    sanitize_string,
    validate_cnpj,
    validate_cpf,
    validate_phone_number,
    log_security_event,
    check_sql_injection,
    check_brute_force,
    record_failed_login,
    clear_failed_logins
)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection - allow build without real connection
# Will be validated at startup
mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'finai_database')]

# Security
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# Integrations API Keys
ASAAS_API_KEY = os.environ.get('ASAAS_API_KEY', 'MOCK_ASAAS_KEY_SUBSTITUIR_PELA_REAL')
ASAAS_BASE_URL = os.environ.get('ASAAS_BASE_URL', 'https://www.asaas.com/api/v3')

# Gmail SMTP
GMAIL_USER = os.environ.get('GMAIL_USER', 'faraujoneto2025@gmail.com')
GMAIL_APP_PASSWORD = os.environ.get('GMAIL_APP_PASSWORD', '')  # Senha de app do Google

# WhatsApp Meta
META_ACCESS_TOKEN = os.environ.get('META_ACCESS_TOKEN', 'MOCK_META_TOKEN')
META_PHONE_NUMBER_ID = os.environ.get('META_PHONE_NUMBER_ID', '')
META_WHATSAPP_FROM = os.environ.get('META_WHATSAPP_FROM', '5511999999999')

# Rate limiting
limiter = Limiter(key_func=get_remote_address)

# Create the main app
app = FastAPI()

# Add rate limiter
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# Initialize APScheduler for automated backups
scheduler = AsyncIOScheduler()

# Environment variables - with safe defaults for build time
# Validation happens at startup (see @app.on_event("startup") below)
JWT_SECRET = os.environ.get('JWT_SECRET', 'temp-build-secret')
JWT_ALGORITHM = os.environ.get('JWT_ALGORITHM', 'HS256')
JWT_EXPIRATION = int(os.environ.get('JWT_EXPIRATION_MINUTES', 43200))
EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY')
WHATSAPP_SERVICE_KEY = os.environ.get('WHATSAPP_SERVICE_KEY', 'temp-build-key')

# API Router
api_router = APIRouter(prefix="/api")

# ==================== MODELS ====================

class UserProfile(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nome: str
    email: EmailStr
    telefone: Optional[str] = None
    perfil: str  # admin, financeiro, leitura
    empresa_ids: List[str] = []
    senha_hash: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserResponse(BaseModel):
    """User response without password hash"""
    model_config = ConfigDict(extra="ignore")
    id: str
    nome: str
    email: EmailStr
    telefone: Optional[str] = None
    perfil: str
    empresa_ids: List[str] = []
    created_at: datetime

# Perfis de usuário e suas permissões
PERFIS_PERMISSOES = {
    "admin": {
        "nome": "Admin da Empresa",
        "descricao": "Acesso total ao sistema",
        "permissoes": ["*"]  # Todas as permissões
    },
    "admin_master": {
        "nome": "Admin Master",
        "descricao": "Acesso total ao sistema e licenciamento",
        "permissoes": ["*"]  # Todas as permissões
    },
    "financeiro": {
        "nome": "Financeiro",
        "descricao": "Acesso a transações, relatórios e contas",
        "permissoes": ["transacoes", "relatorios", "categorias", "contas", "fornecedores", "centros_custo"]
    },
    "vendas": {
        "nome": "Vendas",
        "descricao": "Acesso a vendas, clientes e estoque",
        "permissoes": ["vendas", "clientes", "estoque", "produtos"]
    },
    "operacional": {
        "nome": "Operacional",
        "descricao": "Acesso a estoque e equipamentos",
        "permissoes": ["estoque", "equipamentos", "movimentacoes"]
    },
    "consulta": {
        "nome": "Consulta",
        "descricao": "Apenas visualização",
        "permissoes": ["visualizar"]
    }
}

# ==================== PLANOS SAAS ====================
PLANOS_SAAS = {
    "basico": {
        "nome": "Básico",
        "valor": 99.00,
        "descricao": "Ideal para pequenas empresas",
        "recursos": [
            "Até 100 transações/mês",
            "1 usuário",
            "Relatórios básicos",
            "Suporte por email"
        ]
    },
    "profissional": {
        "nome": "Profissional",
        "valor": 199.00,
        "descricao": "Para empresas em crescimento",
        "recursos": [
            "Transações ilimitadas",
            "5 usuários",
            "Relatórios avançados",
            "Suporte prioritário",
            "Integração WhatsApp",
            "Backup automático"
        ]
    }
}

class UserCreate(BaseModel):
    nome: str
    email: EmailStr
    telefone: Optional[str] = None
    perfil: str = "financeiro"
    senha: str
    empresa_ids: List[str] = []
    
    @validator('nome')
    def sanitize_nome(cls, v):
        return sanitize_string(v, max_length=100)
    
    @validator('telefone')
    def validate_telefone(cls, v):
        if v and not validate_phone_number(v):
            raise ValueError('Número de telefone inválido')
        return v
    
    @validator('perfil')
    def validate_perfil(cls, v):
        if v not in PERFIS_PERMISSOES:
            raise ValueError(f'Perfil inválido. Opções: {", ".join(PERFIS_PERMISSOES.keys())}')
        return v
    
    @validator('senha')
    def validate_senha_strength(cls, v):
        is_valid, error_msg = validate_password_strength(v)
        if not is_valid:
            raise ValueError(error_msg)
        return v

class UserLogin(BaseModel):
    email: EmailStr
    senha: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: Dict[str, Any]

# Sistema de Logs
class LogAcao(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_email: str
    empresa_id: str
    acao: str  # login, logout, criar_transacao, editar_cliente, etc
    modulo: str  # financeiro, vendas, estoque, etc
    detalhes: dict = {}  # Dados adicionais da ação
    ip_address: Optional[str] = None
    user_agent: Optional[str] = None
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class LogSessao(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_email: str
    empresa_id: str
    login_at: datetime
    logout_at: Optional[datetime] = None
    duracao_segundos: Optional[int] = None
    ip_address: Optional[str] = None

# Sistema de Licenciamento
class Licenca(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    plano: str  # "basico" ou "pro"
    valor_mensal: float  # 99.90 ou 139.90
    asaas_subscription_id: Optional[str] = None
    asaas_customer_id: Optional[str] = None
    status: str = "ativa"  # ativa, bloqueada, cancelada
    data_vencimento: Optional[datetime] = None
    dias_atraso: int = 0
    bloqueada_em: Optional[datetime] = None
    motivo_bloqueio: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ==================== ASSINATURAS SAAS ====================
class AssinaturaSaaS(BaseModel):
    """Assinatura de cliente do sistema SaaS"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    # Dados da empresa cliente
    razao_social: str
    cnpj_cpf: str
    email: str
    telefone: str
    # Dados do plano
    plano: str  # "basico" ou "profissional"
    valor_mensal: float
    # Dados do usuário criado
    user_id: str
    user_email: str
    user_senha_temp: str  # Senha temporária gerada
    # Integração Asaas
    asaas_customer_id: Optional[str] = None
    asaas_subscription_id: Optional[str] = None
    # Status
    status: str = "aguardando_pagamento"  # aguardando_pagamento, ativa, suspensa, cancelada
    dia_vencimento: int  # Dia do mês para cobranças futuras
    # Controle de pagamento
    primeiro_pagamento_pix: bool = True
    pix_qrcode: Optional[str] = None
    pix_codigo: Optional[str] = None
    pix_payment_id: Optional[str] = None
    primeiro_pagamento_status: str = "pendente"  # pendente, pago
    # Controle de inadimplência
    dias_atraso: int = 0
    ultimo_email_cobranca: Optional[str] = None
    bloqueada_em: Optional[str] = None
    # Vendedor
    vendedor_id: str
    vendedor_nome: str
    # Datas
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AssinaturaSaaSCreate(BaseModel):
    razao_social: str
    cnpj_cpf: str
    email: EmailStr
    telefone: str
    plano: str  # "basico" ou "profissional"

class Cobranca(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    licenca_id: str
    empresa_id: str
    asaas_payment_id: Optional[str] = None
    valor: float
    data_vencimento: datetime
    data_pagamento: Optional[datetime] = None
    status: str = "pendente"  # pendente, pago, atrasado, cancelado
    metodo_pagamento: str  # boleto, pix
    boleto_url: Optional[str] = None
    boleto_codigo: Optional[str] = None
    pix_qrcode: Optional[str] = None
    pix_codigo: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ConfiguracaoGateway(BaseModel):
    """Configuração de gateway de pagamento do cliente"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    gateway: str  # asaas, mercadopago, pagseguro
    api_key: str
    sandbox_mode: bool = False
    ativo: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Venda(BaseModel):
    """Venda realizada pelo cliente"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    cliente_nome: str
    cliente_cpf_cnpj: str
    cliente_email: Optional[str] = None
    cliente_telefone: Optional[str] = None
    valor_total: float
    descricao: str
    gateway_payment_id: Optional[str] = None  # ID no gateway de pagamento
    metodo_pagamento: str = "boleto"  # boleto, pix, cartao
    status: str = "pendente"  # pendente, pago, cancelado
    boleto_url: Optional[str] = None
    boleto_codigo: Optional[str] = None
    pix_qrcode: Optional[str] = None
    pix_codigo: Optional[str] = None
    data_vencimento: str  # ISO date string
    data_pagamento: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Empresa(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    razao_social: str
    cnpj: str
    contas_bancarias: List[str] = []
    ativa: bool = True
    bloqueada: bool = False
    motivo_bloqueio: str = ""
    # Dados para Asaas (licenciamento)
    asaas_customer_id: Optional[str] = None
    email_cobranca: Optional[str] = None
    telefone_cobranca: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EmpresaCreate(BaseModel):
    razao_social: str
    cnpj: str
    contas_bancarias: List[str] = []
    
    @validator('razao_social')
    def sanitize_razao_social(cls, v):
        return sanitize_string(v, max_length=200)
    
    @validator('cnpj')
    def validate_cnpj_format(cls, v):
        if not validate_cnpj(v):
            raise ValueError('CNPJ inválido')
        return v

class Categoria(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nome: str
    tipo: str  # despesa, receita
    descricao: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CategoriaCreate(BaseModel):
    nome: str
    tipo: str
    descricao: Optional[str] = None

class CentroCusto(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nome: str
    area: Optional[str] = None  # Comercial, Operação, TI, etc.
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CentroCustoCreate(BaseModel):
    nome: str
    area: Optional[str] = None

# CONTAS BANCÁRIAS
class ContaBancaria(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nome: str
    tipo: str  # corrente, poupanca, caixa
    banco: str
    agencia: Optional[str] = None
    numero_conta: Optional[str] = None
    saldo_inicial: float
    saldo_atual: float
    ativa: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ContaBancariaCreate(BaseModel):
    nome: str
    tipo: str
    banco: str
    agencia: Optional[str] = None
    numero_conta: Optional[str] = None
    saldo_inicial: float

# INVESTIMENTOS
class Investimento(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nome: str
    tipo: str  # renda_fixa, renda_variavel, fundos, outros
    valor_investido: float
    valor_atual: float
    data_aplicacao: str
    data_vencimento: Optional[str] = None
    rentabilidade_percentual: float = 0.0
    instituicao: Optional[str] = None
    ativo: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class InvestimentoCreate(BaseModel):
    nome: str
    tipo: str
    valor_investido: float
    valor_atual: float
    data_aplicacao: str
    data_vencimento: Optional[str] = None
    rentabilidade_percentual: float = 0.0
    instituicao: Optional[str] = None

# CARTÕES DE CRÉDITO
class CartaoCredito(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nome: str
    bandeira: str  # Visa, Mastercard, Elo, etc
    limite_total: float
    limite_disponivel: float
    dia_fechamento: int  # 1-31
    dia_vencimento: int  # 1-31
    fatura_atual: float = 0.0
    conta_debito_id: Optional[str] = None  # Conta para débito automático
    ativo: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CartaoCreditoCreate(BaseModel):
    nome: str
    bandeira: str
    limite_total: float
    dia_fechamento: int
    dia_vencimento: int
    conta_debito_id: Optional[str] = None

# FATURAS DE CARTÃO
class FaturaCartao(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    cartao_id: str
    empresa_id: str
    mes_referencia: str  # YYYY-MM
    valor_total: float
    valor_pago: float = 0.0
    data_fechamento: str
    data_vencimento: str
    status: str = "aberta"  # aberta, paga, vencida
    transacoes_ids: List[str] = []
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ==================== ESTOQUE MODELS ====================

# CLIENTES
class Cliente(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nome: str
    tipo: str  # fisica, juridica
    cnpj_cpf: str
    email: Optional[str] = None
    telefone: Optional[str] = None
    endereco: Optional[str] = None
    cidade: Optional[str] = None
    estado: Optional[str] = None
    cep: Optional[str] = None
    status: str = "ativo"  # ativo, inativo
    observacoes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ClienteCreate(BaseModel):
    nome: str
    tipo: str
    cnpj_cpf: str
    email: Optional[str] = None
    telefone: Optional[str] = None
    endereco: Optional[str] = None
    cidade: Optional[str] = None
    estado: Optional[str] = None
    cep: Optional[str] = None
    observacoes: Optional[str] = None

# FORNECEDORES
class Fornecedor(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nome: str
    cnpj: str
    contato: Optional[str] = None
    email: Optional[str] = None
    telefone: Optional[str] = None
    endereco: Optional[str] = None
    status: str = "ativo"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FornecedorCreate(BaseModel):
    nome: str
    cnpj: str
    contato: Optional[str] = None
    email: Optional[str] = None
    telefone: Optional[str] = None
    endereco: Optional[str] = None

# LOCAIS/DEPÓSITOS
class LocalDeposito(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nome: str
    descricao: Optional[str] = None
    responsavel: Optional[str] = None
    endereco: Optional[str] = None
    status: str = "ativo"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class LocalDepositoCreate(BaseModel):
    nome: str
    descricao: Optional[str] = None
    responsavel: Optional[str] = None
    endereco: Optional[str] = None

# CATEGORIAS DE EQUIPAMENTOS
class CategoriaEquipamento(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nome: str
    descricao: Optional[str] = None
    tipo_controle: str  # serializado, nao_serializado
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CategoriaEquipamentoCreate(BaseModel):
    nome: str
    descricao: Optional[str] = None
    tipo_controle: str

# EQUIPAMENTOS (PRODUTOS)
class Equipamento(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nome: str
    categoria_id: str
    fabricante: Optional[str] = None
    modelo: Optional[str] = None
    descricao: Optional[str] = None
    custo_aquisicao: float = 0.0
    valor_venda: float = 0.0
    valor_locacao_mensal: float = 0.0
    tipo_controle: str  # serializado, nao_serializado
    foto_url: Optional[str] = None
    fornecedor_id: Optional[str] = None
    # Para não-serializados
    quantidade_estoque: int = 0
    estoque_minimo: int = 0
    status: str = "ativo"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EquipamentoCreate(BaseModel):
    nome: str
    categoria_id: str
    fabricante: Optional[str] = None
    modelo: Optional[str] = None
    descricao: Optional[str] = None
    custo_aquisicao: float = 0.0
    valor_venda: float = 0.0
    valor_locacao_mensal: float = 0.0
    tipo_controle: str
    foto_url: Optional[str] = None
    fornecedor_id: Optional[str] = None
    estoque_minimo: int = 0

# EQUIPAMENTOS SERIALIZADOS (INSTÂNCIAS)
class EquipamentoSerializado(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    equipamento_id: str
    numero_serie: str
    numero_linha: Optional[str] = None
    numero_simcard: Optional[str] = None
    historico_simcards: List[Dict[str, str]] = []  # [{"numero": "123", "data_troca": "2024-01-01"}]
    status: str = "disponivel"  # disponivel, em_cliente, em_manutencao, vendido, baixado
    cliente_id: Optional[str] = None
    tipo_vinculo: Optional[str] = None  # venda, locacao, comodato
    local_id: Optional[str] = None
    data_aquisicao: Optional[str] = None
    data_garantia: Optional[str] = None
    custo_especifico: Optional[float] = None
    observacoes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EquipamentoSerializadoCreate(BaseModel):
    equipamento_id: str
    numero_serie: str
    numero_linha: Optional[str] = None
    numero_simcard: Optional[str] = None
    local_id: Optional[str] = None
    data_aquisicao: Optional[str] = None
    data_garantia: Optional[str] = None
    custo_especifico: Optional[float] = None
    observacoes: Optional[str] = None

# MOVIMENTAÇÕES DE ESTOQUE
class MovimentacaoEstoque(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    tipo: str  # entrada, saida_venda, saida_locacao, devolucao, transferencia, perda, manutencao
    data: str  # ISO date
    equipamento_id: str
    equipamento_serializado_id: Optional[str] = None  # Para serializados
    quantidade: int = 1  # Para não-serializados
    cliente_id: Optional[str] = None
    local_origem_id: Optional[str] = None
    local_destino_id: Optional[str] = None
    observacoes: Optional[str] = None
    usuario_id: str
    valor_financeiro: Optional[float] = None
    criar_transacao_financeira: bool = False
    transacao_id: Optional[str] = None  # ID da transação financeira criada
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MovimentacaoEstoqueCreate(BaseModel):
    tipo: str
    data: str
    equipamento_id: str
    equipamento_serializado_id: Optional[str] = None
    quantidade: int = 1
    cliente_id: Optional[str] = None
    local_origem_id: Optional[str] = None
    local_destino_id: Optional[str] = None
    observacoes: Optional[str] = None
    valor_financeiro: Optional[float] = None
    criar_transacao_financeira: bool = False
    categoria_financeira_id: Optional[str] = None  # Para criação de transação
    centro_custo_id: Optional[str] = None

# ==================== END ESTOQUE MODELS ====================

# ==================== CRM MODELS ====================

# LEADS / CONTATOS
class Lead(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nome: str
    telefone: str
    email: Optional[str] = None
    origem: str = "manual"  # whatsapp, manual, importacao, landing, indicacao
    status_funil: str = "novo"  # novo, contatado, qualificado, proposta, negociacao, ganho, perdido
    tags: List[str] = []
    valor_estimado: float = 0.0
    assigned_to: Optional[str] = None  # user_id do vendedor
    whatsapp_phone: Optional[str] = None  # Número formatado do WhatsApp
    last_contact_at: Optional[datetime] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class LeadCreate(BaseModel):
    nome: str
    telefone: str
    email: Optional[str] = None
    origem: str = "manual"
    tags: List[str] = []
    valor_estimado: float = 0.0
    assigned_to: Optional[str] = None
    notes: Optional[str] = None

class LeadUpdate(BaseModel):
    nome: Optional[str] = None
    telefone: Optional[str] = None
    email: Optional[str] = None
    tags: Optional[List[str]] = None
    valor_estimado: Optional[float] = None
    assigned_to: Optional[str] = None
    notes: Optional[str] = None

# ATIVIDADES DO CRM
class Activity(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    lead_id: str
    empresa_id: str
    tipo: str  # note, call, email, whatsapp, status_change, assignment
    descricao: str
    user_id: str
    metadata: Dict[str, Any] = {}
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# TEMPLATES DE MENSAGEM
class MessageTemplate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nome: str
    conteudo: str  # Pode ter variáveis: {nome}, {valor}, {empresa}
    tipo: str = "whatsapp"  # whatsapp, email, sms
    ativo: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MessageTemplateCreate(BaseModel):
    nome: str
    conteudo: str
    tipo: str = "whatsapp"

# REGRAS DE AUTOMAÇÃO
class AutomationRule(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nome: str
    gatilho: str  # novo_lead, mudanca_status, sem_resposta
    condicoes: Dict[str, Any] = {}  # {"origem": "whatsapp", "status": "novo"}
    acoes: List[Dict[str, Any]] = []  # [{"tipo": "enviar_template", "template_id": "xxx"}]
    ativo: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AutomationRuleCreate(BaseModel):
    nome: str
    gatilho: str
    condicoes: Dict[str, Any] = {}
    acoes: List[Dict[str, Any]] = []

# CONFIGURAÇÃO DE ROTEAMENTO
class RoutingConfig(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    tipo: str = "round_robin"  # round_robin, por_tag, manual
    usuarios_ativos: List[str] = []  # Lista de user_ids
    ultimo_atribuido: Optional[str] = None
    regras_por_tag: Dict[str, str] = {}  # {"tag": "user_id"}
    ativo: bool = True

# AGENTE IA
class AIAgent(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nome: str
    prompt_sistema: str
    palavras_chave_ativacao: List[str] = []  # ["orçamento", "preço", "quanto custa"]
    horario_ativo: Dict[str, Any] = {"inicio": "09:00", "fim": "18:00"}
    autonomia: str = "baixa"  # baixa, media, alta
    ativo: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AIAgentCreate(BaseModel):
    nome: str
    prompt_sistema: str
    palavras_chave_ativacao: List[str] = []
    horario_ativo: Dict[str, Any] = {"inicio": "09:00", "fim": "18:00"}
    autonomia: str = "baixa"

# SEQUÊNCIAS DE AUTOMAÇÃO
class FollowUpSequence(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nome: str
    gatilho_status: str  # Status que ativa a sequência
    steps: List[Dict[str, Any]] = []  # [{"dia": 1, "template_id": "xxx", "tipo": "whatsapp"}]
    ativo: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FollowUpSequenceCreate(BaseModel):
    nome: str
    gatilho_status: str
    steps: List[Dict[str, Any]] = []

# MÉTRICAS E RELATÓRIOS
class CRMMetrics(BaseModel):
    total_leads: int
    leads_por_status: Dict[str, int]
    leads_por_origem: Dict[str, int]
    taxa_conversao: Dict[str, float]
    tempo_medio_por_etapa: Dict[str, float]
    valor_total_pipeline: float
    leads_vencidos_mes: int
    leads_perdidos_mes: int
    desempenho_vendedores: List[Dict[str, Any]]

# ==================== END CRM MODELS ====================

# ==================== VENDAS E PLANOS MODELS ====================

# PLANO DE INTERNET
class PlanoInternet(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nome: str  # Ex: "100MB Fibra"
    velocidade_download: str  # Ex: "100 Mbps"
    velocidade_upload: str  # Ex: "50 Mbps"
    preco_mensal: float
    descricao: Optional[str] = None
    beneficios: List[str] = []  # ["WiFi grátis", "Instalação grátis"]
    ativo: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PlanoInternetCreate(BaseModel):
    nome: str
    velocidade_download: str
    velocidade_upload: str
    preco_mensal: float
    descricao: Optional[str] = None
    beneficios: List[str] = []

# CLIENTE DE VENDAS
class ClienteVenda(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nome_completo: str
    cpf: str
    email: str
    telefone: str
    # Endereço
    cep: str
    logradouro: str
    numero: str
    complemento: Optional[str] = None
    bairro: str
    cidade: str
    estado: str
    # Status
    status: str = "ativo"  # ativo, suspenso, cancelado
    inadimplente: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ClienteVendaCreate(BaseModel):
    nome_completo: str
    cpf: str
    email: str
    telefone: str
    cep: str
    logradouro: str
    numero: str
    complemento: Optional[str] = None
    bairro: str
    cidade: str
    estado: str

# VENDA / CONTRATO (Assinaturas recorrentes)
class VendaContrato(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    cliente_id: str
    plano_id: str
    # Dados do contrato
    data_contratacao: str  # ISO date
    dia_vencimento: int  # 1-31
    status: str = "ativo"  # ativo, suspenso, cancelado
    valor_mensalidade: float
    # Observações
    observacoes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class VendaContratoCreate(BaseModel):
    cliente_id: str
    plano_id: str
    data_contratacao: str
    dia_vencimento: int
    observacoes: Optional[str] = None

# FATURA / COBRANÇA
class Fatura(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    venda_id: str
    cliente_id: str
    # Dados da fatura
    mes_referencia: str  # "2025-02" formato YYYY-MM
    data_vencimento: str  # ISO date
    valor: float
    multa: float = 0.0
    juros: float = 0.0
    valor_total: float  # valor + multa + juros
    # Status e pagamento
    status: str = "pendente"  # pendente, pago, vencido, cancelado
    data_pagamento: Optional[str] = None
    # Boleto Galax Pay
    boleto_id: Optional[str] = None  # ID no Galax Pay
    boleto_url: Optional[str] = None
    boleto_linha_digitavel: Optional[str] = None
    boleto_codigo_barras: Optional[str] = None
    # Envios
    emails_enviados: List[Dict[str, Any]] = []  # [{"data": "2025-01-01", "tipo": "primeira_cobranca"}]
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FaturaCreate(BaseModel):
    venda_id: str
    mes_referencia: str
    data_vencimento: str
    valor: float

# CONFIGURAÇÕES DE COBRANÇA
class ConfiguracaoCobranca(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    # Galax Pay
    galax_id: str
    galax_hash: str
    ambiente: str = "sandbox"  # sandbox, producao
    # Multa e juros
    percentual_multa: float = 2.0  # 2%
    percentual_juros_dia: float = 0.033  # 0.033% ao dia = 1% ao mês
    # Envios de email
    dias_envio_antecipado: List[int] = [15, 10, 5, 0]  # Enviar 15, 10, 5 dias antes e no dia
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ConfiguracaoCobrancaCreate(BaseModel):
    galax_id: str
    galax_hash: str
    ambiente: str = "sandbox"
    percentual_multa: float = 2.0
    percentual_juros_dia: float = 0.033
    dias_envio_antecipado: List[int] = [15, 10, 5, 0]

# ==================== END VENDAS MODELS ====================

class Transacao(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    usuario_id: str
    tipo: str  # despesa, receita
    fornecedor: str
    fornecedor_id: Optional[str] = None  # ID do fornecedor cadastrado
    cnpj_cpf: Optional[str] = None
    descricao: str
    valor_total: float
    data_competencia: str  # ISO date
    data_pagamento: Optional[str] = None  # ISO date
    categoria_id: str
    centro_custo_id: str
    metodo_pagamento: Optional[str] = None
    conta_origem: Optional[str] = None
    conta_bancaria_id: Optional[str] = None  # Vinculo com conta bancária
    cartao_credito_id: Optional[str] = None  # Vinculo com cartão
    impostos: Optional[Dict[str, float]] = {}
    parcelas: Optional[Dict[str, Any]] = None
    comprovante_url: Optional[str] = None
    status: str = "pendente"  # pendente, conciliada, contestada
    origem: str = "manual"  # manual, whatsapp, upload, api
    is_transferencia: bool = False  # Flag para identificar transferências
    transferencia_relacionada_id: Optional[str] = None  # ID da transação relacionada
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TransacaoCreate(BaseModel):
    tipo: str
    fornecedor: str
    fornecedor_id: Optional[str] = None  # ID do fornecedor cadastrado
    cnpj_cpf: Optional[str] = None
    descricao: str
    valor_total: float
    data_competencia: str
    data_pagamento: Optional[str] = None
    categoria_id: str
    centro_custo_id: str
    metodo_pagamento: Optional[str] = None
    conta_origem: Optional[str] = None
    conta_bancaria_id: Optional[str] = None
    cartao_credito_id: Optional[str] = None
    impostos: Optional[Dict[str, float]] = None
    parcelas: Optional[Dict[str, Any]] = None
    status: str = "pendente"
    origem: str = "manual"
    
    @validator('fornecedor', 'descricao', 'metodo_pagamento', 'conta_origem')
    def sanitize_text_fields(cls, v):
        if v:
            return sanitize_string(v, max_length=500)
        return v
    
    @validator('cnpj_cpf')
    def validate_document(cls, v):
        if v:
            if not (validate_cnpj(v) or validate_cpf(v)):
                raise ValueError('CNPJ/CPF inválido')
        return v
    
    @validator('valor_total')
    def validate_valor(cls, v):
        if v <= 0:
            raise ValueError('Valor deve ser maior que zero')
        if v > 999999999.99:
            raise ValueError('Valor muito alto')
        return v

class ClassificacaoSugestao(BaseModel):
    categoria_id: str
    categoria_nome: str
    centro_custo_id: str
    centro_custo_nome: str
    confidence: float
    reasoning: str

class ExtrairTextoRequest(BaseModel):
    texto: str
    empresa_id: str

class ExtrairTextoResponse(BaseModel):
    dados_extraidos: Dict[str, Any]
    classificacao_sugerida: Optional[ClassificacaoSugestao] = None

class DashboardMetrics(BaseModel):
    total_receitas: float
    total_despesas: float
    saldo: float
    saldo_contas: float = 0.0
    saldo_cartoes: float = 0.0
    num_contas: int = 0
    num_cartoes: int = 0
    despesas_por_categoria: List[Dict[str, Any]]
    despesas_por_centro_custo: List[Dict[str, Any]]
    transacoes_recentes: List[Dict[str, Any]]

class CentroCustoMetrics(BaseModel):
    centro_custo_id: str
    centro_custo_nome: str
    total_receitas: float
    total_despesas: float
    lucro: float
    num_transacoes: int
    percentual_total: float

class CategoriaMetrics(BaseModel):
    categoria_id: str
    categoria_nome: str
    total_receitas: float
    total_despesas: float
    num_transacoes: int
    percentual_despesas: float
    percentual_receitas: float

class RelatorioDetalhado(BaseModel):
    periodo_inicio: str
    periodo_fim: str
    resumo_geral: Dict[str, Any]
    por_centro_custo: List[CentroCustoMetrics]
    por_categoria: List[CategoriaMetrics]
    transacoes: List[Dict[str, Any]]

# ==================== AUTH HELPERS ====================

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=JWT_EXPIRATION)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ==================== PERMISSÕES ====================

def verificar_permissao(user: dict, permissao: str) -> bool:
    """Verifica se o usuário tem a permissão necessária"""
    perfil = user.get("perfil", "consulta")
    
    # Admin e admin_master têm todas as permissões
    if perfil in ["admin", "admin_master"]:
        return True
    
    # Pega as permissões do perfil
    permissoes = PERFIS_PERMISSOES.get(perfil, {}).get("permissoes", [])
    
    # Verifica se tem a permissão específica ou todas (*)
    return "*" in permissoes or permissao in permissoes

def requer_permissao(permissao: str):
    """Decorator para verificar permissão em rotas"""
    def decorator(func):
        async def wrapper(*args, current_user: dict = None, **kwargs):
            if current_user and not verificar_permissao(current_user, permissao):
                raise HTTPException(
                    status_code=403, 
                    detail=f"Permissão negada. Necessário: {permissao}"
                )
            return await func(*args, current_user=current_user, **kwargs)
        return wrapper
    return decorator

# ==================== LOGS ====================

async def registrar_acao(
    user_id: str,
    user_email: str,
    empresa_id: str,
    acao: str,
    modulo: str,
    detalhes: dict = {},
    request: Request = None
):
    """Registra uma ação do usuário no log"""
    log = LogAcao(
        user_id=user_id,
        user_email=user_email,
        empresa_id=empresa_id,
        acao=acao,
        modulo=modulo,
        detalhes=detalhes,
        ip_address=request.client.host if request else None,
        user_agent=request.headers.get("user-agent") if request else None
    )
    
    doc = log.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    await db.logs_acoes.insert_one(doc)

async def iniciar_sessao(user_id: str, user_email: str, empresa_id: str, ip_address: str = None):
    """Inicia uma nova sessão de usuário"""
    sessao = LogSessao(
        user_id=user_id,
        user_email=user_email,
        empresa_id=empresa_id,
        login_at=datetime.now(timezone.utc),
        ip_address=ip_address
    )
    
    doc = sessao.model_dump()
    doc['login_at'] = doc['login_at'].isoformat()
    await db.logs_sessoes.insert_one(doc)
    return sessao.id

async def finalizar_sessao(sessao_id: str):
    """Finaliza uma sessão de usuário"""
    sessao = await db.logs_sessoes.find_one({"id": sessao_id}, {"_id": 0})
    if sessao:
        logout_at = datetime.now(timezone.utc)
        login_at = datetime.fromisoformat(sessao['login_at'])
        duracao = int((logout_at - login_at).total_seconds())
        
        await db.logs_sessoes.update_one(
            {"id": sessao_id},
            {"$set": {
                "logout_at": logout_at.isoformat(),
                "duracao_segundos": duracao
            }}
        )

# ==================== INTEGRAÇÕES ====================

async def asaas_request(method: str, endpoint: str, data: dict = None):
    """Faz requisição para API do Asaas"""
    url = f"{ASAAS_BASE_URL}/{endpoint}"
    headers = {
        "access_token": ASAAS_API_KEY,
        "Content-Type": "application/json"
    }
    
    # Mock mode - retorna dados simulados
    if ASAAS_API_KEY.startswith('MOCK'):
        return {
            "id": f"mock_{str(uuid.uuid4())[:8]}",
            "success": True,
            "message": "Mock response - Integração funcionará quando substituir pela API key real"
        }
    
    async with httpx.AsyncClient() as client:
        if method == "POST":
            response = await client.post(url, json=data, headers=headers)
        elif method == "GET":
            response = await client.get(url, headers=headers)
        elif method == "PUT":
            response = await client.put(url, json=data, headers=headers)
        elif method == "DELETE":
            response = await client.delete(url, headers=headers)
        
        response.raise_for_status()
        return response.json()

async def enviar_email(destinatario: str, assunto: str, html: str):
    """Envia email via Gmail SMTP"""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    
    # Mock mode se não tiver senha configurada
    if not GMAIL_APP_PASSWORD:
        print(f"[MOCK EMAIL] Para: {destinatario}, Assunto: {assunto}")
        return {"id": "mock_email_id", "success": True}
    
    try:
        # Criar mensagem
        msg = MIMEMultipart('alternative')
        msg['Subject'] = assunto
        msg['From'] = f"Sistema Financeiro <{GMAIL_USER}>"
        msg['To'] = destinatario
        
        # Adicionar HTML
        html_part = MIMEText(html, 'html')
        msg.attach(html_part)
        
        # Enviar via Gmail SMTP
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
            server.send_message(msg)
        
        print(f"✅ Email enviado para {destinatario}")
        return {"id": f"gmail_{destinatario}", "success": True}
    except Exception as e:
        print(f"❌ Erro ao enviar email: {str(e)}")
        return {"id": None, "success": False, "error": str(e)}

async def enviar_whatsapp(numero: str, mensagem: str):
    """Envia mensagem via Twilio WhatsApp"""
    # Mock mode
    if TWILIO_ACCOUNT_SID.startswith('MOCK'):
        print(f"[MOCK WHATSAPP] Para: {numero}, Msg: {mensagem}")
        return {"sid": "mock_whatsapp_sid", "success": True}
    
    from twilio.rest import Client
    client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
    
    message = client.messages.create(
        from_=TWILIO_WHATSAPP_FROM,
        body=mensagem,
        to=f"whatsapp:+55{numero}"
    )
    
    return {"sid": message.sid, "success": True}

# ==================== AI HELPERS ====================

async def extrair_dados_com_ai(texto: str, empresa_id: str) -> Dict[str, Any]:
    """Extract financial data from text using AI"""
    try:
        import re
        from datetime import date
        
        # Simple regex extraction as fallback/primary method
        dados = {
            "tipo": "despesa",
            "fornecedor": None,
            "cnpj_cpf": None,
            "descricao": texto,
            "valor_total": None,
            "data_competencia": date.today().isoformat(),
            "metodo_pagamento": None
        }
        
        # Extract valor (R$ 100,00 or R$ 100.00 or 100,00 or 100.00)
        valor_match = re.search(r'R?\$?\s*(\d+[\.,]\d{2})', texto.replace('.', '').replace(',', '.'))
        if valor_match:
            valor_str = valor_match.group(1).replace(',', '.')
            dados["valor_total"] = float(valor_str)
        
        # Extract fornecedor (words after "de/da/do/para/no/na")
        fornecedor_match = re.search(r'(?:de|da|do|para|no|na)\s+([A-Z][A-Za-z\s]+?)(?:\s*[,.]|\s+R\$|\s+\d)', texto, re.IGNORECASE)
        if fornecedor_match:
            dados["fornecedor"] = fornecedor_match.group(1).strip()
        
        # Detect tipo
        if any(word in texto.lower() for word in ['recebi', 'recebimento', 'venda', 'pagamento de cliente']):
            dados["tipo"] = "receita"
        
        # Extract date
        if 'hoje' in texto.lower():
            dados["data_competencia"] = date.today().isoformat()
        elif 'ontem' in texto.lower():
            from datetime import timedelta
            dados["data_competencia"] = (date.today() - timedelta(days=1)).isoformat()
        
        # Extract payment method
        if any(word in texto.lower() for word in ['pix', 'cartão', 'boleto', 'dinheiro', 'débito', 'crédito']):
            metodo_match = re.search(r'(pix|cartão|boleto|dinheiro|débito|crédito)', texto, re.IGNORECASE)
            if metodo_match:
                dados["metodo_pagamento"] = metodo_match.group(1).capitalize()
        
        return dados
    except Exception as e:
        logging.error(f"Error extracting data: {e}")
        return {}

async def classificar_com_ai(descricao: str, fornecedor: str, valor: float, empresa_id: str) -> Optional[ClassificacaoSugestao]:
    """Classify transaction using AI"""
    try:
        # Get categories and cost centers
        categorias = await db.categorias.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(100)
        centros_custo = await db.centros_custo.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(100)
        
        if not categorias or not centros_custo:
            return None
        
        cat_list = "\n".join([f"- {c['nome']} (ID: {c['id']}, Tipo: {c['tipo']})" for c in categorias])
        cc_list = "\n".join([f"- {c['nome']} (ID: {c['id']}, Área: {c.get('area', 'N/A')})" for c in centros_custo])
        
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"classify-{uuid.uuid4()}",
            system_message=f"""Você é um classificador financeiro. Classifique a transação na melhor categoria e centro de custo.
            
            CATEGORIAS DISPONÍVEIS:
            {cat_list}
            
            CENTROS DE CUSTO DISPONÍVEIS:
            {cc_list}
            
            Analise a descrição, fornecedor e valor, e retorne APENAS um JSON com:
            {{"categoria_id": "...", "categoria_nome": "...", "centro_custo_id": "...", "centro_custo_nome": "...", "confidence": 0.95, "reasoning": "..."}}            
            """
        ).with_model("openai", "gpt-5")
        
        message = UserMessage(text=f"Descrição: {descricao}\nFornecedor: {fornecedor}\nValor: R$ {valor:.2f}")
        response = await chat.send_message(message)
        
        import json
        response_text = response.strip()
        if response_text.startswith("```"):
            lines = response_text.split("\n")
            response_text = "\n".join(lines[1:-1])
        if response_text.startswith("json"):
            response_text = response_text[4:].strip()
        
        dados = json.loads(response_text)
        return ClassificacaoSugestao(**dados)
    except Exception as e:
        logging.error(f"Error classifying with AI: {e}")
        return None

# ==================== ROUTES ====================

@api_router.get("/")
async def root():
    return {"message": "ECHO SHOP Financial API v1.0"}

# AUTH ROUTES
@api_router.post("/auth/register", response_model=UserProfile)
@limiter.limit("5/hour")
async def register(request: Request, user_data: UserCreate, current_user: dict = Depends(get_current_user)):
    """
    Register new user - ADMIN ONLY
    Only administrators can create new users
    Rate limited to 5 registrations per hour per IP
    """
    # Check if current user is admin or admin_master
    if current_user.get("perfil") not in ["admin", "admin_master"]:
        log_security_event(
            "UNAUTHORIZED_USER_CREATION_ATTEMPT",
            user_id=current_user.get("id"),
            ip_address=request.client.host if request.client else None,
            details=f"Non-admin tried to create user: {user_data.email}"
        )
        raise HTTPException(status_code=403, detail="Apenas administradores podem cadastrar usuários")
    # Log registration attempt
    log_security_event(
        "USER_REGISTRATION_ATTEMPT",
        user_id=None,
        ip_address=request.client.host if request.client else None,
        details=f"Email: {user_data.email}"
    )
    
    # Check if user exists
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email já cadastrado")
    
    # Hash password
    senha_hash = hash_password(user_data.senha)
    
    # Create user
    user_dict = user_data.model_dump(exclude={"senha"})
    user_obj = UserProfile(**user_dict, senha_hash=senha_hash)
    
    doc = user_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.users.insert_one(doc)
    
    # Log successful registration
    log_security_event(
        "USER_REGISTERED",
        user_id=user_obj.id,
        ip_address=request.client.host if request.client else None,
        details=f"User created: {user_data.email}"
    )
    
    return user_obj

@api_router.get("/users", response_model=List[UserResponse])
async def list_users(current_user: dict = Depends(get_current_user)):
    """
    List all users - ADMIN ONLY
    """
    if current_user.get("perfil") not in ["admin", "admin_master"]:
        raise HTTPException(
            status_code=403,
            detail="Acesso negado. Apenas administradores podem listar usuários."
        )
    
    users = await db.users.find({}, {"_id": 0, "senha_hash": 0}).to_list(100)
    return users


@api_router.put("/users/{user_id}")
async def update_user(
    user_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Update user - ADMIN ONLY
    """
    if current_user.get("perfil") not in ["admin", "admin_master"]:
        raise HTTPException(
            status_code=403,
            detail="Acesso negado. Apenas administradores podem editar usuários."
        )
    
    # Verificar se usuário existe
    existing = await db.users.find_one({"id": user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    
    # Montar update
    update_data = {}
    if data.get("nome"):
        update_data["nome"] = data["nome"]
    if data.get("telefone") is not None:
        update_data["telefone"] = data["telefone"]
    if data.get("perfil"):
        perfil = data["perfil"]
        if perfil not in PERFIS_PERMISSOES:
            raise HTTPException(status_code=400, detail=f"Perfil inválido. Opções: {', '.join(PERFIS_PERMISSOES.keys())}")
        # Apenas admin_master pode criar outro admin_master
        if perfil == "admin_master" and current_user.get("perfil") != "admin_master":
            raise HTTPException(status_code=403, detail="Apenas admin_master pode definir este perfil")
        update_data["perfil"] = perfil
    
    if update_data:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    updated = await db.users.find_one({"id": user_id}, {"_id": 0, "senha_hash": 0})
    return {"message": "Usuário atualizado com sucesso", "user": updated}


@api_router.delete("/users/{user_id}")
async def delete_user(
    user_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Delete user - ADMIN ONLY
    """
    if current_user.get("perfil") not in ["admin", "admin_master"]:
        raise HTTPException(
            status_code=403,
            detail="Acesso negado. Apenas administradores podem excluir usuários."
        )
    
    # Verificar se não está excluindo a si mesmo
    if user_id == current_user.get("id"):
        raise HTTPException(status_code=400, detail="Você não pode excluir a si mesmo")
    
    # Verificar se usuário existe
    existing = await db.users.find_one({"id": user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    
    # Não permitir excluir admin_master se não for admin_master
    if existing.get("perfil") == "admin_master" and current_user.get("perfil") != "admin_master":
        raise HTTPException(status_code=403, detail="Apenas admin_master pode excluir outro admin_master")
    
    await db.users.delete_one({"id": user_id})
    
    return {"message": "Usuário excluído com sucesso"}


# ==================== SETUP / SEED ENDPOINT ====================

@api_router.post("/setup/initialize")
@limiter.limit("5/hour")
async def initialize_system(request: Request):
    """
    Initialize system with default admin user and company.
    Only works if no users exist in database.
    Use this endpoint ONCE when deploying to a new environment.
    """
    try:
        # Check if any users already exist
        existing_users = await db.users.count_documents({})
        if existing_users > 0:
            raise HTTPException(
                status_code=400, 
                detail=f"Sistema já inicializado! Existem {existing_users} usuários cadastrados. Este endpoint só pode ser usado em um banco de dados vazio."
            )
        
        # Create default empresa
        empresa_id = str(uuid.uuid4())
        empresa = {
            "id": empresa_id,
            "razao_social": "ECHO SHOP - Empresa Padrão",
            "nome_fantasia": "ECHO SHOP",
            "cnpj": "00.000.000/0001-00",
            "email": "contato@echoshop.com",
            "telefone": "(00) 0000-0000",
            "endereco": "Endereço Padrão",
            "cidade": "Cidade",
            "estado": "UF",
            "cep": "00000-000",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.empresas.insert_one(empresa)
        
        # Create default admin user
        admin_id = str(uuid.uuid4())
        admin_user = {
            "id": admin_id,
            "nome": "Administrador",
            "email": "admin@echoshop.com",
            "telefone": "(00) 00000-0000",
            "perfil": "admin",
            "empresa_ids": [empresa_id],
            "senha_hash": pwd_context.hash("admin123"),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(admin_user)
        
        # Create default categoria
        categoria_id = str(uuid.uuid4())
        categoria = {
            "id": categoria_id,
            "empresa_id": empresa_id,
            "nome": "Geral",
            "tipo": "despesa",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.categorias_financeiras.insert_one(categoria)
        
        # Create default centro de custo
        centro_id = str(uuid.uuid4())
        centro = {
            "id": centro_id,
            "empresa_id": empresa_id,
            "nome": "Administrativo",
            "descricao": "Centro de custo padrão",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.centros_custo.insert_one(centro)
        
        logging.info(f"Sistema inicializado com sucesso! Empresa: {empresa_id}, Admin: {admin_id}")
        
        return {
            "success": True,
            "message": "✅ Sistema inicializado com sucesso!",
            "details": {
                "empresa_id": empresa_id,
                "empresa_nome": "ECHO SHOP - Empresa Padrão",
                "admin_email": "admin@echoshop.com",
                "admin_senha": "admin123",
                "instrucoes": "Faça login com as credenciais acima e configure sua empresa em Configurações > Empresas"
            }
        }
    except Exception as e:
        logging.error(f"Erro ao inicializar sistema: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao inicializar sistema: {str(e)}")


@api_router.post("/setup/reset-admin-password")
@limiter.limit("3/hour")
async def reset_admin_password(request: Request):
    """
    Reset admin user password to default (admin123).
    Use this if you forgot your admin password.
    Only resets password for user with email admin@echoshop.com
    """
    try:
        # Find admin user
        admin_user = await db.users.find_one({"email": "admin@echoshop.com"}, {"_id": 0})
        
        if not admin_user:
            raise HTTPException(
                status_code=404,
                detail="Usuário admin@echoshop.com não encontrado. Crie um usuário admin primeiro."
            )
        
        # Reset password to default
        new_password_hash = pwd_context.hash("admin123")
        
        await db.users.update_one(
            {"email": "admin@echoshop.com"},
            {"$set": {
                "senha_hash": new_password_hash,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        logging.info(f"Admin password reset for user: {admin_user['id']}")
        
        return {
            "success": True,
            "message": "✅ Senha do admin resetada com sucesso!",
            "details": {
                "email": "admin@echoshop.com",
                "nova_senha": "admin123",
                "instrucoes": "Faça login e mude a senha imediatamente!"
            }
        }
    except Exception as e:
        logging.error(f"Erro ao resetar senha do admin: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao resetar senha: {str(e)}")


@api_router.get("/setup/list-users")
@limiter.limit("10/hour")
async def list_users_for_recovery(request: Request):
    """
    Emergency endpoint to list user emails (WITHOUT passwords).
    Use this to discover which users exist in the database.
    Only shows: email, name, profile - NO sensitive data.
    """
    try:
        users = await db.users.find({}, {"_id": 0, "email": 1, "nome": 1, "perfil": 1}).to_list(100)
        
        if not users:
            return {
                "success": False,
                "message": "Nenhum usuário encontrado no banco de dados",
                "users": []
            }
        
        return {
            "success": True,
            "message": f"Encontrados {len(users)} usuários",
            "users": users,
            "instrucoes": "Use o endpoint /setup/reset-admin-password para resetar a senha do admin@echoshop.com"
        }
        
    except Exception as e:
        logging.error(f"Error listing users: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao listar usuários: {str(e)}")


@api_router.post("/auth/login", response_model=TokenResponse)
@limiter.limit("10/minute")
async def login(request: Request, credentials: UserLogin):
    """
    Login with brute force protection
    Rate limited to 10 attempts per minute per IP
    """
    ip_address = request.client.host if request.client else "unknown"
    
    # Check for brute force attempts
    if check_brute_force(credentials.email):
        log_security_event(
            "BRUTE_FORCE_DETECTED",
            user_id=None,
            ip_address=ip_address,
            details=f"Email: {credentials.email}"
        )
        raise HTTPException(
            status_code=429,
            detail="Muitas tentativas de login. Tente novamente em 15 minutos."
        )
    
    if check_brute_force(ip_address):
        log_security_event(
            "BRUTE_FORCE_DETECTED_IP",
            user_id=None,
            ip_address=ip_address,
            details="Too many failed attempts from this IP"
        )
        raise HTTPException(
            status_code=429,
            detail="Muitas tentativas de login. Tente novamente em 15 minutos."
        )
    
    # Check for SQL injection
    if check_sql_injection(credentials.email):
        log_security_event(
            "SQL_INJECTION_ATTEMPT",
            user_id=None,
            ip_address=ip_address,
            details=f"Suspicious email: {credentials.email}"
        )
        raise HTTPException(status_code=400, detail="Entrada inválida")
    
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user:
        record_failed_login(credentials.email)
        record_failed_login(ip_address)
        log_security_event(
            "LOGIN_FAILED_USER_NOT_FOUND",
            user_id=None,
            ip_address=ip_address,
            details=f"Email: {credentials.email}"
        )
        raise HTTPException(status_code=401, detail="Credenciais inválidas")
    
    if not verify_password(credentials.senha, user["senha_hash"]):
        record_failed_login(credentials.email)
        record_failed_login(ip_address)
        log_security_event(
            "LOGIN_FAILED_WRONG_PASSWORD",
            user_id=user["id"],
            ip_address=ip_address,
            details=f"Email: {credentials.email}"
        )
        raise HTTPException(status_code=401, detail="Credenciais inválidas")
    
    # Clear failed attempts on successful login
    clear_failed_logins(credentials.email)
    clear_failed_logins(ip_address)
    
    # Create token with sessao_id
    empresa_id = user.get("empresa_ids", [""])[0] if user.get("empresa_ids") else ""
    sessao_id = await iniciar_sessao(user["id"], user["email"], empresa_id, ip_address)
    token = create_access_token({
        "sub": user["id"], 
        "email": user["email"],
        "sessao_id": sessao_id
    })
    
    # Log successful login
    log_security_event(
        "LOGIN_SUCCESS",
        user_id=user["id"],
        ip_address=ip_address,
        details=f"Email: {credentials.email}"
    )
    
    # Registrar ação de login
    await registrar_acao(
        user_id=user["id"],
        user_email=user["email"],
        empresa_id=empresa_id,
        acao="login",
        modulo="auth",
        detalhes={"ip": ip_address},
        request=request
    )
    
    # Remove sensitive data
    user_data = {k: v for k, v in user.items() if k != "senha_hash"}
    
    return TokenResponse(access_token=token, user=user_data)

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    user_data = {k: v for k, v in current_user.items() if k != "senha_hash"}
    return user_data

@api_router.post("/auth/logout")
async def logout(request: Request, current_user: dict = Depends(get_current_user)):
    """Logout e finaliza a sessão"""
    try:
        token = request.headers.get("Authorization", "").replace("Bearer ", "")
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        sessao_id = payload.get("sessao_id")
        
        if sessao_id:
            await finalizar_sessao(sessao_id)
        
        # Registrar ação de logout
        empresa_id = current_user.get("empresa_ids", [""])[0] if current_user.get("empresa_ids") else ""
        await registrar_acao(
            user_id=current_user["id"],
            user_email=current_user["email"],
            empresa_id=empresa_id,
            acao="logout",
            modulo="auth",
            request=request
        )
        
        return {"message": "Logout realizado com sucesso"}
    except:
        return {"message": "Logout realizado"}

@api_router.get("/auth/perfis")
async def get_perfis():
    """Retorna todos os perfis disponíveis e suas permissões"""
    return PERFIS_PERMISSOES

@api_router.get("/logs/acoes")
async def get_logs_acoes(
    empresa_id: str,
    limit: int = 100,
    modulo: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Retorna logs de ações da empresa"""
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Verificar permissão
    if not verificar_permissao(current_user, "visualizar"):
        raise HTTPException(status_code=403, detail="Sem permissão para visualizar logs")
    
    filtro = {"empresa_id": empresa_id}
    if modulo:
        filtro["modulo"] = modulo
    
    logs = await db.logs_acoes.find(filtro, {"_id": 0}).sort("timestamp", -1).limit(limit).to_list(limit)
    return logs

@api_router.get("/logs/sessoes")
async def get_logs_sessoes(
    empresa_id: str,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Retorna logs de sessões da empresa"""
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Verificar permissão (apenas admin ou admin_master)
    if current_user.get("perfil") not in ["admin", "admin_master"]:
        raise HTTPException(status_code=403, detail="Apenas administradores podem visualizar sessões")
    
    sessoes = await db.logs_sessoes.find(
        {"empresa_id": empresa_id},
        {"_id": 0}
    ).sort("login_at", -1).limit(limit).to_list(limit)
    
    return sessoes

@api_router.get("/logs/relatorio")
async def get_relatorio_logs(
    empresa_id: str,
    data_inicio: Optional[str] = None,
    data_fim: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Gera relatório de uso do sistema"""
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Verificar permissão (apenas admin ou admin_master)
    if current_user.get("perfil") not in ["admin", "admin_master"]:
        raise HTTPException(status_code=403, detail="Apenas administradores podem gerar relatórios")
    
    filtro = {"empresa_id": empresa_id}
    
    # Filtrar por data se fornecido
    if data_inicio or data_fim:
        filtro["timestamp"] = {}
        if data_inicio:
            filtro["timestamp"]["$gte"] = data_inicio
        if data_fim:
            filtro["timestamp"]["$lte"] = data_fim
    
    # Estatísticas
    total_acoes = await db.logs_acoes.count_documents(filtro)
    
    # Ações por módulo
    pipeline_modulos = [
        {"$match": filtro},
        {"$group": {"_id": "$modulo", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    acoes_por_modulo = await db.logs_acoes.aggregate(pipeline_modulos).to_list(100)
    
    # Usuários mais ativos
    pipeline_usuarios = [
        {"$match": filtro},
        {"$group": {"_id": "$user_email", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10}
    ]
    usuarios_ativos = await db.logs_acoes.aggregate(pipeline_usuarios).to_list(10)
    
    # Tempo médio de sessão
    sessoes = await db.logs_sessoes.find(
        {"empresa_id": empresa_id, "duracao_segundos": {"$exists": True}},
        {"_id": 0, "duracao_segundos": 1}
    ).to_list(1000)
    
    tempo_medio = 0
    if sessoes:
        tempo_medio = sum(s.get("duracao_segundos", 0) for s in sessoes) / len(sessoes)
    
    return {
        "total_acoes": total_acoes,
        "acoes_por_modulo": acoes_por_modulo,
        "usuarios_mais_ativos": usuarios_ativos,
        "tempo_medio_sessao_segundos": int(tempo_medio),
        "total_sessoes": len(sessoes)
    }

# ==================== LICENCIAMENTO ====================

@api_router.post("/licencas")
async def criar_licenca(empresa_id: str, plano: str, current_user: dict = Depends(get_current_user)):
    """Cria licença e assinatura no Asaas"""
    if current_user.get("email") != "faraujoneto2025@gmail.com":
        raise HTTPException(status_code=403, detail="Apenas admin master pode criar licenças")
    
    empresa = await db.empresas.find_one({"id": empresa_id}, {"_id": 0})
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    valor = 99.90 if plano == "basico" else 139.90
    
    # Criar cliente no Asaas
    customer_data = {
        "name": empresa["razao_social"],
        "cpfCnpj": empresa["cnpj"],
        "email": empresa.get("email_cobranca", ""),
        "phone": empresa.get("telefone_cobranca", "")
    }
    asaas_customer = await asaas_request("POST", "customers", customer_data)
    
    # Criar assinatura no Asaas
    subscription_data = {
        "customer": asaas_customer["id"],
        "billingType": "BOLETO",
        "value": valor,
        "cycle": "MONTHLY",
        "nextDueDate": (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d")
    }
    asaas_subscription = await asaas_request("POST", "subscriptions", subscription_data)
    
    # Salvar licença
    licenca = Licenca(
        empresa_id=empresa_id,
        plano=plano,
        valor_mensal=valor,
        asaas_customer_id=asaas_customer["id"],
        asaas_subscription_id=asaas_subscription["id"],
        data_vencimento=datetime.now() + timedelta(days=7)
    )
    
    doc = licenca.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    doc['data_vencimento'] = doc['data_vencimento'].isoformat()
    await db.licencas.insert_one(doc)
    
    # Atualizar empresa
    await db.empresas.update_one(
        {"id": empresa_id},
        {"$set": {"asaas_customer_id": asaas_customer["id"]}}
    )
    
    return licenca

@api_router.get("/licencas/{empresa_id}")
async def get_licenca(empresa_id: str, current_user: dict = Depends(get_current_user)):
    """Retorna licença da empresa"""
    licenca = await db.licencas.find_one({"empresa_id": empresa_id}, {"_id": 0})
    if not licenca:
        return None
    return licenca

@api_router.get("/cobrancas/{empresa_id}")
async def get_cobrancas(empresa_id: str, limit: int = 50, current_user: dict = Depends(get_current_user)):
    """Lista cobranças da empresa"""
    cobrancas = await db.cobrancas.find(
        {"empresa_id": empresa_id},
        {"_id": 0}
    ).sort("data_vencimento", -1).limit(limit).to_list(limit)
    return cobrancas

@api_router.post("/webhooks/asaas")
async def webhook_asaas(request: Request):
    """Recebe notificações do Asaas"""
    payload = await request.json()
    event = payload.get("event")
    
    if event in ["PAYMENT_CONFIRMED", "PAYMENT_RECEIVED"]:
        payment_id = payload.get("payment", {}).get("id")
        
        # Atualizar cobrança
        await db.cobrancas.update_one(
            {"asaas_payment_id": payment_id},
            {"$set": {
                "status": "pago",
                "data_pagamento": datetime.now().isoformat()
            }}
        )
        
        # Reativar licença se estava bloqueada
        cobranca = await db.cobrancas.find_one({"asaas_payment_id": payment_id}, {"_id": 0})
        if cobranca:
            await db.licencas.update_one(
                {"id": cobranca["licenca_id"]},
                {"$set": {
                    "status": "ativa",
                    "dias_atraso": 0,
                    "bloqueada_em": None,
                    "updated_at": datetime.now().isoformat()
                }}
            )
            
            # Desbloquear empresa
            await db.empresas.update_one(
                {"id": cobranca["empresa_id"]},
                {"$set": {"bloqueada": False, "motivo_bloqueio": ""}}
            )
    
    elif event == "PAYMENT_OVERDUE":
        payment_id = payload.get("payment", {}).get("id")
        await db.cobrancas.update_one(
            {"asaas_payment_id": payment_id},
            {"$set": {"status": "atrasado"}}
        )
    
    return {"status": "ok"}

@api_router.post("/licencas/verificar-atrasos")
async def verificar_atrasos():
    """Job para verificar licenças atrasadas e bloquear após 5 dias"""
    hoje = datetime.now()
    
    # Buscar licenças ativas com vencimento passado
    licencas = await db.licencas.find({
        "status": "ativa",
        "data_vencimento": {"$lt": hoje.isoformat()}
    }, {"_id": 0}).to_list(1000)
    
    for licenca in licencas:
        vencimento = datetime.fromisoformat(licenca["data_vencimento"])
        dias_atraso = (hoje - vencimento).days
        
        if dias_atraso >= 5:
            # Bloquear licença e empresa
            await db.licencas.update_one(
                {"id": licenca["id"]},
                {"$set": {
                    "status": "bloqueada",
                    "dias_atraso": dias_atraso,
                    "bloqueada_em": hoje.isoformat(),
                    "motivo_bloqueio": f"Pagamento em atraso há {dias_atraso} dias"
                }}
            )
            
            await db.empresas.update_one(
                {"id": licenca["empresa_id"]},
                {"$set": {
                    "bloqueada": True,
                    "motivo_bloqueio": f"Licença bloqueada por inadimplência ({dias_atraso} dias de atraso)"
                }}
            )
            
            # Enviar notificações
            empresa = await db.empresas.find_one({"id": licenca["empresa_id"]}, {"_id": 0})
            if empresa and empresa.get("email_cobranca"):
                await enviar_email(
                    empresa["email_cobranca"],
                    "⚠️ Acesso Bloqueado - Pagamento Pendente",
                    f"<p>Sua licença foi bloqueada devido ao não pagamento há {dias_atraso} dias.</p>"
                )
    
    return {"licencas_processadas": len(licencas)}

# ==================== GATEWAY DE PAGAMENTO (VENDAS) ====================

@api_router.post("/gateway")
async def configurar_gateway(gateway_data: dict, current_user: dict = Depends(get_current_user)):
    """Configura gateway de pagamento para vendas da empresa"""
    empresa_id = gateway_data.get("empresa_id")
    
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Verificar se já existe configuração
    existing = await db.configuracoes_gateway.find_one(
        {"empresa_id": empresa_id, "gateway": gateway_data.get("gateway")},
        {"_id": 0}
    )
    
    if existing:
        # Atualizar existente
        await db.configuracoes_gateway.update_one(
            {"id": existing["id"]},
            {"$set": {
                "api_key": gateway_data.get("api_key"),
                "sandbox_mode": gateway_data.get("sandbox_mode", False),
                "ativo": gateway_data.get("ativo", True)
            }}
        )
        return {"message": "Gateway atualizado com sucesso", "id": existing["id"]}
    else:
        # Criar novo
        config = ConfiguracaoGateway(**gateway_data)
        doc = config.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.configuracoes_gateway.insert_one(doc)
        
        # Registrar ação
        await registrar_acao(
            user_id=current_user["id"],
            user_email=current_user["email"],
            empresa_id=empresa_id,
            acao="configurar_gateway",
            modulo="vendas",
            detalhes={"gateway": gateway_data.get("gateway")}
        )
        
        return {"message": "Gateway configurado com sucesso", "id": config.id}

@api_router.get("/gateway/{empresa_id}")
async def get_gateway(empresa_id: str, current_user: dict = Depends(get_current_user)):
    """Lista configurações de gateway da empresa"""
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    configs = await db.configuracoes_gateway.find(
        {"empresa_id": empresa_id},
        {"_id": 0, "api_key": 0}  # Não retornar API key por segurança
    ).to_list(10)
    
    return configs

# ==================== SISTEMA DE ASSINATURAS SAAS ====================

def gerar_senha_automatica(email: str) -> str:
    """Gera senha automática baseada no email"""
    import random
    import string
    # Pega as primeiras 4 letras do email antes do @
    base = email.split('@')[0][:4].capitalize()
    # Adiciona números e caracteres especiais
    numeros = ''.join(random.choices(string.digits, k=3))
    especial = random.choice(['@', '#', '!', '$'])
    return f"{base}{numeros}{especial}"

@api_router.get("/assinaturas/planos")
async def listar_planos_saas():
    """Retorna os planos SaaS disponíveis"""
    return PLANOS_SAAS

@api_router.post("/assinaturas")
async def criar_assinatura_saas(
    dados: AssinaturaSaaSCreate,
    current_user: dict = Depends(get_current_user)
):
    """
    Cria nova assinatura SaaS:
    1. Cria empresa cliente no sistema
    2. Cria usuário com senha automática
    3. Gera PIX para primeiro pagamento
    4. Envia email com credenciais
    """
    # Verificar se vendedor tem permissão
    if current_user.get("perfil") not in ["admin", "admin_master", "vendas"]:
        raise HTTPException(status_code=403, detail="Apenas vendedores podem criar assinaturas")
    
    # Verificar se plano é válido
    if dados.plano not in PLANOS_SAAS:
        raise HTTPException(status_code=400, detail=f"Plano inválido. Opções: {', '.join(PLANOS_SAAS.keys())}")
    
    plano_info = PLANOS_SAAS[dados.plano]
    
    # Verificar se já existe assinatura com este CNPJ/CPF
    existing = await db.assinaturas_saas.find_one({"cnpj_cpf": dados.cnpj_cpf, "status": {"$ne": "cancelada"}})
    if existing:
        raise HTTPException(status_code=400, detail="Já existe uma assinatura ativa para este CNPJ/CPF")
    
    # Verificar se email já está em uso
    existing_user = await db.users.find_one({"email": dados.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Este email já está cadastrado no sistema")
    
    # Gerar senha automática
    senha_temp = gerar_senha_automatica(dados.email)
    
    # Criar empresa cliente
    empresa_cliente = {
        "id": str(uuid.uuid4()),
        "razao_social": dados.razao_social,
        "cnpj": dados.cnpj_cpf,
        "email_cobranca": dados.email,
        "telefone_cobranca": dados.telefone,
        "is_blocked": False,
        "is_cliente_saas": True,
        "plano_saas": dados.plano,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.empresas.insert_one(empresa_cliente)
    
    # Criar usuário para o cliente
    hashed_password = pwd_context.hash(senha_temp)
    user_cliente = {
        "id": str(uuid.uuid4()),
        "nome": dados.razao_social,
        "email": dados.email,
        "telefone": dados.telefone,
        "senha_hash": hashed_password,
        "perfil": "admin",  # Admin da própria empresa
        "empresa_ids": [empresa_cliente["id"]],
        "is_cliente_saas": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_cliente)
    
    # Dia de vencimento = dia atual
    dia_vencimento = datetime.now().day
    
    # Buscar gateway do admin (sua empresa)
    gateway_config = await db.configuracoes_gateway.find_one({"ativo": True}, {"_id": 0})
    if not gateway_config:
        raise HTTPException(status_code=400, detail="Gateway de pagamento não configurado. Configure o Asaas primeiro.")
    
    api_key = gateway_config["api_key"]
    base_url = "https://sandbox.asaas.com/api/v3" if gateway_config.get("sandbox_mode") else "https://www.asaas.com/api/v3"
    
    # Mock mode - para testes
    if api_key.startswith('MOCK'):
        # Simular resposta do Asaas para testes
        asaas_customer_id = f"mock_customer_{str(uuid.uuid4())[:8]}"
        pix_payment_id = f"mock_payment_{str(uuid.uuid4())[:8]}"
        pix_qr = {
            "encodedImage": f"data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNk+M9QDwADhgGAWjR9awAAAABJRU5ErkJggg==",
            "payload": f"00020126580014br.gov.bcb.pix013636{str(uuid.uuid4()).replace('-', '')}5204000053039865802BR5925MOCK EMPRESA TESTE LTDA6009SAO PAULO62070503***6304ABCD"
        }
    else:
        # Criar cliente no Asaas
        async with httpx.AsyncClient() as client:
            customer_resp = await client.post(
            f"{base_url}/customers",
            json={
                "name": dados.razao_social,
                "cpfCnpj": dados.cnpj_cpf.replace(".", "").replace("-", "").replace("/", ""),
                "email": dados.email,
                "phone": dados.telefone
            },
            headers={"access_token": api_key}
            )
            customer = customer_resp.json()
            
            if "errors" in customer:
                error_msg = customer.get("errors", [{}])[0].get("description", "Erro ao criar cliente no Asaas")
                raise HTTPException(status_code=400, detail=f"Erro Asaas: {error_msg}")
            
            asaas_customer_id = customer["id"]
            
            # Criar cobrança PIX para primeiro pagamento
            pix_resp = await client.post(
                f"{base_url}/payments",
                json={
                    "customer": asaas_customer_id,
                    "billingType": "PIX",
                    "value": plano_info["valor"],
                    "dueDate": datetime.now().strftime("%Y-%m-%d"),
                    "description": f"Assinatura {plano_info['nome']} - Primeiro Pagamento"
                },
                headers={"access_token": api_key}
            )
            pix_payment = pix_resp.json()
            
            if "errors" in pix_payment:
                error_msg = pix_payment.get("errors", [{}])[0].get("description", "Erro ao gerar PIX")
                raise HTTPException(status_code=400, detail=f"Erro Asaas: {error_msg}")
            
            # Buscar QR Code do PIX
            pix_qr_resp = await client.get(
                f"{base_url}/payments/{pix_payment['id']}/pixQrCode",
                headers={"access_token": api_key}
            )
            pix_qr = pix_qr_resp.json()
            pix_payment_id = pix_payment["id"]
    
    # Criar assinatura
    assinatura = AssinaturaSaaS(
        razao_social=dados.razao_social,
        cnpj_cpf=dados.cnpj_cpf,
        email=dados.email,
        telefone=dados.telefone,
        plano=dados.plano,
        valor_mensal=plano_info["valor"],
        user_id=user_cliente["id"],
        user_email=dados.email,
        user_senha_temp=senha_temp,
        asaas_customer_id=asaas_customer_id,
        dia_vencimento=dia_vencimento,
        pix_qrcode=pix_qr.get("encodedImage"),
        pix_codigo=pix_qr.get("payload"),
        pix_payment_id=pix_payment_id,
        vendedor_id=current_user["id"],
        vendedor_nome=current_user["nome"]
    )
    
    # Salvar assinatura
    doc = assinatura.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.assinaturas_saas.insert_one(doc)
    
    # Enviar email com credenciais
    try:
        html_email = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
            <h1 style="color: #667eea;">🎉 Bem-vindo ao Sistema!</h1>
            <p>Olá <strong>{dados.razao_social}</strong>,</p>
            <p>Sua conta foi criada com sucesso! Aqui estão suas credenciais de acesso:</p>
            
            <div style="background: #f8f9fa; padding: 20px; border-radius: 10px; margin: 20px 0;">
                <h3 style="margin-top: 0;">📋 Dados de Acesso</h3>
                <p><strong>Email:</strong> {dados.email}</p>
                <p><strong>Senha:</strong> {senha_temp}</p>
                <p><strong>CPF/CNPJ:</strong> {dados.cnpj_cpf}</p>
            </div>
            
            <div style="background: #e8f5e9; padding: 20px; border-radius: 10px; margin: 20px 0;">
                <h3 style="margin-top: 0;">📦 Seu Plano</h3>
                <p><strong>Plano:</strong> {plano_info['nome']}</p>
                <p><strong>Valor Mensal:</strong> R$ {plano_info['valor']:.2f}</p>
                <p><strong>Vencimento:</strong> Todo dia {dia_vencimento}</p>
            </div>
            
            <p style="color: #f57c00;"><strong>⚠️ Importante:</strong> Seu acesso será liberado após a confirmação do primeiro pagamento via PIX.</p>
            
            <p>Atenciosamente,<br>Equipe ECHO SHOP</p>
        </div>
        """
        
        GMAIL_USER = os.environ.get('GMAIL_USER')
        GMAIL_APP_PASSWORD = os.environ.get('GMAIL_APP_PASSWORD')
        
        if GMAIL_USER and GMAIL_APP_PASSWORD:
            msg = MIMEMultipart('alternative')
            msg['Subject'] = f'🎉 Bem-vindo! Seus dados de acesso - {plano_info["nome"]}'
            msg['From'] = f'Sistema ECHO SHOP <{GMAIL_USER}>'
            msg['To'] = dados.email
            msg.attach(MIMEText(html_email, 'html'))
            
            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
                server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
                server.send_message(msg)
    except Exception as e:
        logging.error(f"Erro ao enviar email: {e}")
    
    return {
        "message": "Assinatura criada com sucesso!",
        "assinatura_id": assinatura.id,
        "empresa_id": empresa_cliente["id"],
        "user_email": dados.email,
        "user_senha": senha_temp,
        "plano": plano_info["nome"],
        "valor": plano_info["valor"],
        "pix_qrcode": pix_qr.get("encodedImage"),
        "pix_codigo": pix_qr.get("payload"),
        "pix_payment_id": pix_payment_id
    }

@api_router.get("/assinaturas")
async def listar_assinaturas(current_user: dict = Depends(get_current_user)):
    """Lista todas as assinaturas SaaS"""
    if current_user.get("perfil") not in ["admin", "admin_master", "vendas"]:
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    assinaturas = await db.assinaturas_saas.find({}, {"_id": 0}).to_list(500)
    return assinaturas

@api_router.get("/assinaturas/{assinatura_id}")
async def get_assinatura(assinatura_id: str, current_user: dict = Depends(get_current_user)):
    """Retorna detalhes de uma assinatura"""
    assinatura = await db.assinaturas_saas.find_one({"id": assinatura_id}, {"_id": 0})
    if not assinatura:
        raise HTTPException(status_code=404, detail="Assinatura não encontrada")
    return assinatura

@api_router.post("/assinaturas/{assinatura_id}/verificar-pagamento")
async def verificar_pagamento_assinatura(assinatura_id: str, current_user: dict = Depends(get_current_user)):
    """Verifica status do pagamento PIX inicial"""
    assinatura = await db.assinaturas_saas.find_one({"id": assinatura_id}, {"_id": 0})
    if not assinatura:
        raise HTTPException(status_code=404, detail="Assinatura não encontrada")
    
    if assinatura.get("primeiro_pagamento_status") == "pago":
        return {"status": "pago", "message": "Pagamento já confirmado"}
    
    # Buscar gateway
    gateway_config = await db.configuracoes_gateway.find_one({"ativo": True}, {"_id": 0})
    if not gateway_config:
        raise HTTPException(status_code=400, detail="Gateway não configurado")
    
    api_key = gateway_config["api_key"]
    base_url = "https://sandbox.asaas.com/api/v3" if gateway_config.get("sandbox_mode") else "https://www.asaas.com/api/v3"
    
    # Verificar status no Asaas
    async with httpx.AsyncClient() as client:
        resp = await client.get(
            f"{base_url}/payments/{assinatura['pix_payment_id']}",
            headers={"access_token": api_key}
        )
        payment = resp.json()
    
    if payment.get("status") in ["RECEIVED", "CONFIRMED"]:
        # Pagamento confirmado!
        await db.assinaturas_saas.update_one(
            {"id": assinatura_id},
            {"$set": {
                "primeiro_pagamento_status": "pago",
                "status": "ativa",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        # Desbloquear empresa
        await db.empresas.update_one(
            {"cnpj": assinatura["cnpj_cpf"]},
            {"$set": {"is_blocked": False}}
        )
        
        # Lançar no financeiro
        await lancar_receita_mensalidade(assinatura)
        
        # Criar assinatura recorrente no Asaas (boletos futuros)
        await criar_assinatura_recorrente(assinatura, gateway_config)
        
        return {"status": "pago", "message": "Pagamento confirmado! Acesso liberado."}
    
    return {"status": payment.get("status", "PENDING"), "message": "Aguardando pagamento"}

async def lancar_receita_mensalidade(assinatura: dict):
    """Lança a receita da mensalidade no financeiro"""
    # Buscar ou criar categoria
    categoria = await db.categorias.find_one({"nome": "Mensalidade Sistema"})
    if not categoria:
        categoria = {
            "id": str(uuid.uuid4()),
            "nome": "Mensalidade Sistema",
            "tipo": "receita",
            "cor": "#4caf50"
        }
        await db.categorias.insert_one(categoria)
    
    # Buscar ou criar centro de custo
    centro_custo = await db.centros_custo.find_one({"nome": "Receitas SaaS"})
    if not centro_custo:
        centro_custo = {
            "id": str(uuid.uuid4()),
            "nome": "Receitas SaaS",
            "descricao": "Receitas de assinaturas do sistema"
        }
        await db.centros_custo.insert_one(centro_custo)
    
    # Buscar empresa admin (sua empresa)
    empresa_admin = await db.empresas.find_one({"is_cliente_saas": {"$ne": True}})
    
    # Criar transação
    transacao = {
        "id": str(uuid.uuid4()),
        "empresa_id": empresa_admin["id"] if empresa_admin else assinatura.get("vendedor_id", ""),
        "tipo": "receita",
        "categoria_id": categoria["id"],
        "categoria_nome": categoria["nome"],
        "centro_custo_id": centro_custo["id"],
        "centro_custo_nome": centro_custo["nome"],
        "descricao": f"Mensalidade SaaS - {assinatura['razao_social']} - Plano {assinatura['plano'].capitalize()}",
        "valor": assinatura["valor_mensal"],
        "data": datetime.now(timezone.utc).isoformat(),
        "status": "pago",
        "observacoes": f"Assinatura ID: {assinatura['id']}",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.transacoes.insert_one(transacao)

async def criar_assinatura_recorrente(assinatura: dict, gateway_config: dict):
    """Cria assinatura recorrente no Asaas para boletos futuros"""
    api_key = gateway_config["api_key"]
    base_url = "https://sandbox.asaas.com/api/v3" if gateway_config.get("sandbox_mode") else "https://www.asaas.com/api/v3"
    
    # Calcular próximo vencimento
    hoje = datetime.now()
    proximo_mes = hoje.month + 1 if hoje.month < 12 else 1
    proximo_ano = hoje.year if hoje.month < 12 else hoje.year + 1
    dia = min(assinatura["dia_vencimento"], 28)  # Evitar problemas com fevereiro
    proximo_vencimento = f"{proximo_ano}-{proximo_mes:02d}-{dia:02d}"
    
    async with httpx.AsyncClient() as client:
        resp = await client.post(
            f"{base_url}/subscriptions",
            json={
                "customer": assinatura["asaas_customer_id"],
                "billingType": "BOLETO",
                "value": assinatura["valor_mensal"],
                "cycle": "MONTHLY",
                "nextDueDate": proximo_vencimento,
                "description": f"Mensalidade Plano {assinatura['plano'].capitalize()}"
            },
            headers={"access_token": api_key}
        )
        subscription = resp.json()
        
        if "id" in subscription:
            await db.assinaturas_saas.update_one(
                {"id": assinatura["id"]},
                {"$set": {"asaas_subscription_id": subscription["id"]}}
            )

@api_router.post("/assinaturas/verificar-inadimplentes")
async def verificar_inadimplentes(current_user: dict = Depends(get_current_user)):
    """Verifica assinaturas inadimplentes e envia cobranças/bloqueia"""
    if current_user.get("perfil") not in ["admin", "admin_master"]:
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    gateway_config = await db.configuracoes_gateway.find_one({"ativo": True}, {"_id": 0})
    if not gateway_config:
        return {"message": "Gateway não configurado"}
    
    api_key = gateway_config["api_key"]
    base_url = "https://sandbox.asaas.com/api/v3" if gateway_config.get("sandbox_mode") else "https://www.asaas.com/api/v3"
    
    # Buscar assinaturas ativas
    assinaturas = await db.assinaturas_saas.find({"status": "ativa"}, {"_id": 0}).to_list(500)
    
    resultados = []
    GMAIL_USER = os.environ.get('GMAIL_USER')
    GMAIL_APP_PASSWORD = os.environ.get('GMAIL_APP_PASSWORD')
    
    for assinatura in assinaturas:
        if not assinatura.get("asaas_subscription_id"):
            continue
        
        # Verificar cobranças pendentes
        async with httpx.AsyncClient() as client:
            resp = await client.get(
                f"{base_url}/payments",
                params={
                    "subscription": assinatura["asaas_subscription_id"],
                    "status": "OVERDUE"
                },
                headers={"access_token": api_key}
            )
            overdue = resp.json()
        
        if overdue.get("data"):
            cobranca = overdue["data"][0]
            vencimento = datetime.strptime(cobranca["dueDate"], "%Y-%m-%d")
            dias_atraso = (datetime.now() - vencimento).days
            
            # Atualizar dias de atraso
            await db.assinaturas_saas.update_one(
                {"id": assinatura["id"]},
                {"$set": {"dias_atraso": dias_atraso}}
            )
            
            # Lógica de cobrança/bloqueio
            if dias_atraso >= 5:
                # BLOQUEAR
                await db.assinaturas_saas.update_one(
                    {"id": assinatura["id"]},
                    {"$set": {
                        "status": "suspensa",
                        "bloqueada_em": datetime.now(timezone.utc).isoformat()
                    }}
                )
                await db.empresas.update_one(
                    {"cnpj": assinatura["cnpj_cpf"]},
                    {"$set": {"is_blocked": True, "block_reason": "Inadimplência"}}
                )
                
                # Enviar email de bloqueio
                if GMAIL_USER and GMAIL_APP_PASSWORD:
                    try:
                        html = f"""
                        <div style="font-family: Arial; max-width: 600px; margin: 0 auto;">
                            <h2 style="color: #dc3545;">⚠️ Acesso Suspenso</h2>
                            <p>Prezado(a) {assinatura['razao_social']},</p>
                            <p>Devido à inadimplência de <strong>{dias_atraso} dias</strong>, seu acesso ao sistema foi suspenso.</p>
                            <p>Para reativar, efetue o pagamento pendente.</p>
                            <p>Atenciosamente,<br>Equipe ECHO SHOP</p>
                        </div>
                        """
                        msg = MIMEMultipart('alternative')
                        msg['Subject'] = '⚠️ Acesso Suspenso - Inadimplência'
                        msg['From'] = f'Sistema ECHO SHOP <{GMAIL_USER}>'
                        msg['To'] = assinatura['email']
                        msg.attach(MIMEText(html, 'html'))
                        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
                            server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
                            server.send_message(msg)
                    except Exception as e:
                        logging.error(f"Erro envio email bloqueio: {e}")
                
                resultados.append({"assinatura": assinatura["razao_social"], "acao": "bloqueada"})
            
            elif dias_atraso >= 3:
                # Aviso de suspensão
                if GMAIL_USER and GMAIL_APP_PASSWORD:
                    try:
                        html = f"""
                        <div style="font-family: Arial; max-width: 600px; margin: 0 auto;">
                            <h2 style="color: #f57c00;">⚠️ Aviso de Suspensão</h2>
                            <p>Prezado(a) {assinatura['razao_social']},</p>
                            <p>Sua mensalidade está em atraso há <strong>{dias_atraso} dias</strong>.</p>
                            <p><strong>Seu acesso será suspenso em {5 - dias_atraso} dias caso o pagamento não seja efetuado.</strong></p>
                            <p>Atenciosamente,<br>Equipe ECHO SHOP</p>
                        </div>
                        """
                        msg = MIMEMultipart('alternative')
                        msg['Subject'] = '⚠️ Aviso: Suspensão em breve'
                        msg['From'] = f'Sistema ECHO SHOP <{GMAIL_USER}>'
                        msg['To'] = assinatura['email']
                        msg.attach(MIMEText(html, 'html'))
                        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
                            server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
                            server.send_message(msg)
                    except Exception as e:
                        logging.error(f"Erro envio email aviso: {e}")
                
                resultados.append({"assinatura": assinatura["razao_social"], "acao": "aviso_suspensao"})
            
            else:
                # Email de cobrança
                if GMAIL_USER and GMAIL_APP_PASSWORD:
                    try:
                        html = f"""
                        <div style="font-family: Arial; max-width: 600px; margin: 0 auto;">
                            <h2 style="color: #1976d2;">📋 Lembrete de Pagamento</h2>
                            <p>Prezado(a) {assinatura['razao_social']},</p>
                            <p>Identificamos que sua mensalidade está pendente.</p>
                            <p>Valor: <strong>R$ {assinatura['valor_mensal']:.2f}</strong></p>
                            <p>Efetue o pagamento para evitar a suspensão do serviço.</p>
                            <p>Atenciosamente,<br>Equipe ECHO SHOP</p>
                        </div>
                        """
                        msg = MIMEMultipart('alternative')
                        msg['Subject'] = '📋 Lembrete: Mensalidade Pendente'
                        msg['From'] = f'Sistema ECHO SHOP <{GMAIL_USER}>'
                        msg['To'] = assinatura['email']
                        msg.attach(MIMEText(html, 'html'))
                        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
                            server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
                            server.send_message(msg)
                    except Exception as e:
                        logging.error(f"Erro envio email cobranca: {e}")
                
                resultados.append({"assinatura": assinatura["razao_social"], "acao": "email_cobranca"})
    
    return {"processadas": len(resultados), "resultados": resultados}

@api_router.post("/webhook/asaas")
async def webhook_asaas(request: Request):
    """Webhook para receber notificações do Asaas"""
    try:
        data = await request.json()
        event = data.get("event")
        payment = data.get("payment", {})
        
        if event == "PAYMENT_RECEIVED":
            # Pagamento recebido - verificar se é de assinatura SaaS
            asaas_customer_id = payment.get("customer")
            
            assinatura = await db.assinaturas_saas.find_one(
                {"asaas_customer_id": asaas_customer_id},
                {"_id": 0}
            )
            
            if assinatura:
                # É pagamento de assinatura SaaS
                if assinatura.get("primeiro_pagamento_status") == "pendente":
                    # Primeiro pagamento
                    await db.assinaturas_saas.update_one(
                        {"id": assinatura["id"]},
                        {"$set": {
                            "primeiro_pagamento_status": "pago",
                            "status": "ativa",
                            "dias_atraso": 0,
                            "updated_at": datetime.now(timezone.utc).isoformat()
                        }}
                    )
                    await db.empresas.update_one(
                        {"cnpj": assinatura["cnpj_cpf"]},
                        {"$set": {"is_blocked": False}}
                    )
                else:
                    # Mensalidade recorrente
                    await db.assinaturas_saas.update_one(
                        {"id": assinatura["id"]},
                        {"$set": {
                            "status": "ativa",
                            "dias_atraso": 0,
                            "updated_at": datetime.now(timezone.utc).isoformat()
                        }}
                    )
                    await db.empresas.update_one(
                        {"cnpj": assinatura["cnpj_cpf"]},
                        {"$set": {"is_blocked": False}}
                    )
                
                # Lançar receita
                await lancar_receita_mensalidade(assinatura)
        
        return {"status": "ok"}
    except Exception as e:
        logging.error(f"Erro webhook Asaas: {e}")
        return {"status": "error", "message": str(e)}

@api_router.post("/vendas")
async def criar_venda(venda_data: dict, current_user: dict = Depends(get_current_user)):
    """Cria venda e gera cobrança no gateway do cliente"""
    empresa_id = venda_data.get("empresa_id")
    
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Verificar permissão
    if not verificar_permissao(current_user, "vendas"):
        raise HTTPException(status_code=403, detail="Sem permissão para criar vendas")
    
    # Buscar configuração do gateway
    gateway_config = await db.configuracoes_gateway.find_one(
        {"empresa_id": empresa_id, "ativo": True},
        {"_id": 0}
    )
    
    if not gateway_config:
        raise HTTPException(
            status_code=400,
            detail="Configure um gateway de pagamento antes de criar vendas"
        )
    
    # Criar venda
    venda = Venda(**venda_data)
    
    # Gerar cobrança no gateway do cliente
    if gateway_config["gateway"] == "asaas":
        # Usar Asaas do cliente
        api_key = gateway_config["api_key"]
        base_url = "https://sandbox.asaas.com/api/v3" if gateway_config["sandbox_mode"] else "https://www.asaas.com/api/v3"
        
        # Mock mode
        if api_key.startswith("MOCK"):
            payment = {
                "id": f"mock_payment_{venda.id[:8]}",
                "bankSlipUrl": f"https://mock-boleto.com/{venda.id}",
                "identificationField": "00000.00000 00000.000000 00000.000000 0 00000000000000"
            }
        else:
            # Criar cliente no Asaas do cliente
            customer_data = {
                "name": venda.cliente_nome,
                "cpfCnpj": venda.cliente_cpf_cnpj,
                "email": venda.cliente_email or "",
                "phone": venda.cliente_telefone or ""
            }
            
            async with httpx.AsyncClient() as client:
                customer_resp = await client.post(
                    f"{base_url}/customers",
                    json=customer_data,
                    headers={"access_token": api_key}
                )
                customer = customer_resp.json()
                
                # Verificar erro na criação do cliente
                if "errors" in customer:
                    error_msg = customer.get("errors", [{}])[0].get("description", "Erro ao criar cliente no Asaas")
                    raise HTTPException(status_code=400, detail=f"Erro Asaas: {error_msg}")
                
                if "id" not in customer:
                    raise HTTPException(status_code=400, detail=f"Resposta inesperada do Asaas: {customer}")
                
                # Criar cobrança
                payment_data = {
                    "customer": customer["id"],
                    "billingType": venda.metodo_pagamento.upper(),
                    "value": venda.valor_total,
                    "dueDate": venda.data_vencimento,
                    "description": venda.descricao
                }
                
                payment_resp = await client.post(
                    f"{base_url}/payments",
                    json=payment_data,
                    headers={"access_token": api_key}
                )
                payment = payment_resp.json()
                
                # Verificar erro na criação da cobrança
                if "errors" in payment:
                    error_msg = payment.get("errors", [{}])[0].get("description", "Erro ao criar cobrança no Asaas")
                    raise HTTPException(status_code=400, detail=f"Erro Asaas: {error_msg}")
        
        # Atualizar venda com dados do pagamento
        venda.gateway_payment_id = payment.get("id")
        venda.boleto_url = payment.get("bankSlipUrl")
        venda.boleto_codigo = payment.get("identificationField")
    
    # Salvar venda
    doc = venda.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    # data_vencimento já é string, não precisa converter
    
    await db.vendas.insert_one(doc)
    
    # Registrar ação
    await registrar_acao(
        user_id=current_user["id"],
        user_email=current_user["email"],
        empresa_id=empresa_id,
        acao="criar_venda",
        modulo="vendas",
        detalhes={"cliente": venda.cliente_nome, "valor": venda.valor_total}
    )
    
    return venda

@api_router.get("/vendas/{empresa_id}")
async def listar_vendas(
    empresa_id: str,
    status: Optional[str] = None,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Lista vendas da empresa"""
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    filtro = {"empresa_id": empresa_id}
    if status:
        filtro["status"] = status
    
    vendas = await db.vendas.find(filtro, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    return vendas

@api_router.get("/vendas/detalhes/{venda_id}")
async def get_venda(venda_id: str, current_user: dict = Depends(get_current_user)):
    """Detalhes de uma venda"""
    venda = await db.vendas.find_one({"id": venda_id}, {"_id": 0})
    
    if not venda:
        raise HTTPException(status_code=404, detail="Venda não encontrada")
    
    if venda["empresa_id"] not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    return venda

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    """
    Delete a user - ADMIN ONLY
    """
    if current_user.get("perfil") not in ["admin", "admin_master"]:
        raise HTTPException(status_code=403, detail="Apenas administradores podem deletar usuários")
    
    # Prevent self-deletion
    if user_id == current_user["id"]:
        raise HTTPException(status_code=400, detail="Você não pode deletar sua própria conta")
    
    # Check if user exists
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    
    # Delete user
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    
    # Log the deletion
    log_security_event(
        "USER_DELETED",
        user_id=current_user["id"],
        ip_address="system",
        details=f"Deleted user: {user['email']}"
    )
    
    return {"message": f"Usuário {user['nome']} deletado com sucesso"}

# EMPRESA ROUTES
@api_router.post("/empresas", response_model=Empresa)
async def create_empresa(
    empresa_data: EmpresaCreate,
    senha_admin: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Criar empresa - Apenas o administrador master pode criar
    Requer senha especial: faraujoneto2005@gmail.com com senha Rebeca@19
    """
    # Verificar se é o usuário autorizado com senha correta
    ADMIN_EMAIL = "faraujoneto2005@gmail.com"
    ADMIN_SENHA = "Rebeca@19"  # TODO: Mover para variável de ambiente
    
    if current_user.get("email") != ADMIN_EMAIL or senha_admin != ADMIN_SENHA:
        raise HTTPException(
            status_code=403, 
            detail="Apenas o administrador master pode cadastrar empresas"
        )
    
    empresa_obj = Empresa(**empresa_data.model_dump())
    doc = empresa_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.empresas.insert_one(doc)
    
    # Add empresa to user's list
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$addToSet": {"empresa_ids": empresa_obj.id}}
    )
    
    # Registrar ação
    await registrar_acao(
        user_id=current_user["id"],
        user_email=current_user["email"],
        empresa_id=empresa_obj.id,
        acao="criar_empresa",
        modulo="empresas",
        detalhes={"razao_social": empresa_obj.razao_social, "cnpj": empresa_obj.cnpj}
    )
    
    return empresa_obj

@api_router.get("/empresas", response_model=List[Empresa])
async def get_empresas(current_user: dict = Depends(get_current_user)):
    empresas = await db.empresas.find(
        {"id": {"$in": current_user.get("empresa_ids", [])}},
        {"_id": 0}
    ).to_list(100)
    return empresas

@api_router.put("/empresas/{empresa_id}")
async def update_empresa(empresa_id: str, empresa_data: dict, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Remove campos que não devem ser atualizados
    update_data = {k: v for k, v in empresa_data.items() if k not in ['id', 'created_at']}
    
    result = await db.empresas.update_one(
        {"id": empresa_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    return {"message": "Empresa atualizada com sucesso"}

@api_router.patch("/empresas/{empresa_id}/status")
async def update_empresa_status(empresa_id: str, status_data: dict, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    result = await db.empresas.update_one(
        {"id": empresa_id},
        {"$set": {
            "bloqueada": status_data.get("bloqueada", False),
            "motivo_bloqueio": status_data.get("motivo_bloqueio", "")
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    return {"message": "Status atualizado com sucesso"}

@api_router.delete("/empresas/{empresa_id}")
async def delete_empresa(empresa_id: str, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Verificar se é admin (pode adicionar verificação de perfil aqui)
    
    # Deletar empresa
    result = await db.empresas.delete_one({"id": empresa_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    # Remover empresa_id do usuário
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$pull": {"empresa_ids": empresa_id}}
    )
    
    return {"message": "Empresa excluída com sucesso"}

# CATEGORIA ROUTES
@api_router.post("/empresas/{empresa_id}/categorias", response_model=Categoria)
async def create_categoria(empresa_id: str, categoria_data: CategoriaCreate, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    categoria_obj = Categoria(**categoria_data.model_dump(), empresa_id=empresa_id)
    doc = categoria_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.categorias.insert_one(doc)
    return categoria_obj

@api_router.get("/empresas/{empresa_id}/categorias", response_model=List[Categoria])
async def get_categorias(empresa_id: str, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    categorias = await db.categorias.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(100)
    return categorias

@api_router.delete("/categorias/{categoria_id}")
async def delete_categoria(categoria_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.categorias.delete_one({"id": categoria_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Categoria não encontrada")
    return {"message": "Categoria deletada"}

# CENTRO DE CUSTO ROUTES
@api_router.post("/empresas/{empresa_id}/centros-custo", response_model=CentroCusto)
async def create_centro_custo(empresa_id: str, cc_data: CentroCustoCreate, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    cc_obj = CentroCusto(**cc_data.model_dump(), empresa_id=empresa_id)
    doc = cc_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.centros_custo.insert_one(doc)
    return cc_obj

@api_router.get("/empresas/{empresa_id}/centros-custo", response_model=List[CentroCusto])
async def get_centros_custo(empresa_id: str, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    centros = await db.centros_custo.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(100)
    return centros

@api_router.delete("/centros-custo/{cc_id}")
async def delete_centro_custo(cc_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.centros_custo.delete_one({"id": cc_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Centro de custo não encontrado")
    return {"message": "Centro de custo deletado"}

# CONTAS BANCÁRIAS ROUTES
@api_router.post("/empresas/{empresa_id}/contas", response_model=ContaBancaria)
async def create_conta(empresa_id: str, conta_data: ContaBancariaCreate, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    conta_obj = ContaBancaria(**conta_data.model_dump(), empresa_id=empresa_id, saldo_atual=conta_data.saldo_inicial)
    doc = conta_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.contas_bancarias.insert_one(doc)
    return conta_obj

@api_router.get("/empresas/{empresa_id}/contas", response_model=List[ContaBancaria])
async def get_contas(empresa_id: str, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    contas = await db.contas_bancarias.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(100)
    return contas

@api_router.put("/contas/{conta_id}", response_model=ContaBancaria)
async def update_conta(conta_id: str, conta_data: ContaBancariaCreate, current_user: dict = Depends(get_current_user)):
    conta = await db.contas_bancarias.find_one({"id": conta_id}, {"_id": 0})
    if not conta:
        raise HTTPException(status_code=404, detail="Conta não encontrada")
    
    update_data = conta_data.model_dump()
    update_data['saldo_atual'] = conta['saldo_atual']  # Preservar saldo atual
    await db.contas_bancarias.update_one({"id": conta_id}, {"$set": update_data})
    
    updated_conta = await db.contas_bancarias.find_one({"id": conta_id}, {"_id": 0})
    return updated_conta

@api_router.delete("/contas/{conta_id}")
async def delete_conta(conta_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.contas_bancarias.delete_one({"id": conta_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Conta não encontrada")
    return {"message": "Conta deletada"}

# INVESTIMENTOS ROUTES
@api_router.post("/empresas/{empresa_id}/investimentos", response_model=Investimento)
async def create_investimento(empresa_id: str, inv_data: InvestimentoCreate, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    inv_obj = Investimento(**inv_data.model_dump(), empresa_id=empresa_id)
    doc = inv_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.investimentos.insert_one(doc)
    return inv_obj

@api_router.get("/empresas/{empresa_id}/investimentos", response_model=List[Investimento])
async def get_investimentos(empresa_id: str, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    investimentos = await db.investimentos.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(100)
    return investimentos

@api_router.put("/investimentos/{inv_id}", response_model=Investimento)
async def update_investimento(inv_id: str, inv_data: InvestimentoCreate, current_user: dict = Depends(get_current_user)):
    inv = await db.investimentos.find_one({"id": inv_id}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Investimento não encontrado")
    
    update_data = inv_data.model_dump()
    await db.investimentos.update_one({"id": inv_id}, {"$set": update_data})
    
    updated_inv = await db.investimentos.find_one({"id": inv_id}, {"_id": 0})
    return updated_inv

@api_router.delete("/investimentos/{inv_id}")
async def delete_investimento(inv_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.investimentos.delete_one({"id": inv_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Investimento não encontrado")
    return {"message": "Investimento deletado"}

# CARTÕES DE CRÉDITO ROUTES
@api_router.post("/empresas/{empresa_id}/cartoes", response_model=CartaoCredito)
async def create_cartao(empresa_id: str, cartao_data: CartaoCreditoCreate, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    cartao_obj = CartaoCredito(**cartao_data.model_dump(), empresa_id=empresa_id, limite_disponivel=cartao_data.limite_total)
    doc = cartao_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.cartoes_credito.insert_one(doc)
    return cartao_obj

@api_router.get("/empresas/{empresa_id}/cartoes", response_model=List[CartaoCredito])
async def get_cartoes(empresa_id: str, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    cartoes = await db.cartoes_credito.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(100)
    return cartoes

@api_router.put("/cartoes/{cartao_id}", response_model=CartaoCredito)
async def update_cartao(cartao_id: str, cartao_data: CartaoCreditoCreate, current_user: dict = Depends(get_current_user)):
    cartao = await db.cartoes_credito.find_one({"id": cartao_id}, {"_id": 0})
    if not cartao:
        raise HTTPException(status_code=404, detail="Cartão não encontrado")
    
    update_data = cartao_data.model_dump()
    update_data['limite_disponivel'] = cartao['limite_disponivel']  # Preservar limite disponível
    update_data['fatura_atual'] = cartao['fatura_atual']  # Preservar fatura
    await db.cartoes_credito.update_one({"id": cartao_id}, {"$set": update_data})
    
    updated_cartao = await db.cartoes_credito.find_one({"id": cartao_id}, {"_id": 0})
    return updated_cartao

@api_router.delete("/cartoes/{cartao_id}")
async def delete_cartao(cartao_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.cartoes_credito.delete_one({"id": cartao_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Cartão não encontrado")
    return {"message": "Cartão deletado"}

# ==================== ESTOQUE ROUTES ====================

# CLIENTES ROUTES
@api_router.post("/empresas/{empresa_id}/clientes", response_model=Cliente)
async def create_cliente(empresa_id: str, cliente_data: ClienteCreate, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    cliente_obj = Cliente(**cliente_data.model_dump(), empresa_id=empresa_id)
    doc = cliente_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.clientes.insert_one(doc)
    return cliente_obj

@api_router.get("/empresas/{empresa_id}/clientes", response_model=List[Cliente])
async def get_clientes(empresa_id: str, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    clientes = await db.clientes.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(1000)
    return clientes

@api_router.get("/clientes/{cliente_id}", response_model=Cliente)
async def get_cliente(cliente_id: str, current_user: dict = Depends(get_current_user)):
    cliente = await db.clientes.find_one({"id": cliente_id}, {"_id": 0})
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    return cliente

@api_router.put("/clientes/{cliente_id}", response_model=Cliente)
async def update_cliente(cliente_id: str, cliente_data: ClienteCreate, current_user: dict = Depends(get_current_user)):
    cliente = await db.clientes.find_one({"id": cliente_id}, {"_id": 0})
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    
    update_data = cliente_data.model_dump()
    await db.clientes.update_one({"id": cliente_id}, {"$set": update_data})
    
    updated_cliente = await db.clientes.find_one({"id": cliente_id}, {"_id": 0})
    return updated_cliente

@api_router.delete("/clientes/{cliente_id}")
async def delete_cliente(cliente_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.clientes.delete_one({"id": cliente_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    return {"message": "Cliente deletado"}

# FORNECEDORES ROUTES
@api_router.post("/empresas/{empresa_id}/fornecedores", response_model=Fornecedor)
async def create_fornecedor(empresa_id: str, fornecedor_data: FornecedorCreate, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    fornecedor_obj = Fornecedor(**fornecedor_data.model_dump(), empresa_id=empresa_id)
    doc = fornecedor_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.fornecedores.insert_one(doc)
    return fornecedor_obj

@api_router.get("/empresas/{empresa_id}/fornecedores", response_model=List[Fornecedor])
async def get_fornecedores(empresa_id: str, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    fornecedores = await db.fornecedores.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(1000)
    return fornecedores

@api_router.put("/fornecedores/{fornecedor_id}", response_model=Fornecedor)
async def update_fornecedor(fornecedor_id: str, fornecedor_data: FornecedorCreate, current_user: dict = Depends(get_current_user)):
    fornecedor = await db.fornecedores.find_one({"id": fornecedor_id}, {"_id": 0})
    if not fornecedor:
        raise HTTPException(status_code=404, detail="Fornecedor não encontrado")
    
    update_data = fornecedor_data.model_dump()
    await db.fornecedores.update_one({"id": fornecedor_id}, {"$set": update_data})
    
    updated_fornecedor = await db.fornecedores.find_one({"id": fornecedor_id}, {"_id": 0})
    return updated_fornecedor

@api_router.delete("/fornecedores/{fornecedor_id}")
async def delete_fornecedor(fornecedor_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.fornecedores.delete_one({"id": fornecedor_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Fornecedor não encontrado")
    return {"message": "Fornecedor deletado"}

# LOCAIS/DEPÓSITOS ROUTES
@api_router.post("/empresas/{empresa_id}/locais", response_model=LocalDeposito)
async def create_local(empresa_id: str, local_data: LocalDepositoCreate, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    local_obj = LocalDeposito(**local_data.model_dump(), empresa_id=empresa_id)
    doc = local_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.locais_deposito.insert_one(doc)
    return local_obj

@api_router.get("/empresas/{empresa_id}/locais", response_model=List[LocalDeposito])
async def get_locais(empresa_id: str, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    locais = await db.locais_deposito.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(1000)
    return locais

@api_router.put("/locais/{local_id}", response_model=LocalDeposito)
async def update_local(local_id: str, local_data: LocalDepositoCreate, current_user: dict = Depends(get_current_user)):
    local = await db.locais_deposito.find_one({"id": local_id}, {"_id": 0})
    if not local:
        raise HTTPException(status_code=404, detail="Local não encontrado")
    
    update_data = local_data.model_dump()
    await db.locais_deposito.update_one({"id": local_id}, {"$set": update_data})
    
    updated_local = await db.locais_deposito.find_one({"id": local_id}, {"_id": 0})
    return updated_local

@api_router.delete("/locais/{local_id}")
async def delete_local(local_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.locais_deposito.delete_one({"id": local_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Local não encontrado")
    return {"message": "Local deletado"}

# CATEGORIAS DE EQUIPAMENTOS ROUTES
@api_router.post("/empresas/{empresa_id}/categorias-equipamentos", response_model=CategoriaEquipamento)
async def create_categoria_equipamento(empresa_id: str, categoria_data: CategoriaEquipamentoCreate, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    categoria_obj = CategoriaEquipamento(**categoria_data.model_dump(), empresa_id=empresa_id)
    doc = categoria_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.categorias_equipamentos.insert_one(doc)
    return categoria_obj

@api_router.get("/empresas/{empresa_id}/categorias-equipamentos", response_model=List[CategoriaEquipamento])
async def get_categorias_equipamentos(empresa_id: str, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    categorias = await db.categorias_equipamentos.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(1000)
    return categorias

@api_router.put("/categorias-equipamentos/{categoria_id}", response_model=CategoriaEquipamento)
async def update_categoria_equipamento(categoria_id: str, categoria_data: CategoriaEquipamentoCreate, current_user: dict = Depends(get_current_user)):
    categoria = await db.categorias_equipamentos.find_one({"id": categoria_id}, {"_id": 0})
    if not categoria:
        raise HTTPException(status_code=404, detail="Categoria não encontrada")
    
    update_data = categoria_data.model_dump()
    await db.categorias_equipamentos.update_one({"id": categoria_id}, {"$set": update_data})
    
    updated_categoria = await db.categorias_equipamentos.find_one({"id": categoria_id}, {"_id": 0})
    return updated_categoria

@api_router.delete("/categorias-equipamentos/{categoria_id}")
async def delete_categoria_equipamento(categoria_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.categorias_equipamentos.delete_one({"id": categoria_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Categoria não encontrada")
    return {"message": "Categoria deletada"}

# EQUIPAMENTOS ROUTES
@api_router.post("/empresas/{empresa_id}/equipamentos", response_model=Equipamento)
async def create_equipamento(empresa_id: str, equipamento_data: EquipamentoCreate, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    equipamento_obj = Equipamento(**equipamento_data.model_dump(), empresa_id=empresa_id)
    doc = equipamento_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.equipamentos.insert_one(doc)
    return equipamento_obj

@api_router.get("/empresas/{empresa_id}/equipamentos", response_model=List[Equipamento])
async def get_equipamentos(empresa_id: str, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    equipamentos = await db.equipamentos.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(1000)
    return equipamentos

@api_router.get("/equipamentos/{equipamento_id}", response_model=Equipamento)
async def get_equipamento(equipamento_id: str, current_user: dict = Depends(get_current_user)):
    equipamento = await db.equipamentos.find_one({"id": equipamento_id}, {"_id": 0})
    if not equipamento:
        raise HTTPException(status_code=404, detail="Equipamento não encontrado")
    return equipamento

@api_router.put("/equipamentos/{equipamento_id}", response_model=Equipamento)
async def update_equipamento(equipamento_id: str, equipamento_data: EquipamentoCreate, current_user: dict = Depends(get_current_user)):
    equipamento = await db.equipamentos.find_one({"id": equipamento_id}, {"_id": 0})
    if not equipamento:
        raise HTTPException(status_code=404, detail="Equipamento não encontrado")
    
    update_data = equipamento_data.model_dump()
    # Preservar quantidade_estoque atual
    update_data['quantidade_estoque'] = equipamento['quantidade_estoque']
    
    await db.equipamentos.update_one({"id": equipamento_id}, {"$set": update_data})
    
    updated_equipamento = await db.equipamentos.find_one({"id": equipamento_id}, {"_id": 0})
    return updated_equipamento

@api_router.delete("/equipamentos/{equipamento_id}")
async def delete_equipamento(equipamento_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.equipamentos.delete_one({"id": equipamento_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Equipamento não encontrado")
    return {"message": "Equipamento deletado"}

# EQUIPAMENTOS SERIALIZADOS ROUTES
@api_router.post("/empresas/{empresa_id}/equipamentos-serializados", response_model=EquipamentoSerializado)
async def create_equipamento_serializado(empresa_id: str, eq_serial_data: EquipamentoSerializadoCreate, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Verificar se número de série já existe
    existing = await db.equipamentos_serializados.find_one({"numero_serie": eq_serial_data.numero_serie, "empresa_id": empresa_id})
    if existing:
        raise HTTPException(status_code=400, detail="Número de série já cadastrado")
    
    eq_serial_obj = EquipamentoSerializado(**eq_serial_data.model_dump(), empresa_id=empresa_id)
    doc = eq_serial_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.equipamentos_serializados.insert_one(doc)
    return eq_serial_obj

@api_router.get("/empresas/{empresa_id}/equipamentos-serializados", response_model=List[EquipamentoSerializado])
async def get_equipamentos_serializados(
    empresa_id: str, 
    status: Optional[str] = None,
    cliente_id: Optional[str] = None,
    equipamento_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    query = {"empresa_id": empresa_id}
    if status:
        query["status"] = status
    if cliente_id:
        query["cliente_id"] = cliente_id
    if equipamento_id:
        query["equipamento_id"] = equipamento_id
    
    equipamentos = await db.equipamentos_serializados.find(query, {"_id": 0}).to_list(1000)
    return equipamentos

@api_router.get("/equipamentos-serializados/{eq_serial_id}", response_model=EquipamentoSerializado)
async def get_equipamento_serializado(eq_serial_id: str, current_user: dict = Depends(get_current_user)):
    eq_serial = await db.equipamentos_serializados.find_one({"id": eq_serial_id}, {"_id": 0})
    if not eq_serial:
        raise HTTPException(status_code=404, detail="Equipamento serializado não encontrado")
    return eq_serial

@api_router.put("/equipamentos-serializados/{eq_serial_id}", response_model=EquipamentoSerializado)
async def update_equipamento_serializado(eq_serial_id: str, eq_serial_data: Dict[str, Any], current_user: dict = Depends(get_current_user)):
    eq_serial = await db.equipamentos_serializados.find_one({"id": eq_serial_id}, {"_id": 0})
    if not eq_serial:
        raise HTTPException(status_code=404, detail="Equipamento serializado não encontrado")
    
    # Se mudou o SIM card, adicionar ao histórico
    if 'numero_simcard' in eq_serial_data and eq_serial_data['numero_simcard'] != eq_serial.get('numero_simcard'):
        if 'historico_simcards' not in eq_serial:
            eq_serial['historico_simcards'] = []
        
        if eq_serial.get('numero_simcard'):
            eq_serial['historico_simcards'].append({
                "numero": eq_serial['numero_simcard'],
                "data_troca": datetime.now(timezone.utc).isoformat()
            })
        eq_serial_data['historico_simcards'] = eq_serial['historico_simcards']
    
    await db.equipamentos_serializados.update_one({"id": eq_serial_id}, {"$set": eq_serial_data})
    
    updated_eq_serial = await db.equipamentos_serializados.find_one({"id": eq_serial_id}, {"_id": 0})
    return updated_eq_serial

@api_router.delete("/equipamentos-serializados/{eq_serial_id}")
async def delete_equipamento_serializado(eq_serial_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.equipamentos_serializados.delete_one({"id": eq_serial_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Equipamento serializado não encontrado")
    return {"message": "Equipamento serializado deletado"}

# MOVIMENTAÇÕES DE ESTOQUE ROUTES
@api_router.post("/empresas/{empresa_id}/movimentacoes", response_model=MovimentacaoEstoque)
async def create_movimentacao(empresa_id: str, mov_data: MovimentacaoEstoqueCreate, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Buscar equipamento
    equipamento = await db.equipamentos.find_one({"id": mov_data.equipamento_id}, {"_id": 0})
    if not equipamento:
        raise HTTPException(status_code=404, detail="Equipamento não encontrado")
    
    mov_obj = MovimentacaoEstoque(
        **mov_data.model_dump(exclude={'categoria_financeira_id', 'centro_custo_id'}),
        empresa_id=empresa_id,
        usuario_id=current_user["id"]
    )
    doc = mov_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    # Processar movimentação baseado no tipo
    if equipamento['tipo_controle'] == 'serializado':
        # Para serializados, atualizar status do equipamento
        if mov_data.equipamento_serializado_id:
            eq_serial = await db.equipamentos_serializados.find_one({"id": mov_data.equipamento_serializado_id}, {"_id": 0})
            if not eq_serial:
                raise HTTPException(status_code=404, detail="Equipamento serializado não encontrado")
            
            update_data = {}
            if mov_data.tipo == "saida_venda":
                update_data = {"status": "vendido", "cliente_id": mov_data.cliente_id, "tipo_vinculo": "venda"}
            elif mov_data.tipo == "saida_locacao":
                update_data = {"status": "em_cliente", "cliente_id": mov_data.cliente_id, "tipo_vinculo": "locacao"}
            elif mov_data.tipo == "devolucao":
                update_data = {"status": "disponivel", "cliente_id": None, "tipo_vinculo": None}
            elif mov_data.tipo == "manutencao":
                update_data = {"status": "em_manutencao"}
            elif mov_data.tipo == "perda":
                update_data = {"status": "baixado"}
            
            if update_data:
                await db.equipamentos_serializados.update_one({"id": mov_data.equipamento_serializado_id}, {"$set": update_data})
    
    else:
        # Para não-serializados, atualizar quantidade em estoque
        if mov_data.tipo == "entrada":
            await db.equipamentos.update_one(
                {"id": mov_data.equipamento_id},
                {"$inc": {"quantidade_estoque": mov_data.quantidade}}
            )
        elif mov_data.tipo in ["saida_venda", "saida_locacao", "perda"]:
            # Verificar se há estoque suficiente
            if equipamento['quantidade_estoque'] < mov_data.quantidade:
                raise HTTPException(status_code=400, detail="Estoque insuficiente")
            
            await db.equipamentos.update_one(
                {"id": mov_data.equipamento_id},
                {"$inc": {"quantidade_estoque": -mov_data.quantidade}}
            )
        elif mov_data.tipo == "devolucao":
            await db.equipamentos.update_one(
                {"id": mov_data.equipamento_id},
                {"$inc": {"quantidade_estoque": mov_data.quantidade}}
            )
    
    # Criar transação financeira se solicitado
    if mov_data.criar_transacao_financeira and mov_data.valor_financeiro and mov_data.categoria_financeira_id and mov_data.centro_custo_id:
        transacao_tipo = "receita" if mov_data.tipo in ["saida_venda", "saida_locacao"] else "despesa"
        
        transacao_obj = Transacao(
            empresa_id=empresa_id,
            usuario_id=current_user["id"],
            tipo=transacao_tipo,
            fornecedor=mov_data.cliente_id or "Sistema de Estoque",
            descricao=f"Movimentação estoque: {mov_data.tipo} - {equipamento['nome']}",
            valor_total=mov_data.valor_financeiro,
            data_competencia=mov_data.data,
            categoria_id=mov_data.categoria_financeira_id,
            centro_custo_id=mov_data.centro_custo_id,
            status="conciliada",
            origem="estoque"
        )
        
        trans_doc = transacao_obj.model_dump()
        trans_doc['created_at'] = trans_doc['created_at'].isoformat()
        
        await db.transacoes.insert_one(trans_doc)
        mov_obj.transacao_id = transacao_obj.id
        doc['transacao_id'] = transacao_obj.id
    
    await db.movimentacoes_estoque.insert_one(doc)
    return mov_obj

@api_router.get("/empresas/{empresa_id}/movimentacoes", response_model=List[MovimentacaoEstoque])
async def get_movimentacoes(
    empresa_id: str,
    tipo: Optional[str] = None,
    equipamento_id: Optional[str] = None,
    cliente_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    query = {"empresa_id": empresa_id}
    if tipo:
        query["tipo"] = tipo
    if equipamento_id:
        query["equipamento_id"] = equipamento_id
    if cliente_id:
        query["cliente_id"] = cliente_id
    
    movimentacoes = await db.movimentacoes_estoque.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return movimentacoes

# ==================== END ESTOQUE ROUTES ====================

# ==================== CRM ROUTES ====================

# LEADS ROUTES
@api_router.post("/empresas/{empresa_id}/leads", response_model=Lead)
async def create_lead(empresa_id: str, lead_data: LeadCreate, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    lead_obj = Lead(**lead_data.model_dump(), empresa_id=empresa_id)
    doc = lead_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    if doc.get('last_contact_at'):
        doc['last_contact_at'] = doc['last_contact_at'].isoformat()
    
    await db.leads.insert_one(doc)
    
    # Registrar atividade
    activity = Activity(
        lead_id=lead_obj.id,
        empresa_id=empresa_id,
        tipo="note",
        descricao=f"Lead criado: {lead_obj.nome}",
        user_id=current_user["id"]
    )
    activity_doc = activity.model_dump()
    activity_doc['created_at'] = activity_doc['created_at'].isoformat()
    await db.activities.insert_one(activity_doc)
    
    return lead_obj

@api_router.get("/empresas/{empresa_id}/leads", response_model=List[Lead])
async def get_leads(
    empresa_id: str,
    status_funil: Optional[str] = None,
    assigned_to: Optional[str] = None,
    origem: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    query = {"empresa_id": empresa_id}
    if status_funil:
        query["status_funil"] = status_funil
    if assigned_to:
        query["assigned_to"] = assigned_to
    if origem:
        query["origem"] = origem
    
    leads = await db.leads.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return leads

@api_router.get("/leads/{lead_id}", response_model=Lead)
async def get_lead(lead_id: str, current_user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead não encontrado")
    return lead

@api_router.put("/leads/{lead_id}", response_model=Lead)
async def update_lead(lead_id: str, lead_data: LeadUpdate, current_user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead não encontrado")
    
    update_data = {k: v for k, v in lead_data.model_dump(exclude_unset=True).items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.leads.update_one({"id": lead_id}, {"$set": update_data})
    
    # Registrar atividade
    activity = Activity(
        lead_id=lead_id,
        empresa_id=lead['empresa_id'],
        tipo="note",
        descricao=f"Lead atualizado",
        user_id=current_user["id"],
        metadata=update_data
    )
    activity_doc = activity.model_dump()
    activity_doc['created_at'] = activity_doc['created_at'].isoformat()
    await db.activities.insert_one(activity_doc)
    
    updated_lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    return updated_lead

@api_router.delete("/leads/{lead_id}")
async def delete_lead(lead_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.leads.delete_one({"id": lead_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lead não encontrado")
    return {"message": "Lead deletado"}

@api_router.patch("/leads/{lead_id}/status")
async def update_lead_status(lead_id: str, status: str, current_user: dict = Depends(get_current_user)):
    """Atualizar status do lead no funil (para Kanban drag-and-drop)"""
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead não encontrado")
    
    old_status = lead.get('status_funil')
    
    await db.leads.update_one(
        {"id": lead_id},
        {"$set": {
            "status_funil": status,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Registrar atividade
    activity = Activity(
        lead_id=lead_id,
        empresa_id=lead['empresa_id'],
        tipo="status_change",
        descricao=f"Status alterado de '{old_status}' para '{status}'",
        user_id=current_user["id"],
        metadata={"old_status": old_status, "new_status": status}
    )
    activity_doc = activity.model_dump()
    activity_doc['created_at'] = activity_doc['created_at'].isoformat()
    await db.activities.insert_one(activity_doc)
    
    return {"message": "Status atualizado", "new_status": status}

@api_router.patch("/leads/{lead_id}/assign")
async def assign_lead(lead_id: str, user_id: str, current_user: dict = Depends(get_current_user)):
    """Atribuir lead a um vendedor"""
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead não encontrado")
    
    old_assigned = lead.get('assigned_to')
    
    await db.leads.update_one(
        {"id": lead_id},
        {"$set": {
            "assigned_to": user_id,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Registrar atividade
    activity = Activity(
        lead_id=lead_id,
        empresa_id=lead['empresa_id'],
        tipo="assignment",
        descricao=f"Lead atribuído ao usuário {user_id}",
        user_id=current_user["id"],
        metadata={"old_assigned": old_assigned, "new_assigned": user_id}
    )
    activity_doc = activity.model_dump()
    activity_doc['created_at'] = activity_doc['created_at'].isoformat()
    await db.activities.insert_one(activity_doc)
    
    return {"message": "Lead atribuído", "assigned_to": user_id}

@api_router.get("/leads/{lead_id}/activities", response_model=List[Activity])
async def get_lead_activities(lead_id: str, current_user: dict = Depends(get_current_user)):
    """Obter histórico de atividades do lead"""
    activities = await db.activities.find({"lead_id": lead_id}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return activities

@api_router.get("/leads/{lead_id}/whatsapp-messages")
async def get_lead_whatsapp_messages(lead_id: str, current_user: dict = Depends(get_current_user)):
    """Obter histórico de mensagens WhatsApp do lead"""
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead não encontrado")
    
    # Buscar transações do WhatsApp com o telefone do lead
    telefone = lead.get('whatsapp_phone') or lead.get('telefone')
    transacoes = await db.transacoes.find({
        "empresa_id": lead['empresa_id'],
        "origem": {"$in": ["whatsapp", "whatsapp_bot"]}
    }, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Filtrar por telefone se disponível
    # Note: Isso é uma aproximação. Idealmente teríamos uma coleção separada de mensagens
    return {
        "lead_id": lead_id,
        "telefone": telefone,
        "messages": [],  # TODO: Implementar busca real de mensagens WhatsApp
        "transacoes_relacionadas": transacoes[:10]
    }

# TEMPLATES ROUTES
@api_router.post("/empresas/{empresa_id}/templates", response_model=MessageTemplate)
async def create_template(empresa_id: str, template_data: MessageTemplateCreate, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    template_obj = MessageTemplate(**template_data.model_dump(), empresa_id=empresa_id)
    doc = template_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.message_templates.insert_one(doc)
    return template_obj

@api_router.get("/empresas/{empresa_id}/templates", response_model=List[MessageTemplate])
async def get_templates(empresa_id: str, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    templates = await db.message_templates.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(1000)
    return templates

@api_router.delete("/templates/{template_id}")
async def delete_template(template_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.message_templates.delete_one({"id": template_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Template não encontrado")
    return {"message": "Template deletado"}

@api_router.post("/leads/{lead_id}/send-message")
async def send_message_to_lead(
    lead_id: str,
    message: str,
    template_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Enviar mensagem para lead via WhatsApp"""
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead não encontrado")
    
    telefone = lead.get('whatsapp_phone') or lead.get('telefone')
    if not telefone:
        raise HTTPException(status_code=400, detail="Lead não possui telefone cadastrado")
    
    # Se template_id fornecido, buscar e interpolar variáveis
    if template_id:
        template = await db.message_templates.find_one({"id": template_id}, {"_id": 0})
        if template:
            message = template['conteudo']
            # Interpolar variáveis
            message = message.replace('{nome}', lead.get('nome', ''))
            message = message.replace('{valor}', str(lead.get('valor_estimado', 0)))
            empresa = await db.empresas.find_one({"id": lead['empresa_id']}, {"_id": 0})
            if empresa:
                message = message.replace('{empresa}', empresa.get('razao_social', ''))
    
    # TODO: Integrar com serviço WhatsApp para enviar mensagem real
    # Por enquanto, apenas registrar atividade
    activity = Activity(
        lead_id=lead_id,
        empresa_id=lead['empresa_id'],
        tipo="whatsapp",
        descricao=f"Mensagem enviada: {message[:50]}...",
        user_id=current_user["id"],
        metadata={"telefone": telefone, "message": message}
    )
    activity_doc = activity.model_dump()
    activity_doc['created_at'] = activity_doc['created_at'].isoformat()
    await db.activities.insert_one(activity_doc)
    
    return {"message": "Mensagem enviada", "telefone": telefone, "conteudo": message}

# AUTOMAÇÃO ROUTES
@api_router.post("/empresas/{empresa_id}/automation-rules", response_model=AutomationRule)
async def create_automation_rule(empresa_id: str, rule_data: AutomationRuleCreate, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    rule_obj = AutomationRule(**rule_data.model_dump(), empresa_id=empresa_id)
    doc = rule_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.automation_rules.insert_one(doc)
    return rule_obj

@api_router.get("/empresas/{empresa_id}/automation-rules", response_model=List[AutomationRule])
async def get_automation_rules(empresa_id: str, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    rules = await db.automation_rules.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(1000)
    return rules

# ROTEAMENTO ROUTES
@api_router.get("/empresas/{empresa_id}/routing-config", response_model=RoutingConfig)
async def get_routing_config(empresa_id: str, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    config = await db.routing_configs.find_one({"empresa_id": empresa_id}, {"_id": 0})
    if not config:
        # Criar configuração padrão
        config_obj = RoutingConfig(empresa_id=empresa_id)
        doc = config_obj.model_dump()
        await db.routing_configs.insert_one(doc)
        return config_obj
    return config

@api_router.put("/empresas/{empresa_id}/routing-config", response_model=RoutingConfig)
async def update_routing_config(
    empresa_id: str,
    tipo: str,
    usuarios_ativos: List[str],
    current_user: dict = Depends(get_current_user)
):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    await db.routing_configs.update_one(
        {"empresa_id": empresa_id},
        {"$set": {"tipo": tipo, "usuarios_ativos": usuarios_ativos, "ativo": True}},
        upsert=True
    )
    
    config = await db.routing_configs.find_one({"empresa_id": empresa_id}, {"_id": 0})
    return config

async def apply_routing(empresa_id: str, lead_id: str):
    """Aplicar roteamento automático ao lead"""
    config = await db.routing_configs.find_one({"empresa_id": empresa_id, "ativo": True}, {"_id": 0})
    if not config or not config.get('usuarios_ativos'):
        return None
    
    if config['tipo'] == 'round_robin':
        usuarios = config['usuarios_ativos']
        ultimo = config.get('ultimo_atribuido')
        
        # Encontrar próximo usuário
        if not ultimo or ultimo not in usuarios:
            proximo = usuarios[0]
        else:
            idx = usuarios.index(ultimo)
            proximo = usuarios[(idx + 1) % len(usuarios)]
        
        # Atualizar lead e config
        await db.leads.update_one({"id": lead_id}, {"$set": {"assigned_to": proximo}})
        await db.routing_configs.update_one(
            {"empresa_id": empresa_id},
            {"$set": {"ultimo_atribuido": proximo}}
        )
        
        return proximo
    
    return None

# AGENTE IA ROUTES
@api_router.post("/empresas/{empresa_id}/ai-agents", response_model=AIAgent)
async def create_ai_agent(empresa_id: str, agent_data: AIAgentCreate, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    agent_obj = AIAgent(**agent_data.model_dump(), empresa_id=empresa_id)
    doc = agent_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.ai_agents.insert_one(doc)
    return agent_obj

@api_router.get("/empresas/{empresa_id}/ai-agents", response_model=List[AIAgent])
async def get_ai_agents(empresa_id: str, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    agents = await db.ai_agents.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(1000)
    return agents

@api_router.put("/ai-agents/{agent_id}", response_model=AIAgent)
async def update_ai_agent(agent_id: str, agent_data: AIAgentCreate, current_user: dict = Depends(get_current_user)):
    agent = await db.ai_agents.find_one({"id": agent_id}, {"_id": 0})
    if not agent:
        raise HTTPException(status_code=404, detail="Agente não encontrado")
    
    update_data = agent_data.model_dump()
    await db.ai_agents.update_one({"id": agent_id}, {"$set": update_data})
    
    updated_agent = await db.ai_agents.find_one({"id": agent_id}, {"_id": 0})
    return updated_agent

@api_router.delete("/ai-agents/{agent_id}")
async def delete_ai_agent(agent_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.ai_agents.delete_one({"id": agent_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Agente não encontrado")
    return {"message": "Agente deletado"}

async def process_ai_response(empresa_id: str, lead_id: str, message: str):
    """Processar mensagem com agente de IA e gerar resposta"""
    # Buscar agente ativo
    agent = await db.ai_agents.find_one({"empresa_id": empresa_id, "ativo": True}, {"_id": 0})
    if not agent:
        return None
    
    # Verificar palavras-chave de ativação
    palavras_chave = agent.get('palavras_chave_ativacao', [])
    if palavras_chave:
        message_lower = message.lower()
        if not any(palavra.lower() in message_lower for palavra in palavras_chave):
            return None
    
    # Buscar lead
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    
    # TODO: Integrar com Emergent LLM para resposta real
    # Por enquanto, resposta simples baseada em regras
    prompt = f"{agent['prompt_sistema']}\n\nMensagem do cliente: {message}\nNome: {lead.get('nome', 'Cliente')}"
    
    # Resposta padrão por enquanto
    if any(palavra in message.lower() for palavra in ['preço', 'valor', 'quanto', 'orçamento']):
        resposta = f"Olá {lead.get('nome', 'Cliente')}! Obrigado pelo interesse. Nossa equipe entrará em contato em breve com um orçamento personalizado. Poderia me informar mais detalhes sobre suas necessidades?"
    elif any(palavra in message.lower() for palavra in ['horário', 'atendimento', 'funciona']):
        resposta = "Nosso horário de atendimento é de segunda a sexta, das 9h às 18h. Como posso ajudá-lo?"
    else:
        resposta = f"Olá! Recebi sua mensagem. Um de nossos consultores entrará em contato em breve. Há algo específico que gostaria de saber?"
    
    return resposta

# MÉTRICAS E DASHBOARD CRM
@api_router.get("/empresas/{empresa_id}/crm/metrics", response_model=CRMMetrics)
async def get_crm_metrics(empresa_id: str, current_user: dict = Depends(get_current_user)):
    """Obter métricas e KPIs do CRM"""
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Buscar todos os leads
    leads = await db.leads.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(10000)
    
    # Total de leads
    total_leads = len(leads)
    
    # Leads por status
    leads_por_status = {}
    for lead in leads:
        status = lead.get('status_funil', 'novo')
        leads_por_status[status] = leads_por_status.get(status, 0) + 1
    
    # Leads por origem
    leads_por_origem = {}
    for lead in leads:
        origem = lead.get('origem', 'manual')
        leads_por_origem[origem] = leads_por_origem.get(origem, 0) + 1
    
    # Taxa de conversão por etapa
    taxa_conversao = {}
    total = total_leads if total_leads > 0 else 1
    for status, count in leads_por_status.items():
        taxa_conversao[status] = (count / total) * 100
    
    # Valor total do pipeline
    valor_total_pipeline = sum(lead.get('valor_estimado', 0) for lead in leads if lead.get('status_funil') not in ['ganho', 'perdido'])
    
    # Leads ganhos e perdidos no mês atual
    now = datetime.now(timezone.utc)
    mes_atual = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
    
    leads_vencidos_mes = len([l for l in leads if l.get('status_funil') == 'ganho' and l.get('updated_at', '') >= mes_atual])
    leads_perdidos_mes = len([l for l in leads if l.get('status_funil') == 'perdido' and l.get('updated_at', '') >= mes_atual])
    
    # Desempenho por vendedor
    vendedores_stats = {}
    for lead in leads:
        assigned_to = lead.get('assigned_to')
        if assigned_to:
            if assigned_to not in vendedores_stats:
                vendedores_stats[assigned_to] = {
                    'user_id': assigned_to,
                    'total_leads': 0,
                    'leads_ganhos': 0,
                    'leads_perdidos': 0,
                    'valor_ganho': 0
                }
            vendedores_stats[assigned_to]['total_leads'] += 1
            if lead.get('status_funil') == 'ganho':
                vendedores_stats[assigned_to]['leads_ganhos'] += 1
                vendedores_stats[assigned_to]['valor_ganho'] += lead.get('valor_estimado', 0)
            elif lead.get('status_funil') == 'perdido':
                vendedores_stats[assigned_to]['leads_perdidos'] += 1
    
    desempenho_vendedores = list(vendedores_stats.values())
    
    # Tempo médio por etapa (simplificado)
    tempo_medio_por_etapa = {
        "novo": 1.5,
        "contatado": 2.0,
        "qualificado": 3.5,
        "proposta": 5.0,
        "negociacao": 7.0
    }
    
    return CRMMetrics(
        total_leads=total_leads,
        leads_por_status=leads_por_status,
        leads_por_origem=leads_por_origem,
        taxa_conversao=taxa_conversao,
        tempo_medio_por_etapa=tempo_medio_por_etapa,
        valor_total_pipeline=valor_total_pipeline,
        leads_vencidos_mes=leads_vencidos_mes,
        leads_perdidos_mes=leads_perdidos_mes,
        desempenho_vendedores=desempenho_vendedores
    )

# RELATÓRIOS EXPORTÁVEIS
@api_router.get("/empresas/{empresa_id}/crm/export/leads")
async def export_leads_csv(empresa_id: str, current_user: dict = Depends(get_current_user)):
    """Exportar leads para CSV"""
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    leads = await db.leads.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(10000)
    
    # Criar CSV
    import io
    import csv
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=['id', 'nome', 'telefone', 'email', 'origem', 'status_funil', 'valor_estimado', 'assigned_to', 'created_at'])
    writer.writeheader()
    
    for lead in leads:
        writer.writerow({
            'id': lead.get('id'),
            'nome': lead.get('nome'),
            'telefone': lead.get('telefone'),
            'email': lead.get('email', ''),
            'origem': lead.get('origem'),
            'status_funil': lead.get('status_funil'),
            'valor_estimado': lead.get('valor_estimado', 0),
            'assigned_to': lead.get('assigned_to', ''),
            'created_at': lead.get('created_at')
        })
    
    output.seek(0)
    return Response(content=output.getvalue(), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=leads.csv"})

# SEQUÊNCIAS DE FOLLOW-UP
@api_router.post("/empresas/{empresa_id}/follow-up-sequences", response_model=FollowUpSequence)
async def create_sequence(empresa_id: str, sequence_data: FollowUpSequenceCreate, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    sequence_obj = FollowUpSequence(**sequence_data.model_dump(), empresa_id=empresa_id)
    doc = sequence_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.follow_up_sequences.insert_one(doc)
    return sequence_obj

@api_router.get("/empresas/{empresa_id}/follow-up-sequences", response_model=List[FollowUpSequence])
async def get_sequences(empresa_id: str, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    sequences = await db.follow_up_sequences.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(1000)
    return sequences

@api_router.delete("/follow-up-sequences/{sequence_id}")
async def delete_sequence(sequence_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.follow_up_sequences.delete_one({"id": sequence_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Sequência não encontrada")
    return {"message": "Sequência deletada"}

# COMPLIANCE LGPD
@api_router.get("/leads/{lead_id}/export-data")
async def export_lead_data(lead_id: str, current_user: dict = Depends(get_current_user)):
    """Exportar todos os dados de um lead (LGPD)"""
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead não encontrado")
    
    # Buscar atividades
    activities = await db.activities.find({"lead_id": lead_id}, {"_id": 0}).to_list(1000)
    
    # Montar pacote de dados
    export_data = {
        "lead": lead,
        "activities": activities,
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "data_subject_rights": "Este é seu pacote completo de dados conforme LGPD"
    }
    
    return export_data

@api_router.delete("/leads/{lead_id}/gdpr-delete")
async def delete_lead_gdpr(lead_id: str, current_user: dict = Depends(get_current_user)):
    """Deletar permanentemente todos os dados de um lead (LGPD)"""
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead não encontrado")
    
    # Deletar lead
    await db.leads.delete_one({"id": lead_id})
    
    # Deletar atividades
    await db.activities.delete_many({"lead_id": lead_id})
    
    # Registrar log de exclusão
    logging.info(f"LGPD: Lead {lead_id} deletado permanentemente por {current_user['id']}")
    
    return {
        "message": "Todos os dados do lead foram deletados permanentemente",
        "lead_id": lead_id,
        "deleted_at": datetime.now(timezone.utc).isoformat()
    }

# ==================== END CRM ROUTES ====================

# TRANSACAO ROUTES
@api_router.post("/empresas/{empresa_id}/transacoes", response_model=Transacao)
async def create_transacao(empresa_id: str, transacao_data: TransacaoCreate, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    transacao_obj = Transacao(**transacao_data.model_dump(), empresa_id=empresa_id, usuario_id=current_user["id"])
    doc = transacao_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.transacoes.insert_one(doc)
    
    # Atualizar saldo da conta bancária se vinculado
    if transacao_data.conta_bancaria_id:
        if transacao_data.tipo == "receita":
            await db.contas_bancarias.update_one(
                {"id": transacao_data.conta_bancaria_id},
                {"$inc": {"saldo_atual": transacao_data.valor_total}}
            )
        else:  # despesa
            await db.contas_bancarias.update_one(
                {"id": transacao_data.conta_bancaria_id},
                {"$inc": {"saldo_atual": -transacao_data.valor_total}}
            )
    
    # Atualizar fatura do cartão se vinculado
    if transacao_data.cartao_credito_id and transacao_data.tipo == "despesa":
        await db.cartoes_credito.update_one(
            {"id": transacao_data.cartao_credito_id},
            {
                "$inc": {
                    "fatura_atual": transacao_data.valor_total,
                    "limite_disponivel": -transacao_data.valor_total
                }
            }
        )
    
    return transacao_obj

@api_router.get("/empresas/{empresa_id}/transacoes", response_model=List[Transacao])
async def get_transacoes(
    empresa_id: str, 
    current_user: dict = Depends(get_current_user),
    categoria_id: Optional[str] = None,
    centro_custo_id: Optional[str] = None,
    conta_bancaria_id: Optional[str] = None,
    fornecedor_id: Optional[str] = None,
    tipo: Optional[str] = None,
    status: Optional[str] = None
):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Build query with filters
    query = {"empresa_id": empresa_id}
    
    if categoria_id:
        query["categoria_id"] = categoria_id
    if centro_custo_id:
        query["centro_custo_id"] = centro_custo_id
    if conta_bancaria_id:
        query["conta_bancaria_id"] = conta_bancaria_id
    if fornecedor_id:
        query["fornecedor_id"] = fornecedor_id
    if tipo:
        query["tipo"] = tipo
    if status:
        query["status"] = status
    
    transacoes = await db.transacoes.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return transacoes

@api_router.delete("/transacoes/{transacao_id}")
async def delete_transacao(transacao_id: str, current_user: dict = Depends(get_current_user)):
    """Delete transaction and reverse balance updates"""
    
    # First, fetch the transaction to get its details before deletion
    transacao = await db.transacoes.find_one({"id": transacao_id})
    
    if not transacao:
        raise HTTPException(status_code=404, detail="Transação não encontrada")
    
    # Verify user has access to this transaction (empresa_id check)
    empresa_id = transacao.get("empresa_id")
    if not empresa_id:
        raise HTTPException(status_code=404, detail="Transação sem empresa associada")
    
    # Verify user belongs to this empresa
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado a esta transação")
    
    # Reverse the balance update if transaction has conta_bancaria_id
    conta_bancaria_id = transacao.get("conta_bancaria_id")
    if conta_bancaria_id:
        tipo = transacao.get("tipo")
        valor_total = transacao.get("valor_total", 0)
        
        # Reverse the balance change (opposite of what was done on creation)
        if tipo == "receita":
            # On creation we added (+), so now subtract (-)
            await db.contas_bancarias.update_one(
                {"id": conta_bancaria_id},
                {"$inc": {"saldo_atual": -valor_total}}
            )
        elif tipo == "despesa":
            # On creation we subtracted (-), so now add back (+)
            await db.contas_bancarias.update_one(
                {"id": conta_bancaria_id},
                {"$inc": {"saldo_atual": valor_total}}
            )
    
    # Now delete the transaction
    result = await db.transacoes.delete_one({"id": transacao_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=500, detail="Erro ao deletar transação")
    
    return {
        "message": "Transação deletada e saldos atualizados com sucesso",
        "saldo_revertido": conta_bancaria_id is not None
    }

# Transferência entre contas
class TransferenciaRequest(BaseModel):
    conta_origem_id: str
    conta_destino_id: str
    valor: float
    descricao: str
    data_transferencia: str

@api_router.post("/empresas/{empresa_id}/transferencias")
async def criar_transferencia(
    empresa_id: str,
    transferencia: TransferenciaRequest,
    current_user: dict = Depends(get_current_user)
):
    """Create transfer between accounts"""
    
    # Validate accounts exist
    conta_origem = await db.contas_bancarias.find_one({"id": transferencia.conta_origem_id}, {"_id": 0})
    conta_destino = await db.contas_bancarias.find_one({"id": transferencia.conta_destino_id}, {"_id": 0})
    
    if not conta_origem:
        raise HTTPException(status_code=404, detail="Conta de origem não encontrada")
    if not conta_destino:
        raise HTTPException(status_code=404, detail="Conta de destino não encontrada")
    
    if transferencia.valor <= 0:
        raise HTTPException(status_code=400, detail="Valor deve ser maior que zero")
    
    # Check if origem has sufficient balance
    if conta_origem.get("saldo_atual", 0) < transferencia.valor:
        raise HTTPException(status_code=400, detail="Saldo insuficiente na conta de origem")
    
    # Get or create default categoria and centro_custo for transfers
    categoria_transferencia = await db.categorias_financeiras.find_one(
        {"empresa_id": empresa_id, "nome": "Transferências"}
    )
    if not categoria_transferencia:
        categoria_transferencia = {
            "id": str(uuid.uuid4()),
            "empresa_id": empresa_id,
            "nome": "Transferências",
            "tipo": "despesa",
            "created_at": datetime.now(timezone.utc)
        }
        await db.categorias_financeiras.insert_one(categoria_transferencia)
    
    centro_custo_default = await db.centros_custo.find_one({"empresa_id": empresa_id})
    centro_custo_id = centro_custo_default.get("id") if centro_custo_default else "default"
    
    # Create two transactions: one despesa (origem) and one receita (destino)
    # Transaction 1: Despesa na conta origem
    transacao_saida = {
        "id": str(uuid.uuid4()),
        "empresa_id": empresa_id,
        "usuario_id": current_user.get("id", "system"),
        "tipo": "despesa",
        "fornecedor": "Transferência",
        "descricao": f"Transferência para {conta_destino.get('nome', 'conta')} - {transferencia.descricao}",
        "valor_total": transferencia.valor,
        "data_competencia": transferencia.data_transferencia,
        "categoria_id": categoria_transferencia["id"],
        "centro_custo_id": centro_custo_id,
        "conta_bancaria_id": transferencia.conta_origem_id,
        "is_transferencia": True,
        "transferencia_relacionada_id": None,  # Will be updated
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    # Transaction 2: Receita na conta destino
    transacao_entrada = {
        "id": str(uuid.uuid4()),
        "empresa_id": empresa_id,
        "usuario_id": current_user.get("id", "system"),
        "tipo": "receita",
        "fornecedor": "Transferência",
        "descricao": f"Transferência de {conta_origem.get('nome', 'conta')} - {transferencia.descricao}",
        "valor_total": transferencia.valor,
        "data_competencia": transferencia.data_transferencia,
        "categoria_id": categoria_transferencia["id"],
        "centro_custo_id": centro_custo_id,
        "conta_bancaria_id": transferencia.conta_destino_id,
        "is_transferencia": True,
        "transferencia_relacionada_id": transacao_saida["id"],
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    # Link transactions
    transacao_saida["transferencia_relacionada_id"] = transacao_entrada["id"]
    
    # Update balances
    await db.contas_bancarias.update_one(
        {"id": transferencia.conta_origem_id},
        {"$inc": {"saldo_atual": -transferencia.valor}}
    )
    
    await db.contas_bancarias.update_one(
        {"id": transferencia.conta_destino_id},
        {"$inc": {"saldo_atual": transferencia.valor}}
    )
    
    # Insert transactions
    await db.transacoes.insert_one(transacao_saida)
    await db.transacoes.insert_one(transacao_entrada)
    
    return {
        "message": "Transferência realizada com sucesso",
        "transacao_saida_id": transacao_saida["id"],
        "transacao_entrada_id": transacao_entrada["id"],
        "valor": transferencia.valor,
        "conta_origem": conta_origem.get("nome"),
        "conta_destino": conta_destino.get("nome")
    }

@api_router.patch("/transacoes/{transacao_id}/status")
async def update_transacao_status(
    transacao_id: str, 
    status: str = Body(..., embed=True),
    current_user: dict = Depends(get_current_user)
):
    """Update transaction status (pendente, concluido, cancelado)"""
    if status not in ["pendente", "concluido", "cancelado"]:
        raise HTTPException(status_code=400, detail="Status inválido. Use: pendente, concluido ou cancelado")
    
    # Check if transaction exists and user has access
    transacao = await db.transacoes.find_one({"id": transacao_id}, {"_id": 0})
    if not transacao:
        raise HTTPException(status_code=404, detail="Transação não encontrada")
    
    if transacao["empresa_id"] not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Update status
    result = await db.transacoes.update_one(
        {"id": transacao_id},
        {"$set": {"status": status}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=400, detail="Não foi possível atualizar o status")
    
    return {"message": f"Status atualizado para {status}", "transacao_id": transacao_id, "status": status}

# AI ROUTES
@api_router.post("/ai/extrair-texto", response_model=ExtrairTextoResponse)
async def extrair_texto(request: ExtrairTextoRequest, current_user: dict = Depends(get_current_user)):
    if request.empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Extract data
    dados = await extrair_dados_com_ai(request.texto, request.empresa_id)
    
    if not dados:
        raise HTTPException(status_code=400, detail="Não foi possível extrair dados do texto")
    
    # Classify if we have description and fornecedor
    classificacao = None
    if dados.get("descricao") and dados.get("fornecedor") and dados.get("valor_total"):
        classificacao = await classificar_com_ai(
            dados["descricao"],
            dados["fornecedor"],
            dados["valor_total"],
            request.empresa_id
        )
    
    return ExtrairTextoResponse(dados_extraidos=dados, classificacao_sugerida=classificacao)

@api_router.post("/ai/classificar")
async def classificar_transacao(
    descricao: str = Form(...),
    fornecedor: str = Form(...),
    valor: float = Form(...),
    empresa_id: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    classificacao = await classificar_com_ai(descricao, fornecedor, valor, empresa_id)
    if not classificacao:
        raise HTTPException(status_code=400, detail="Não foi possível classificar")
    
    return classificacao

# DASHBOARD ROUTES
@api_router.get("/empresas/{empresa_id}/dashboard", response_model=DashboardMetrics)
async def get_dashboard(empresa_id: str, current_user: dict = Depends(get_current_user)):
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Get all transactions
    transacoes = await db.transacoes.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(1000)
    
    # Calculate metrics
    total_receitas = sum(t["valor_total"] for t in transacoes if t["tipo"] == "receita")
    total_despesas = sum(t["valor_total"] for t in transacoes if t["tipo"] == "despesa")
    saldo = total_receitas - total_despesas
    
    # Get bank accounts balance
    contas = await db.contas_bancarias.find({"empresa_id": empresa_id, "ativa": True}, {"_id": 0}).to_list(1000)
    saldo_contas = sum(c.get("saldo_atual", 0) for c in contas)
    num_contas = len(contas)
    
    # Get credit cards balance
    cartoes = await db.cartoes_credito.find({"empresa_id": empresa_id, "ativo": True}, {"_id": 0}).to_list(1000)
    saldo_cartoes = sum(c.get("limite_disponivel", 0) for c in cartoes)
    num_cartoes = len(cartoes)
    
    # Despesas por categoria
    cat_map = {}
    for t in transacoes:
        if t["tipo"] == "despesa":
            cat_id = t.get("categoria_id")
            if cat_id:
                if cat_id not in cat_map:
                    cat = await db.categorias.find_one({"id": cat_id}, {"_id": 0})
                    cat_map[cat_id] = {"nome": cat["nome"] if cat else "Desconhecido", "valor": 0}
                cat_map[cat_id]["valor"] += t["valor_total"]
    
    despesas_por_categoria = [{"categoria": v["nome"], "valor": v["valor"]} for v in cat_map.values()]
    despesas_por_categoria.sort(key=lambda x: x["valor"], reverse=True)
    
    # Despesas por centro de custo
    cc_map = {}
    for t in transacoes:
        if t["tipo"] == "despesa":
            cc_id = t.get("centro_custo_id")
            if cc_id:
                if cc_id not in cc_map:
                    cc = await db.centros_custo.find_one({"id": cc_id}, {"_id": 0})
                    cc_map[cc_id] = {"nome": cc["nome"] if cc else "Desconhecido", "valor": 0}
                cc_map[cc_id]["valor"] += t["valor_total"]
    
    despesas_por_centro_custo = [{"centro_custo": v["nome"], "valor": v["valor"]} for v in cc_map.values()]
    despesas_por_centro_custo.sort(key=lambda x: x["valor"], reverse=True)
    
    # Recent transactions - handle mixed datetime/string types and timezone awareness
    def get_created_at_timestamp(x):
        created = x.get("created_at", "")
        if isinstance(created, datetime):
            # Make timezone-aware if naive
            if created.tzinfo is None:
                return created.replace(tzinfo=timezone.utc)
            return created
        elif isinstance(created, str):
            try:
                dt = datetime.fromisoformat(created.replace('Z', '+00:00'))
                # Make timezone-aware if naive
                if dt.tzinfo is None:
                    return dt.replace(tzinfo=timezone.utc)
                return dt
            except:
                return datetime.min.replace(tzinfo=timezone.utc)
        return datetime.min.replace(tzinfo=timezone.utc)
    
    transacoes_recentes = sorted(transacoes, key=get_created_at_timestamp, reverse=True)[:10]
    
    return DashboardMetrics(
        total_receitas=total_receitas,
        total_despesas=total_despesas,
        saldo=saldo,
        saldo_contas=saldo_contas,
        saldo_cartoes=saldo_cartoes,
        num_contas=num_contas,
        num_cartoes=num_cartoes,
        despesas_por_categoria=despesas_por_categoria,
        despesas_por_centro_custo=despesas_por_centro_custo,
        transacoes_recentes=transacoes_recentes
    )

@api_router.get("/empresas/{empresa_id}/relatorios", response_model=RelatorioDetalhado)
async def get_relatorio_detalhado(
    empresa_id: str,
    periodo_inicio: Optional[str] = None,
    periodo_fim: Optional[str] = None,
    tipo_periodo: str = "mensal",  # mensal, anual, personalizado
    current_user: dict = Depends(get_current_user)
):
    """
    Gerar relatório detalhado com análise por centro de custo e categoria
    """
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Definir período baseado no tipo
    hoje = datetime.now(timezone.utc)
    if tipo_periodo == "mensal":
        periodo_inicio = hoje.replace(day=1).isoformat()
        periodo_fim = hoje.isoformat()
    elif tipo_periodo == "anual":
        periodo_inicio = hoje.replace(month=1, day=1).isoformat()
        periodo_fim = hoje.isoformat()
    # Para personalizado, usar os parâmetros fornecidos
    
    # Buscar transações do período
    query = {"empresa_id": empresa_id}
    if periodo_inicio and periodo_fim:
        query["data_competencia"] = {
            "$gte": periodo_inicio.split("T")[0],
            "$lte": periodo_fim.split("T")[0]
        }
    
    transacoes = await db.transacoes.find(query, {"_id": 0}).to_list(10000)
    
    # Calcular resumo geral
    total_receitas = sum(t["valor_total"] for t in transacoes if t["tipo"] == "receita")
    total_despesas = sum(t["valor_total"] for t in transacoes if t["tipo"] == "despesa")
    lucro = total_receitas - total_despesas
    
    resumo_geral = {
        "total_receitas": total_receitas,
        "total_despesas": total_despesas,
        "lucro": lucro,
        "num_transacoes": len(transacoes),
        "ticket_medio": (total_receitas + total_despesas) / len(transacoes) if transacoes else 0
    }
    
    # Análise por Centro de Custo
    cc_map = {}
    for t in transacoes:
        cc_id = t.get("centro_custo_id")
        if cc_id:
            if cc_id not in cc_map:
                cc = await db.centros_custo.find_one({"id": cc_id}, {"_id": 0})
                cc_map[cc_id] = {
                    "id": cc_id,
                    "nome": cc["nome"] if cc else "Desconhecido",
                    "receitas": 0,
                    "despesas": 0,
                    "num_transacoes": 0
                }
            
            cc_map[cc_id]["num_transacoes"] += 1
            if t["tipo"] == "receita":
                cc_map[cc_id]["receitas"] += t["valor_total"]
            else:
                cc_map[cc_id]["despesas"] += t["valor_total"]
    
    por_centro_custo = []
    for cc_id, data in cc_map.items():
        lucro_cc = data["receitas"] - data["despesas"]
        percentual = ((data["receitas"] + data["despesas"]) / (total_receitas + total_despesas) * 100) if (total_receitas + total_despesas) > 0 else 0
        
        por_centro_custo.append(CentroCustoMetrics(
            centro_custo_id=cc_id,
            centro_custo_nome=data["nome"],
            total_receitas=data["receitas"],
            total_despesas=data["despesas"],
            lucro=lucro_cc,
            num_transacoes=data["num_transacoes"],
            percentual_total=round(percentual, 2)
        ))
    
    por_centro_custo.sort(key=lambda x: abs(x.total_receitas + x.total_despesas), reverse=True)
    
    # Análise por Categoria
    cat_map = {}
    for t in transacoes:
        cat_id = t.get("categoria_id")
        if cat_id:
            if cat_id not in cat_map:
                cat = await db.categorias.find_one({"id": cat_id}, {"_id": 0})
                cat_map[cat_id] = {
                    "id": cat_id,
                    "nome": cat["nome"] if cat else "Desconhecido",
                    "receitas": 0,
                    "despesas": 0,
                    "num_transacoes": 0
                }
            
            cat_map[cat_id]["num_transacoes"] += 1
            if t["tipo"] == "receita":
                cat_map[cat_id]["receitas"] += t["valor_total"]
            else:
                cat_map[cat_id]["despesas"] += t["valor_total"]
    
    por_categoria = []
    for cat_id, data in cat_map.items():
        perc_despesas = (data["despesas"] / total_despesas * 100) if total_despesas > 0 else 0
        perc_receitas = (data["receitas"] / total_receitas * 100) if total_receitas > 0 else 0
        
        por_categoria.append(CategoriaMetrics(
            categoria_id=cat_id,
            categoria_nome=data["nome"],
            total_receitas=data["receitas"],
            total_despesas=data["despesas"],
            num_transacoes=data["num_transacoes"],
            percentual_despesas=round(perc_despesas, 2),
            percentual_receitas=round(perc_receitas, 2)
        ))
    
    por_categoria.sort(key=lambda x: x.total_despesas + x.total_receitas, reverse=True)
    
    return RelatorioDetalhado(
        periodo_inicio=periodo_inicio or hoje.isoformat(),
        periodo_fim=periodo_fim or hoje.isoformat(),
        resumo_geral=resumo_geral,
        por_centro_custo=por_centro_custo,
        por_categoria=por_categoria,
        transacoes=transacoes[:100]  # Limitar a 100 transações para performance
    )

@api_router.get("/empresas/{empresa_id}/relatorios/export/csv")
async def export_relatorio_csv(
    empresa_id: str,
    periodo_inicio: Optional[str] = None,
    periodo_fim: Optional[str] = None,
    tipo_periodo: str = "mensal",
    current_user: dict = Depends(get_current_user)
):
    """Export relatório para CSV"""
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Reutilizar lógica do endpoint de relatórios
    hoje = datetime.now(timezone.utc)
    if tipo_periodo == "mensal":
        periodo_inicio = hoje.replace(day=1).isoformat()
        periodo_fim = hoje.isoformat()
    elif tipo_periodo == "anual":
        periodo_inicio = hoje.replace(month=1, day=1).isoformat()
        periodo_fim = hoje.isoformat()
    
    query = {"empresa_id": empresa_id}
    if periodo_inicio and periodo_fim:
        query["data_competencia"] = {
            "$gte": periodo_inicio.split("T")[0],
            "$lte": periodo_fim.split("T")[0]
        }
    
    transacoes = await db.transacoes.find(query, {"_id": 0}).to_list(10000)
    
    # Criar CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Cabeçalho
    writer.writerow(['ID', 'Data', 'Tipo', 'Fornecedor', 'Categoria', 'Centro de Custo', 'Valor', 'Status', 'Origem'])
    
    # Dados
    for t in transacoes:
        writer.writerow([
            t.get('id', ''),
            t.get('data_competencia', ''),
            t.get('tipo', ''),
            t.get('fornecedor', ''),
            t.get('categoria_id', ''),
            t.get('centro_custo_id', ''),
            t.get('valor_total', 0),
            t.get('status', ''),
            t.get('origem', '')
        ])
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=relatorio_{tipo_periodo}.csv"}
    )

@api_router.get("/empresas/{empresa_id}/relatorios/export/excel")
async def export_relatorio_excel(
    empresa_id: str,
    periodo_inicio: Optional[str] = None,
    periodo_fim: Optional[str] = None,
    tipo_periodo: str = "mensal",
    current_user: dict = Depends(get_current_user)
):
    """Export relatório para Excel"""
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Reutilizar lógica do endpoint de relatórios
    hoje = datetime.now(timezone.utc)
    if tipo_periodo == "mensal":
        periodo_inicio = hoje.replace(day=1).isoformat()
        periodo_fim = hoje.isoformat()
    elif tipo_periodo == "anual":
        periodo_inicio = hoje.replace(month=1, day=1).isoformat()
        periodo_fim = hoje.isoformat()
    
    query = {"empresa_id": empresa_id}
    if periodo_inicio and periodo_fim:
        query["data_competencia"] = {
            "$gte": periodo_inicio.split("T")[0],
            "$lte": periodo_fim.split("T")[0]
        }
    
    transacoes = await db.transacoes.find(query, {"_id": 0}).to_list(10000)
    
    # Calcular métricas
    total_receitas = sum(t["valor_total"] for t in transacoes if t["tipo"] == "receita")
    total_despesas = sum(t["valor_total"] for t in transacoes if t["tipo"] == "despesa")
    
    # Criar workbook
    wb = Workbook()
    
    # Sheet 1: Resumo
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    
    # Título
    ws_resumo['A1'] = 'RELATÓRIO FINANCEIRO'
    ws_resumo['A1'].font = Font(size=16, bold=True)
    ws_resumo.merge_cells('A1:D1')
    
    ws_resumo['A3'] = 'Total de Receitas:'
    ws_resumo['B3'] = f'R$ {total_receitas:,.2f}'
    ws_resumo['B3'].font = Font(color="008000")
    
    ws_resumo['A4'] = 'Total de Despesas:'
    ws_resumo['B4'] = f'R$ {total_despesas:,.2f}'
    ws_resumo['B4'].font = Font(color="FF0000")
    
    ws_resumo['A5'] = 'Lucro/Prejuízo:'
    ws_resumo['B5'] = f'R$ {(total_receitas - total_despesas):,.2f}'
    ws_resumo['B5'].font = Font(color="008000" if (total_receitas - total_despesas) >= 0 else "FF0000", bold=True)
    
    # Sheet 2: Transações
    ws_trans = wb.create_sheet("Transações")
    
    # Cabeçalho
    headers = ['ID', 'Data', 'Tipo', 'Fornecedor', 'Categoria', 'Centro Custo', 'Valor', 'Status', 'Origem']
    ws_trans.append(headers)
    
    # Estilo do cabeçalho
    for cell in ws_trans[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Dados
    for t in transacoes:
        ws_trans.append([
            t.get('id', '')[:8] + '...',
            t.get('data_competencia', ''),
            t.get('tipo', ''),
            t.get('fornecedor', ''),
            t.get('categoria_id', '')[:8] + '...' if t.get('categoria_id') else '',
            t.get('centro_custo_id', '')[:8] + '...' if t.get('centro_custo_id') else '',
            t.get('valor_total', 0),
            t.get('status', ''),
            t.get('origem', '')
        ])
    
    # Salvar em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=relatorio_{tipo_periodo}.xlsx"}
    )

@api_router.get("/empresas/{empresa_id}/relatorios/export/pdf")
async def export_relatorio_pdf(
    empresa_id: str,
    periodo_inicio: Optional[str] = None,
    periodo_fim: Optional[str] = None,
    tipo_periodo: str = "mensal",
    current_user: dict = Depends(get_current_user)
):
    """Export relatório para PDF"""
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Reutilizar lógica do endpoint de relatórios
    hoje = datetime.now(timezone.utc)
    if tipo_periodo == "mensal":
        periodo_inicio = hoje.replace(day=1).isoformat()
        periodo_fim = hoje.isoformat()
    elif tipo_periodo == "anual":
        periodo_inicio = hoje.replace(month=1, day=1).isoformat()
        periodo_fim = hoje.isoformat()
    
    query = {"empresa_id": empresa_id}
    if periodo_inicio and periodo_fim:
        query["data_competencia"] = {
            "$gte": periodo_inicio.split("T")[0],
            "$lte": periodo_fim.split("T")[0]
        }
    
    transacoes = await db.transacoes.find(query, {"_id": 0}).to_list(10000)
    
    # Calcular métricas
    total_receitas = sum(t["valor_total"] for t in transacoes if t["tipo"] == "receita")
    total_despesas = sum(t["valor_total"] for t in transacoes if t["tipo"] == "despesa")
    lucro = total_receitas - total_despesas
    
    # Criar PDF
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4))
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Título
    elements.append(Paragraph("RELATÓRIO FINANCEIRO - ECHO SHOP", title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Resumo
    resumo_data = [
        ['RESUMO FINANCEIRO', ''],
        ['Total de Receitas', f'R$ {total_receitas:,.2f}'],
        ['Total de Despesas', f'R$ {total_despesas:,.2f}'],
        ['Lucro/Prejuízo', f'R$ {lucro:,.2f}'],
        ['Número de Transações', str(len(transacoes))]
    ]
    
    resumo_table = Table(resumo_data, colWidths=[3*inch, 2*inch])
    resumo_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(resumo_table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Transações (primeiras 50)
    if transacoes:
        elements.append(Paragraph("TRANSAÇÕES (Primeiras 50)", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))
        
        trans_data = [['Data', 'Tipo', 'Fornecedor', 'Valor', 'Status']]
        for t in transacoes[:50]:
            trans_data.append([
                t.get('data_competencia', '')[:10],
                t.get('tipo', ''),
                t.get('fornecedor', '')[:20],
                f'R$ {t.get("valor_total", 0):,.2f}',
                t.get('status', '')
            ])
        
        trans_table = Table(trans_data, colWidths=[1.2*inch, 1*inch, 2.5*inch, 1.3*inch, 1*inch])
        trans_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(trans_table)
    
    doc.build(elements)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=relatorio_{tipo_periodo}.pdf"}
    )

# IMPORT ROUTES
@api_router.post("/empresas/{empresa_id}/transacoes/import/csv")
async def import_csv(
    empresa_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Importar transações de arquivo CSV"""
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Arquivo deve ser CSV")
    
    try:
        # Ler CSV
        contents = await file.read()
        df = pd.read_csv(io.BytesIO(contents))
        
        # Validar colunas esperadas
        required_columns = ['data', 'tipo', 'fornecedor', 'valor']
        if not all(col in df.columns for col in required_columns):
            return {
                "status": "error",
                "message": f"CSV deve conter as colunas: {', '.join(required_columns)}",
                "columns_found": list(df.columns)
            }
        
        # Buscar categorias e centros de custo default
        categorias = await db.categorias.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(100)
        centros_custo = await db.centros_custo.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(100)
        
        default_categoria_id = categorias[0]["id"] if categorias else None
        default_centro_custo_id = centros_custo[0]["id"] if centros_custo else None
        
        # Processar cada linha
        imported_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                transacao = {
                    "id": str(uuid.uuid4()),
                    "empresa_id": empresa_id,
                    "usuario_id": current_user["id"],
                    "tipo": str(row['tipo']).lower(),
                    "fornecedor": str(row['fornecedor']),
                    "valor_total": float(row['valor']),
                    "data_competencia": str(row['data']),
                    "categoria_id": str(row.get('categoria_id', default_categoria_id)) if 'categoria_id' in row else default_categoria_id,
                    "centro_custo_id": str(row.get('centro_custo_id', default_centro_custo_id)) if 'centro_custo_id' in row else default_centro_custo_id,
                    "descricao": str(row.get('descricao', '')),
                    "status": "concluido",
                    "origem": "import_csv",
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.transacoes.insert_one(transacao)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Linha {index + 2}: {str(e)}")
        
        return {
            "status": "success",
            "imported": imported_count,
            "total": len(df),
            "errors": errors[:10]  # Limitar a 10 erros
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar CSV: {str(e)}")

@api_router.post("/empresas/{empresa_id}/transacoes/import/excel")
async def import_excel(
    empresa_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Importar transações de arquivo Excel"""
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser Excel (.xlsx ou .xls)")
    
    try:
        # Ler Excel
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        # Validar colunas esperadas
        required_columns = ['data', 'tipo', 'fornecedor', 'valor']
        if not all(col in df.columns for col in required_columns):
            return {
                "status": "error",
                "message": f"Excel deve conter as colunas: {', '.join(required_columns)}",
                "columns_found": list(df.columns)
            }
        
        # Buscar categorias e centros de custo default
        categorias = await db.categorias.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(100)
        centros_custo = await db.centros_custo.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(100)
        
        default_categoria_id = categorias[0]["id"] if categorias else None
        default_centro_custo_id = centros_custo[0]["id"] if centros_custo else None
        
        # Processar cada linha
        imported_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                transacao = {
                    "id": str(uuid.uuid4()),
                    "empresa_id": empresa_id,
                    "usuario_id": current_user["id"],
                    "tipo": str(row['tipo']).lower(),
                    "fornecedor": str(row['fornecedor']),
                    "valor_total": float(row['valor']),
                    "data_competencia": str(row['data'])[:10],
                    "categoria_id": str(row.get('categoria_id', default_categoria_id)) if 'categoria_id' in row else default_categoria_id,
                    "centro_custo_id": str(row.get('centro_custo_id', default_centro_custo_id)) if 'centro_custo_id' in row else default_centro_custo_id,
                    "descricao": str(row.get('descricao', '')),
                    "status": "concluido",
                    "origem": "import_excel",
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.transacoes.insert_one(transacao)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Linha {index + 2}: {str(e)}")
        
        return {
            "status": "success",
            "imported": imported_count,
            "total": len(df),
            "errors": errors[:10]
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar Excel: {str(e)}")

@api_router.post("/empresas/{empresa_id}/transacoes/import/ofx")
async def import_ofx(
    empresa_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Importar transações de arquivo OFX (extrato bancário)"""
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    if not file.filename.endswith('.ofx'):
        raise HTTPException(status_code=400, detail="Arquivo deve ser OFX")
    
    try:
        # Ler OFX
        contents = await file.read()
        ofx = OfxParser.parse(io.BytesIO(contents))
        
        # Buscar categorias e centros de custo default
        categorias = await db.categorias.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(100)
        centros_custo = await db.centros_custo.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(100)
        
        default_categoria_id = categorias[0]["id"] if categorias else None
        default_centro_custo_id = centros_custo[0]["id"] if centros_custo else None
        
        # Processar transações
        imported_count = 0
        errors = []
        
        account = ofx.account
        for transaction in account.statement.transactions:
            try:
                # Determinar tipo baseado no valor
                tipo = "receita" if transaction.amount > 0 else "despesa"
                valor_abs = abs(transaction.amount)
                
                transacao = {
                    "id": str(uuid.uuid4()),
                    "empresa_id": empresa_id,
                    "usuario_id": current_user["id"],
                    "tipo": tipo,
                    "fornecedor": transaction.payee or "Não informado",
                    "valor_total": valor_abs,
                    "data_competencia": transaction.date.strftime("%Y-%m-%d"),
                    "categoria_id": default_categoria_id,
                    "centro_custo_id": default_centro_custo_id,
                    "descricao": transaction.memo or "",
                    "status": "concluido",
                    "origem": "import_ofx",
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.transacoes.insert_one(transacao)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Transação {transaction.id}: {str(e)}")
        
        return {
            "status": "success",
            "imported": imported_count,
            "account": account.number if hasattr(account, 'number') else "N/A",
            "errors": errors[:10]
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar OFX: {str(e)}")

@api_router.post("/empresas/{empresa_id}/transacoes/import/pdf")
async def import_pdf(
    empresa_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Importar transações de arquivo PDF (usando OCR via IA)"""
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Arquivo deve ser PDF")
    
    try:
        # Ler PDF
        contents = await file.read()
        pdf_reader = PdfReader(io.BytesIO(contents))
        
        # Extrair texto de todas as páginas
        texto_completo = ""
        for page in pdf_reader.pages:
            texto_completo += page.extract_text() + "\n"
        
        if not texto_completo.strip():
            raise HTTPException(status_code=400, detail="Não foi possível extrair texto do PDF")
        
        # Usar IA para extrair dados financeiros
        emergent_key = os.environ.get("EMERGENT_LLM_KEY")
        if not emergent_key:
            raise HTTPException(status_code=500, detail="Chave de IA não configurada")
        
        llm = LlmChat(
            platform="openai",
            model="gpt-4o-mini",
            api_key=emergent_key
        )
        
        prompt = f"""Analise este extrato/fatura e extraia todas as transações financeiras.
        
Texto do PDF:
{texto_completo[:3000]}

Retorne um JSON com lista de transações no formato:
{{
  "transacoes": [
    {{
      "data": "YYYY-MM-DD",
      "tipo": "receita" ou "despesa",
      "fornecedor": "nome do fornecedor/cliente",
      "valor": número (apenas valor numérico),
      "descricao": "descrição breve"
    }}
  ]
}}

Retorne APENAS o JSON, sem texto adicional."""

        response = llm.run([UserMessage(content=prompt)])
        
        # Parse resposta
        import json
        try:
            # Tentar extrair JSON da resposta
            resposta_texto = response.content[0]["text"]
            # Remover markdown se presente
            if "```json" in resposta_texto:
                resposta_texto = resposta_texto.split("```json")[1].split("```")[0]
            elif "```" in resposta_texto:
                resposta_texto = resposta_texto.split("```")[1].split("```")[0]
            
            dados = json.loads(resposta_texto.strip())
            transacoes_extraidas = dados.get("transacoes", [])
        except:
            raise HTTPException(status_code=500, detail="Não foi possível processar resposta da IA")
        
        # Buscar categorias e centros de custo default
        categorias = await db.categorias.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(100)
        centros_custo = await db.centros_custo.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(100)
        
        default_categoria_id = categorias[0]["id"] if categorias else None
        default_centro_custo_id = centros_custo[0]["id"] if centros_custo else None
        
        # Inserir transações
        imported_count = 0
        errors = []
        
        for trans in transacoes_extraidas:
            try:
                transacao = {
                    "id": str(uuid.uuid4()),
                    "empresa_id": empresa_id,
                    "usuario_id": current_user["id"],
                    "tipo": trans.get("tipo", "despesa"),
                    "fornecedor": trans.get("fornecedor", "Não informado"),
                    "valor_total": float(trans.get("valor", 0)),
                    "data_competencia": trans.get("data", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
                    "categoria_id": default_categoria_id,
                    "centro_custo_id": default_centro_custo_id,
                    "descricao": trans.get("descricao", ""),
                    "status": "pendente",  # Pendente para revisão
                    "origem": "import_pdf",
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.transacoes.insert_one(transacao)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Erro: {str(e)}")
        
        return {
            "status": "success",
            "imported": imported_count,
            "total": len(transacoes_extraidas),
            "errors": errors[:10],
            "message": "Transações importadas como 'pendente' para revisão"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar PDF: {str(e)}")

# AI ANALYSIS ROUTES
@api_router.get("/empresas/{empresa_id}/ai/analise-financeira")
async def analise_financeira_ai(
    empresa_id: str,
    periodo_dias: int = 30,
    current_user: dict = Depends(get_current_user)
):
    """Análise financeira completa com IA: padrões, sugestões e insights"""
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    try:
        # Buscar dados dos últimos N dias
        data_inicio = (datetime.now(timezone.utc) - timedelta(days=periodo_dias)).strftime("%Y-%m-%d")
        
        transacoes = await db.transacoes.find({
            "empresa_id": empresa_id,
            "data_competencia": {"$gte": data_inicio}
        }, {"_id": 0}).to_list(1000)
        
        if not transacoes:
            return {
                "status": "no_data",
                "message": "Não há transações suficientes para análise"
            }
        
        # Calcular métricas básicas
        total_receitas = sum(t["valor_total"] for t in transacoes if t["tipo"] == "receita")
        total_despesas = sum(t["valor_total"] for t in transacoes if t["tipo"] == "despesa")
        
        # Agrupar por categoria
        despesas_por_categoria = {}
        for t in transacoes:
            if t["tipo"] == "despesa":
                cat_id = t.get("categoria_id", "sem_categoria")
                if cat_id not in despesas_por_categoria:
                    cat = await db.categorias.find_one({"id": cat_id}, {"_id": 0})
                    despesas_por_categoria[cat_id] = {
                        "nome": cat["nome"] if cat else "Sem categoria",
                        "valor": 0,
                        "transacoes": 0
                    }
                despesas_por_categoria[cat_id]["valor"] += t["valor_total"]
                despesas_por_categoria[cat_id]["transacoes"] += 1
        
        # Preparar dados para IA
        resumo_financeiro = f"""
Análise Financeira - Últimos {periodo_dias} dias:

RESUMO GERAL:
- Total de Receitas: R$ {total_receitas:,.2f}
- Total de Despesas: R$ {total_despesas:,.2f}
- Saldo: R$ {(total_receitas - total_despesas):,.2f}
- Total de Transações: {len(transacoes)}

DESPESAS POR CATEGORIA:
"""
        for cat_data in sorted(despesas_por_categoria.values(), key=lambda x: x["valor"], reverse=True):
            percentual = (cat_data["valor"] / total_despesas * 100) if total_despesas > 0 else 0
            resumo_financeiro += f"- {cat_data['nome']}: R$ {cat_data['valor']:,.2f} ({percentual:.1f}%) - {cat_data['transacoes']} transações\n"
        
        # Tentar usar IA, mas fornecer fallback se falhar
        analise_texto = ""
        ia_disponivel = False
        
        try:
            emergent_key = os.environ.get("EMERGENT_LLM_KEY")
            if emergent_key:
                llm = LlmChat(
                    api_key=emergent_key,
                    session_id=f"analise-{empresa_id}",
                    system_message="Você é um consultor financeiro especializado."
                ).with_model("openai", "gpt-4o-mini")
                
                prompt = f"""{resumo_financeiro}

Como consultor financeiro especializado, analise estes dados e forneça:

1. **PADRÕES IDENTIFICADOS**: Principais tendências de gastos
2. **OPORTUNIDADES DE OTIMIZAÇÃO**: 3-5 sugestões práticas para reduzir custos
3. **ALERTAS**: Categorias com gastos acima do esperado
4. **RECOMENDAÇÕES**: Ações prioritárias para melhorar a saúde financeira

Seja objetivo, prático e focado em ações concretas. Use formato claro com bullets."""

                user_message = UserMessage(text=prompt)
                response = await llm.send_message(user_message)
                analise_texto = response
                ia_disponivel = True
        except Exception as ia_error:
            # Fallback: Análise estatística básica
            analise_texto = f"""**ANÁLISE ESTATÍSTICA AUTOMÁTICA**

**RESUMO FINANCEIRO:**
- Receitas: R$ {total_receitas:,.2f}
- Despesas: R$ {total_despesas:,.2f}
- Saldo: R$ {(total_receitas - total_despesas):,.2f}
- Status: {"✅ Positivo" if total_receitas > total_despesas else "⚠️ Negativo"}

**PRINCIPAIS CATEGORIAS DE DESPESAS:**
"""
            top_categorias = sorted(despesas_por_categoria.values(), key=lambda x: x["valor"], reverse=True)[:5]
            for i, cat in enumerate(top_categorias, 1):
                percentual = (cat["valor"] / total_despesas * 100) if total_despesas > 0 else 0
                analise_texto += f"\n{i}. {cat['nome']}: R$ {cat['valor']:,.2f} ({percentual:.1f}%)"
            
            analise_texto += "\n\n**ALERTAS:**"
            if total_despesas > total_receitas:
                deficit = total_despesas - total_receitas
                analise_texto += f"\n- ⚠️ Déficit de R$ {deficit:,.2f} no período"
            
            if despesas_por_categoria:
                maior_cat = max(despesas_por_categoria.values(), key=lambda x: x["valor"])
                percentual_maior = (maior_cat["valor"] / total_despesas * 100) if total_despesas > 0 else 0
                if percentual_maior > 40:
                    analise_texto += f"\n- ⚠️ Categoria '{maior_cat['nome']}' representa {percentual_maior:.1f}% das despesas"
            
            analise_texto += "\n\n**RECOMENDAÇÕES:**"
            analise_texto += "\n- Revisar as 3 maiores categorias de despesas"
            analise_texto += "\n- Buscar alternativas para reduzir custos recorrentes"
            if total_receitas > 0:
                analise_texto += f"\n- Manter controle para não exceder receitas (margem: {((total_receitas - total_despesas) / total_receitas * 100):.1f}%)"
            
            analise_texto += "\n\n_Nota: Análise com IA temporariamente indisponível. Usando análise estatística._"
        
        return {
            "status": "success",
            "periodo_dias": periodo_dias,
            "metricas": {
                "total_receitas": total_receitas,
                "total_despesas": total_despesas,
                "saldo": total_receitas - total_despesas,
                "num_transacoes": len(transacoes)
            },
            "analise_ia": analise_texto,
            "despesas_por_categoria": despesas_por_categoria
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro na análise: {str(e)}")

@api_router.get("/empresas/{empresa_id}/ai/anomalias")
async def detectar_anomalias(
    empresa_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Detectar anomalias e gastos incomuns usando IA"""
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    try:
        # Buscar últimos 60 dias de transações
        data_inicio = (datetime.now(timezone.utc) - timedelta(days=60)).strftime("%Y-%m-%d")
        
        transacoes = await db.transacoes.find({
            "empresa_id": empresa_id,
            "tipo": "despesa",
            "data_competencia": {"$gte": data_inicio}
        }, {"_id": 0}).to_list(1000)
        
        if len(transacoes) < 10:
            return {
                "status": "insufficient_data",
                "message": "Necessário pelo menos 10 transações para análise de anomalias",
                "anomalias": []
            }
        
        # Calcular estatísticas por categoria
        stats_por_categoria = {}
        for t in transacoes:
            cat_id = t.get("categoria_id", "sem_categoria")
            if cat_id not in stats_por_categoria:
                stats_por_categoria[cat_id] = []
            stats_por_categoria[cat_id].append(t["valor_total"])
        
        # Detectar valores outliers (acima de 2 desvios padrão)
        import statistics
        anomalias = []
        
        for cat_id, valores in stats_por_categoria.items():
            if len(valores) < 3:
                continue
            
            media = statistics.mean(valores)
            desvio = statistics.stdev(valores) if len(valores) > 1 else 0
            
            for t in transacoes:
                if t.get("categoria_id") == cat_id:
                    if t["valor_total"] > media + (2 * desvio) and desvio > 0:
                        cat = await db.categorias.find_one({"id": cat_id}, {"_id": 0})
                        anomalias.append({
                            "transacao_id": t["id"],
                            "data": t["data_competencia"],
                            "fornecedor": t["fornecedor"],
                            "valor": t["valor_total"],
                            "categoria": cat["nome"] if cat else "Sem categoria",
                            "media_categoria": round(media, 2),
                            "desvio": round((t["valor_total"] - media) / desvio, 2) if desvio > 0 else 0,
                            "tipo_alerta": "valor_acima_media"
                        })
        
        # Ordenar por desvio (mais críticos primeiro)
        anomalias.sort(key=lambda x: x["desvio"], reverse=True)
        
        return {
            "status": "success",
            "num_anomalias": len(anomalias),
            "anomalias": anomalias[:10],  # Top 10
            "message": f"Encontradas {len(anomalias)} transações com valores acima do padrão"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro na detecção: {str(e)}")

@api_router.get("/empresas/{empresa_id}/ai/previsao-fluxo")
async def previsao_fluxo_caixa(
    empresa_id: str,
    dias_futuros: int = 30,
    current_user: dict = Depends(get_current_user)
):
    """Prever fluxo de caixa futuro usando IA"""
    if empresa_id not in current_user.get("empresa_ids", []):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    try:
        # Buscar histórico dos últimos 90 dias
        data_inicio = (datetime.now(timezone.utc) - timedelta(days=90)).strftime("%Y-%m-%d")
        
        transacoes = await db.transacoes.find({
            "empresa_id": empresa_id,
            "data_competencia": {"$gte": data_inicio}
        }, {"_id": 0}).to_list(1000)
        
        if len(transacoes) < 15:
            return {
                "status": "insufficient_data",
                "message": "Necessário pelo menos 15 transações para previsão"
            }
        
        # Calcular médias mensais
        total_receitas = sum(t["valor_total"] for t in transacoes if t["tipo"] == "receita")
        total_despesas = sum(t["valor_total"] for t in transacoes if t["tipo"] == "despesa")
        
        media_receitas_mensal = (total_receitas / 90) * 30
        media_despesas_mensal = (total_despesas / 90) * 30
        
        # Preparar dados para IA
        resumo = f"""
HISTÓRICO FINANCEIRO (últimos 90 dias):
- Total Receitas: R$ {total_receitas:,.2f}
- Total Despesas: R$ {total_despesas:,.2f}
- Média Mensal Receitas: R$ {media_receitas_mensal:,.2f}
- Média Mensal Despesas: R$ {media_despesas_mensal:,.2f}
- Total de Transações: {len(transacoes)}
"""
        
        # Tentar usar IA com fallback
        previsao_texto = ""
        
        try:
            emergent_key = os.environ.get("EMERGENT_LLM_KEY")
            if emergent_key:
                llm = LlmChat(
                    api_key=emergent_key,
                    session_id=f"analise-{empresa_id}",
                    system_message="Você é um consultor financeiro especializado."
                ).with_model("openai", "gpt-4o-mini")
                
                prompt = f"""{resumo}

Como analista financeiro, baseado neste histórico, forneça uma previsão para os próximos {dias_futuros} dias:

1. **CENÁRIO PROVÁVEL**: Estimativa de receitas e despesas
2. **FATORES DE RISCO**: O que pode impactar negativamente
3. **OPORTUNIDADES**: O que pode melhorar os resultados
4. **RECOMENDAÇÕES**: Ações para os próximos {dias_futuros} dias

Seja realista e baseie-se nos dados históricos."""

                response = llm.run([UserMessage(content=prompt)])
                previsao_texto = response.content[0]["text"]
        except Exception:
            # Fallback: Análise baseada em médias
            # Calcular previsão simples (proporcional)
            fator_dias = dias_futuros / 30
            previsao_receitas = media_receitas_mensal * fator_dias
            previsao_despesas = media_despesas_mensal * fator_dias
            
            previsao_texto = f"""**PREVISÃO ESTATÍSTICA - Próximos {dias_futuros} dias**

**CENÁRIO PROVÁVEL (baseado em médias):**
- Receitas Estimadas: R$ {previsao_receitas:,.2f}
- Despesas Estimadas: R$ {previsao_despesas:,.2f}
- Saldo Projetado: R$ {(previsao_receitas - previsao_despesas):,.2f}

**TENDÊNCIA:**
"""
            if previsao_receitas > previsao_despesas:
                margem = ((previsao_receitas - previsao_despesas) / previsao_receitas * 100)
                previsao_texto += f"- ✅ Projeção positiva com margem de {margem:.1f}%"
            else:
                deficit = previsao_despesas - previsao_receitas
                previsao_texto += f"- ⚠️ Projeção de déficit de R$ {deficit:,.2f}"
            
            previsao_texto += f"""

**FATORES A CONSIDERAR:**
- Baseado em média de {len(transacoes)} transações dos últimos 90 dias
- Sazonalidade e eventos específicos não considerados
- Recomenda-se revisão quinzenal da previsão

**RECOMENDAÇÕES:**
- Manter reserva de emergência equivalente a 1 mês de despesas
- Monitorar receitas semanalmente
- Revisar despesas fixas mensalmente

_Nota: Análise com IA temporariamente indisponível. Usando projeção estatística simples._"""
        
        # Calcular previsão simples (proporcional) para retorno
        fator_dias = dias_futuros / 30
        previsao_receitas = media_receitas_mensal * fator_dias
        previsao_despesas = media_despesas_mensal * fator_dias
        
        return {
            "status": "success",
            "periodo_analise_dias": 90,
            "dias_futuros": dias_futuros,
            "previsao_numerica": {
                "receitas_estimadas": round(previsao_receitas, 2),
                "despesas_estimadas": round(previsao_despesas, 2),
                "saldo_estimado": round(previsao_receitas - previsao_despesas, 2)
            },
            "previsao_ia": previsao_texto,
            "historico": {
                "media_receitas_mensal": round(media_receitas_mensal, 2),
                "media_despesas_mensal": round(media_despesas_mensal, 2)
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro na previsão: {str(e)}")

# WHATSAPP MESSAGE PROCESSING (internal service)
class WhatsAppMessageRequest(BaseModel):
    phone_number: str
    sender_name: str
    message: str

class WhatsAppAudioRequest(BaseModel):
    audio_base64: str

@api_router.post("/whatsapp/transcribe")
async def transcribe_audio(request: WhatsAppAudioRequest):
    """Transcribe audio - placeholder for now"""
    try:
        # For MVP, return a message asking to send text
        # In production, integrate with Whisper API or Speech-to-Text service
        return {
            "transcription": "Por favor, envie mensagem de texto com: fornecedor, valor e descrição. Exemplo: Paguei R$ 100 para a empresa ABC hoje"
        }
    except Exception as e:
        logging.error(f"Error transcribing audio: {e}")
        return {"transcription": None}

@api_router.post("/whatsapp/process")
async def process_whatsapp_message(request: WhatsAppMessageRequest):
    """Process WhatsApp messages - Internal service endpoint"""
    try:
        # Normalize phone number for comparison (remove all non-digits)
        normalized_phone = ''.join(filter(str.isdigit, request.phone_number))
        
        # Try to find user by normalized phone number
        users_cursor = db.users.find({}, {"_id": 0})
        user = None
        async for u in users_cursor:
            u_phone = ''.join(filter(str.isdigit, u.get("telefone", "")))
            if u_phone == normalized_phone:
                user = u
                break
        
        # Determine which empresa to use
        empresa = None
        if user and user.get("empresa_ids"):
            # User found - use their first empresa
            empresa_id = user["empresa_ids"][0]
            empresa = await db.empresas.find_one({"id": empresa_id}, {"_id": 0})
        else:
            # No user found - try environment variable for default empresa
            default_empresa_id = os.environ.get("WHATSAPP_DEFAULT_EMPRESA_ID")
            if default_empresa_id:
                empresa = await db.empresas.find_one({"id": default_empresa_id}, {"_id": 0})
            
            # If still no empresa, use most recently created one (more likely to be active)
            if not empresa:
                empresas_list = await db.empresas.find({}, {"_id": 0}).sort("created_at", -1).limit(1).to_list(1)
                if empresas_list:
                    empresa = empresas_list[0]
        
        if not empresa:
            return {
                "response_message": "❌ Nenhuma empresa cadastrada no sistema. Configure uma empresa primeiro."
            }
        
        # WhatsApp agora é usado APENAS para CRM - Não cria mais transações financeiras
        response_text = "✅ Mensagem recebida!\\n\\n"
        response_text += "Olá! Sua mensagem foi registrada no nosso sistema.\\n"
        response_text += "Nossa equipe entrará em contato em breve.\\n"
        
        # Get or create empresa-specific WhatsApp user for CRM
        whatsapp_email = f"whatsapp-bot-{empresa['id'][:8]}@echoshop.com"
        default_user = await db.users.find_one({"email": whatsapp_email}, {"_id": 0})
        if not default_user:
            # Create empresa-specific WhatsApp bot user
            default_user = {
                "id": str(uuid.uuid4()),
                "nome": f"WhatsApp Bot - {empresa.get('razao_social', 'ECHO SHOP')}",
                "email": whatsapp_email,
                "telefone": request.phone_number,
                "perfil": "crm",
                "empresa_ids": [empresa["id"]],
                "senha_hash": pwd_context.hash("whatsapp-bot-user"),
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.users.insert_one(default_user)
        
        # ==================== CRM: Criar/Atualizar Lead ====================
        try:
            # Buscar lead existente pelo telefone
            existing_lead = await db.leads.find_one({
                "empresa_id": empresa["id"],
                "$or": [
                    {"telefone": request.phone_number},
                    {"whatsapp_phone": request.phone_number}
                ]
            }, {"_id": 0})
            
            if existing_lead:
                # Atualizar lead existente
                await db.leads.update_one(
                    {"id": existing_lead["id"]},
                    {"$set": {
                        "last_contact_at": datetime.now(timezone.utc).isoformat(),
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
                lead_id = existing_lead["id"]
                response_text += f"\\n👤 Lead atualizado!"
            else:
                # Criar novo lead
                lead_data = {
                    "id": str(uuid.uuid4()),
                    "empresa_id": empresa["id"],
                    "nome": request.contact_name or f"Contato {request.phone_number}",
                    "telefone": request.phone_number,
                    "whatsapp_phone": request.phone_number,
                    "email": None,
                    "origem": "whatsapp",
                    "status_funil": "novo",
                    "tags": ["whatsapp"],
                    "valor_estimado": 0.0,
                    "assigned_to": None,
                    "last_contact_at": datetime.now(timezone.utc).isoformat(),
                    "notes": f"Lead criado automaticamente via WhatsApp. Primeira mensagem: {request.message[:100]}",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.leads.insert_one(lead_data)
                lead_id = lead_data["id"]
                
                # Aplicar roteamento automático
                assigned_user = await apply_routing(empresa["id"], lead_id)
                
                # Registrar atividade
                activity = {
                    "id": str(uuid.uuid4()),
                    "lead_id": lead_id,
                    "empresa_id": empresa["id"],
                    "tipo": "whatsapp",
                    "descricao": f"Lead criado via WhatsApp: {request.message[:100]}",
                    "user_id": default_user["id"],
                    "metadata": {"telefone": request.phone_number, "primeira_mensagem": request.message},
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.activities.insert_one(activity)
                
                if assigned_user:
                    response_text += f"\\n🎯 Novo lead criado e atribuído automaticamente!"
                else:
                    response_text += f"\\n🎯 Novo lead criado!"
                
                # Tentar resposta automática do agente IA
                ai_response = await process_ai_response(empresa["id"], lead_id, request.message)
                if ai_response:
                    response_text += f"\\n\\n🤖 {ai_response}"
                
        except Exception as e:
            logging.error(f"Error creating/updating lead: {e}")
            response_text += f"\\n⚠️ Lead não foi criado automaticamente."
        # ==================== END CRM ====================
        
        return {
            "response_message": response_text
        }
        
    except Exception as e:
        logging.error(f"Error processing WhatsApp message: {e}")
        return {
            "response_message": f"❌ Erro ao processar mensagem: {str(e)}"
        }

# WhatsApp Service Proxy Routes
@api_router.get("/whatsapp/status")
@limiter.limit("30/minute")
async def whatsapp_status(request: Request, current_user: dict = Depends(get_current_user)):
    """Proxy to WhatsApp service status"""
    try:
        import httpx
        whatsapp_url = os.environ.get("WHATSAPP_SERVICE_URL", "http://localhost:8002")
        async with httpx.AsyncClient() as client:
            response = await client.get(f"{whatsapp_url}/status", timeout=5.0)
            return response.json()
    except Exception as e:
        logging.error(f"Error getting WhatsApp status: {e}")
        return {"status": "service_offline", "phone_number": None, "has_qr": False}

@api_router.get("/whatsapp/qr")
async def whatsapp_qr(current_user: dict = Depends(get_current_user)):
    """Proxy to WhatsApp service QR code"""
    try:
        import httpx
        whatsapp_url = os.environ.get("WHATSAPP_SERVICE_URL", "http://localhost:8002")
        async with httpx.AsyncClient() as client:
            response = await client.get(f"{whatsapp_url}/qr", timeout=5.0)
            return response.json()
    except Exception as e:
        logging.error(f"Error getting WhatsApp QR: {e}")
        raise HTTPException(status_code=404, detail="QR Code não disponível")

@api_router.post("/whatsapp/reconnect")
async def whatsapp_reconnect(current_user: dict = Depends(get_current_user)):
    """Proxy to WhatsApp service reconnect"""
    try:
        import httpx
        whatsapp_url = os.environ.get("WHATSAPP_SERVICE_URL", "http://localhost:8002")
        async with httpx.AsyncClient() as client:
            response = await client.post(f"{whatsapp_url}/reconnect", timeout=10.0)
            return response.json()
    except Exception as e:
        logging.error(f"Error reconnecting WhatsApp: {e}")
        raise HTTPException(status_code=500, detail="Erro ao reconectar")

@api_router.post("/whatsapp/disconnect")
async def whatsapp_disconnect(current_user: dict = Depends(get_current_user)):
    """Proxy to WhatsApp service disconnect"""
    try:
        import httpx
        whatsapp_url = os.environ.get("WHATSAPP_SERVICE_URL", "http://localhost:8002")
        async with httpx.AsyncClient() as client:
            response = await client.post(f"{whatsapp_url}/disconnect", timeout=10.0)
            return response.json()
    except Exception as e:
        logging.error(f"Error disconnecting WhatsApp: {e}")
        raise HTTPException(status_code=500, detail="Erro ao desconectar")

# ==================== GOOGLE DRIVE BACKUP ENDPOINTS ====================

# DEPRECATED: Old service account method - no longer used
# Kept for reference only. Use get_drive_service(user_id) with OAuth instead.
def get_drive_service_deprecated():
    """DEPRECATED: Create and return Google Drive service using service account"""
    service_account_path = os.environ.get("GOOGLE_SERVICE_ACCOUNT_PATH", "/app/backend/service_account.json")
    
    if not os.path.exists(service_account_path):
        raise HTTPException(status_code=500, detail="Google Service Account não configurado")
    
    SCOPES = ['https://www.googleapis.com/auth/drive.file']
    credentials = service_account.Credentials.from_service_account_file(
        service_account_path, scopes=SCOPES)
    
    return build('drive', 'v3', credentials=credentials)

async def export_all_data():
    """Export all database data to JSON format"""
    backup_data = {
        "backup_date": datetime.now(timezone.utc).isoformat(),
        "database": os.environ.get("DB_NAME", "finai_database"),
        "collections": {}
    }
    
    # Collections to backup
    collections = [
        "empresas", "users", "categorias_financeiras", "centros_custo",
        "transacoes", "contas_bancarias", "investimentos", "cartoes_credito",
        "clientes", "fornecedores", "locais", "categorias_equipamentos",
        "equipamentos", "equipamentos_serializados", "movimentacoes_estoque",
        "contatos_crm", "conversas_crm", "mensagens_crm", "agentes_ia",
        "funil_stages", "templates_mensagem", "automacoes_crm", "auditoria_crm",
        "consentimentos_crm", "vendas_clientes", "planos_internet", "assinaturas"
    ]
    
    for collection_name in collections:
        try:
            collection = db[collection_name]
            documents = await collection.find().to_list(length=None)
            
            # Convert ObjectId and datetime to strings
            for doc in documents:
                if "_id" in doc:
                    doc["_id"] = str(doc["_id"])
                for key, value in doc.items():
                    if isinstance(value, datetime):
                        doc[key] = value.isoformat()
            
            backup_data["collections"][collection_name] = documents
            logging.info(f"Exported {len(documents)} documents from {collection_name}")
        except Exception as e:
            logging.error(f"Error exporting {collection_name}: {e}")
            backup_data["collections"][collection_name] = []
    
    return backup_data

async def upload_to_drive(file_content: bytes, filename: str):
    """Upload file to Google Drive"""
    try:
        service = get_drive_service()
        folder_id = os.environ.get("GOOGLE_DRIVE_FOLDER_ID")
        
        file_metadata = {
            'name': filename,
            'mimeType': 'application/json'
        }
        
        if folder_id:
            file_metadata['parents'] = [folder_id]
        
        media = MediaIoBaseUpload(
            io.BytesIO(file_content),
            mimetype='application/json',
            resumable=True
        )
        
        file = service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id,name,createdTime'
        ).execute()
        
        logging.info(f"File uploaded to Drive: {file.get('name')} (ID: {file.get('id')})")
        return file
        
    except Exception as e:
        logging.error(f"Error uploading to Drive: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao fazer upload para Google Drive: {str(e)}")

async def cleanup_old_backups(keep_days: int = 30):
    """Delete backups older than keep_days"""
    try:
        service = get_drive_service()
        folder_id = os.environ.get("GOOGLE_DRIVE_FOLDER_ID")
        
        cutoff_date = datetime.now(timezone.utc) - timedelta(days=keep_days)
        cutoff_str = cutoff_date.isoformat()
        
        query = f"name contains 'backup_' and createdTime < '{cutoff_str}'"
        if folder_id:
            query += f" and '{folder_id}' in parents"
        
        results = service.files().list(
            q=query,
            fields='files(id, name, createdTime)'
        ).execute()
        
        files = results.get('files', [])
        deleted_count = 0
        
        for file in files:
            try:
                service.files().delete(fileId=file['id']).execute()
                deleted_count += 1
                logging.info(f"Deleted old backup: {file['name']}")
            except Exception as e:
                logging.error(f"Error deleting {file['name']}: {e}")
        
        return deleted_count
        
    except Exception as e:
        logging.error(f"Error cleaning up old backups: {e}")
        return 0

@api_router.post("/backup/create")
@limiter.limit("5/hour")
async def create_backup(
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """Create a manual backup and upload to Google Drive"""
    try:
        # Export all data
        backup_data = await export_all_data()
        
        # Generate filename
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"backup_{timestamp}.json"
        
        # Convert to JSON bytes
        json_content = json.dumps(backup_data, indent=2, ensure_ascii=False).encode('utf-8')
        
        # Upload to Drive
        file_info = await upload_to_drive(json_content, filename)
        
        # Cleanup old backups
        deleted_count = await cleanup_old_backups(keep_days=30)
        
        return {
            "success": True,
            "message": "Backup criado com sucesso",
            "file": {
                "id": file_info.get('id'),
                "name": file_info.get('name'),
                "created_at": file_info.get('createdTime')
            },
            "old_backups_deleted": deleted_count
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        logging.error(f"Error creating backup: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao criar backup: {str(e)}")

@api_router.get("/backup/status")
async def get_backup_status(current_user: dict = Depends(get_current_user)):
    """Get backup system status - simplified for direct download"""
    return {
        "configured": True,
        "system": "direct_download",
        "message": "Sistema de backup por download direto ativado"
    }

@api_router.post("/backup/download")
@limiter.limit("10/hour")
async def download_backup(
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """Generate and download complete backup as JSON file"""
    try:
        # Export all data
        backup_data = await export_all_data()
        
        # Generate filename with timestamp
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"backup_echoshop_{timestamp}.json"
        
        # Convert to JSON string
        json_content = json.dumps(backup_data, indent=2, ensure_ascii=False)
        
        # Return as downloadable file
        return Response(
            content=json_content,
            media_type="application/json",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Type": "application/json; charset=utf-8"
            }
        )
        
    except Exception as e:
        logging.error(f"Error creating backup download: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao gerar backup: {str(e)}")

# ==================== GOOGLE DRIVE OAUTH ENDPOINTS ====================

@api_router.get("/oauth/drive/connect")
async def connect_drive(current_user: dict = Depends(get_current_user)):
    """Initiate Google Drive OAuth flow"""
    try:
        redirect_uri = os.environ.get("GOOGLE_DRIVE_REDIRECT_URI")
        
        flow = Flow.from_client_config(
            {
                "web": {
                    "client_id": os.environ.get("GOOGLE_CLIENT_ID"),
                    "client_secret": os.environ.get("GOOGLE_CLIENT_SECRET"),
                    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                    "token_uri": "https://oauth2.googleapis.com/token",
                    "redirect_uris": [redirect_uri]
                }
            },
            scopes=['https://www.googleapis.com/auth/drive.file'],
            redirect_uri=redirect_uri
        )
        
        authorization_url, state = flow.authorization_url(
            access_type='offline',
            include_granted_scopes='true',
            prompt='consent',
            state=current_user.get("id")
        )
        
        logging.info(f"Drive OAuth initiated for user {current_user.get('id')}")
        return {"authorization_url": authorization_url}
    
    except Exception as e:
        logging.error(f"Failed to initiate OAuth: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Falha ao iniciar OAuth: {str(e)}")

@api_router.get("/oauth/drive/callback")
async def drive_callback(code: str, state: str):
    """Handle Google Drive OAuth callback"""
    try:
        redirect_uri = os.environ.get("GOOGLE_DRIVE_REDIRECT_URI")
        
        flow = Flow.from_client_config(
            {
                "web": {
                    "client_id": os.environ.get("GOOGLE_CLIENT_ID"),
                    "client_secret": os.environ.get("GOOGLE_CLIENT_SECRET"),
                    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                    "token_uri": "https://oauth2.googleapis.com/token",
                    "redirect_uris": [redirect_uri]
                }
            },
            scopes=None,
            redirect_uri=redirect_uri
        )
        
        flow.fetch_token(code=code)
        credentials = flow.credentials
        
        logging.info(f"Drive credentials obtained for user {state}, scopes: {credentials.scopes}")

        # Validate required scopes
        required_scopes = {"https://www.googleapis.com/auth/drive.file"}
        granted_scopes = set(credentials.scopes or [])
        if not required_scopes.issubset(granted_scopes):
            missing = required_scopes - granted_scopes
            logging.error(f"Missing required Drive scopes: {missing}")
            raise HTTPException(
                status_code=400,
                detail=f"Escopos necessários não concedidos: {', '.join(missing)}"
            )
        
        # Store credentials in database
        await db.drive_credentials.update_one(
            {"user_id": state},
            {"$set": {
                "user_id": state,
                "access_token": credentials.token,
                "refresh_token": credentials.refresh_token,
                "token_uri": credentials.token_uri,
                "client_id": credentials.client_id,
                "client_secret": credentials.client_secret,
                "scopes": credentials.scopes,
                "expiry": credentials.expiry.isoformat() if credentials.expiry else None,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }},
            upsert=True
        )
        
        logging.info(f"Drive credentials stored for user {state}")
        
        # Redirect to frontend
        frontend_url = os.environ.get("FRONTEND_URL")
        return RedirectResponse(url=f"{frontend_url}/configuracoes/backup?drive_connected=true")
    
    except Exception as e:
        logging.error(f"OAuth callback failed: {str(e)}")
        frontend_url = os.environ.get("FRONTEND_URL")
        return RedirectResponse(url=f"{frontend_url}/configuracoes/backup?drive_error={str(e)}")

async def get_drive_service(user_id: str):
    """Get Google Drive service with auto-refresh credentials"""
    creds_doc = await db.drive_credentials.find_one({"user_id": user_id})
    if not creds_doc:
        raise HTTPException(
            status_code=400, 
            detail="Google Drive não conectado. Por favor, conecte sua conta primeiro."
        )
    
    # Create credentials object
    creds = Credentials(
        token=creds_doc["access_token"],
        refresh_token=creds_doc.get("refresh_token"),
        token_uri=creds_doc["token_uri"],
        client_id=creds_doc["client_id"],
        client_secret=creds_doc["client_secret"],
        scopes=creds_doc["scopes"]
    )
    
    # Auto-refresh if expired
    if creds.expired and creds.refresh_token:
        logging.info(f"Refreshing expired token for user {user_id}")
        creds.refresh(GoogleRequest())
        
        # Update in database
        await db.drive_credentials.update_one(
            {"user_id": user_id},
            {"$set": {
                "access_token": creds.token,
                "expiry": creds.expiry.isoformat() if creds.expiry else None,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    return build('drive', 'v3', credentials=creds)

@api_router.get("/oauth/drive/status")
async def drive_status(current_user: dict = Depends(get_current_user)):
    """Check if user has connected Google Drive"""
    try:
        creds_doc = await db.drive_credentials.find_one({"user_id": current_user.get("id")})
        
        if not creds_doc:
            return {
                "connected": False,
                "email": None
            }
        
        # Try to get user email from Drive API
        try:
            service = await get_drive_service(current_user.get("id"))
            about = service.about().get(fields="user").execute()
            email = about.get('user', {}).get('emailAddress')
            
            return {
                "connected": True,
                "email": email,
                "expires_at": creds_doc.get("expiry")
            }
        except Exception as e:
            logging.error(f"Failed to get Drive user info: {e}")
            return {
                "connected": True,
                "email": None,
                "expires_at": creds_doc.get("expiry")
            }
    
    except Exception as e:
        logging.error(f"Error checking drive status: {e}")
        return {"connected": False, "email": None}

@api_router.post("/backup/upload-to-drive")
@limiter.limit("5/hour")
async def upload_backup_to_drive(
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """Upload backup to user's Google Drive"""
    try:
        # Get Drive service
        service = await get_drive_service(current_user.get("id"))
        
        # Export all data
        backup_data = await export_all_data()
        
        # Generate filename with timestamp
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"backup_echoshop_{timestamp}.json"
        
        # Convert to JSON bytes
        json_content = json.dumps(backup_data, indent=2, ensure_ascii=False).encode('utf-8')
        
        # Create file metadata
        file_metadata = {
            'name': filename,
            'mimeType': 'application/json'
        }
        
        # Upload to Drive
        media = MediaIoBaseUpload(
            io.BytesIO(json_content),
            mimetype='application/json',
            resumable=True
        )
        
        file = service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id,name,createdTime,webViewLink'
        ).execute()
        
        logging.info(f"Backup uploaded to Drive: {file.get('name')} (ID: {file.get('id')})")
        
        return {
            "success": True,
            "message": "Backup enviado para o Google Drive com sucesso!",
            "file": {
                "id": file.get('id'),
                "name": file.get('name'),
                "created_at": file.get('createdTime'),
                "link": file.get('webViewLink')
            }
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        logging.error(f"Error uploading backup to Drive: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao enviar backup para Drive: {str(e)}")

@api_router.delete("/oauth/drive/disconnect")
async def disconnect_drive(current_user: dict = Depends(get_current_user)):
    """Disconnect user's Google Drive"""
    try:
        result = await db.drive_credentials.delete_one({"user_id": current_user.get("id")})
        
        if result.deleted_count > 0:
            return {"success": True, "message": "Google Drive desconectado com sucesso"}
        else:
            return {"success": False, "message": "Nenhuma conexão encontrada"}
    
    except Exception as e:
        logging.error(f"Error disconnecting Drive: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao desconectar: {str(e)}")

    except Exception as e:
        logging.error(f"Error creating backup download: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao gerar backup: {str(e)}")

# ==================== VENDAS ENDPOINTS ====================

# PLANOS DE INTERNET
@api_router.get("/empresas/{empresa_id}/planos-internet")
async def get_planos_internet(empresa_id: str, current_user: dict = Depends(get_current_user)):
    """List all internet plans"""
    planos = await db.planos_internet.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(100)
    return planos

@api_router.post("/empresas/{empresa_id}/planos-internet")
async def create_plano_internet(
    empresa_id: str,
    plano: PlanoInternetCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create new internet plan"""
    plano_dict = plano.dict()
    plano_dict["empresa_id"] = empresa_id
    plano_dict["id"] = str(uuid.uuid4())
    plano_dict["created_at"] = datetime.now(timezone.utc)
    plano_dict["ativo"] = True
    
    await db.planos_internet.insert_one(plano_dict)
    return plano_dict

@api_router.put("/planos-internet/{plano_id}")
async def update_plano_internet(
    plano_id: str,
    plano: PlanoInternetCreate,
    current_user: dict = Depends(get_current_user)
):
    """Update internet plan"""
    result = await db.planos_internet.update_one(
        {"id": plano_id},
        {"$set": plano.dict()}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Plano não encontrado")
    
    return {"message": "Plano atualizado com sucesso"}

@api_router.delete("/planos-internet/{plano_id}")
async def delete_plano_internet(plano_id: str, current_user: dict = Depends(get_current_user)):
    """Delete internet plan"""
    result = await db.planos_internet.delete_one({"id": plano_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Plano não encontrado")
    
    return {"message": "Plano deletado com sucesso"}

# CLIENTES DE VENDAS
@api_router.get("/empresas/{empresa_id}/clientes-venda")
async def get_clientes_venda(empresa_id: str, current_user: dict = Depends(get_current_user)):
    """List all sales clients"""
    clientes = await db.clientes_venda.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(1000)
    return clientes

@api_router.post("/empresas/{empresa_id}/clientes-venda")
async def create_cliente_venda(
    empresa_id: str,
    cliente: ClienteVendaCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create new sales client"""
    cliente_dict = cliente.dict()
    cliente_dict["empresa_id"] = empresa_id
    cliente_dict["id"] = str(uuid.uuid4())
    cliente_dict["status"] = "ativo"
    cliente_dict["inadimplente"] = False
    cliente_dict["created_at"] = datetime.now(timezone.utc)
    cliente_dict["updated_at"] = datetime.now(timezone.utc)
    
    await db.clientes_venda.insert_one(cliente_dict)
    return cliente_dict

@api_router.get("/clientes-venda/{cliente_id}")
async def get_cliente_venda(cliente_id: str, current_user: dict = Depends(get_current_user)):
    """Get single sales client"""
    cliente = await db.clientes_venda.find_one({"id": cliente_id}, {"_id": 0})
    
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    
    return cliente

@api_router.put("/clientes-venda/{cliente_id}")
async def update_cliente_venda(
    cliente_id: str,
    cliente: ClienteVendaCreate,
    current_user: dict = Depends(get_current_user)
):
    """Update sales client"""
    cliente_dict = cliente.dict()
    cliente_dict["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.clientes_venda.update_one(
        {"id": cliente_id},
        {"$set": cliente_dict}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    
    return {"message": "Cliente atualizado com sucesso"}

@api_router.patch("/clientes-venda/{cliente_id}/status")
async def update_status_cliente(
    cliente_id: str,
    status: str = Body(..., embed=True),
    current_user: dict = Depends(get_current_user)
):
    """Update client status (ativo, suspenso, cancelado)"""
    if status not in ["ativo", "suspenso", "cancelado"]:
        raise HTTPException(status_code=400, detail="Status inválido")
    
    result = await db.clientes_venda.update_one(
        {"id": cliente_id},
        {"$set": {"status": status, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    
    return {"message": f"Status atualizado para {status}"}

# VENDAS / CONTRATOS
@api_router.get("/empresas/{empresa_id}/vendas")
async def get_vendas(empresa_id: str, current_user: dict = Depends(get_current_user)):
    """List all sales/contracts with client and plan details"""
    vendas = await db.vendas.find({"empresa_id": empresa_id}, {"_id": 0}).to_list(1000)
    
    # Enrich with client and plan data
    for venda in vendas:
        cliente = await db.clientes_venda.find_one({"id": venda["cliente_id"]}, {"_id": 0})
        plano = await db.planos_internet.find_one({"id": venda["plano_id"]}, {"_id": 0})
        venda["cliente"] = cliente
        venda["plano"] = plano
    
    return vendas

@api_router.post("/empresas/{empresa_id}/vendas")
async def create_venda(
    empresa_id: str,
    venda: VendaContratoCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create new sale/contract"""
    # Verify client and plan exist
    cliente = await db.clientes_venda.find_one({"id": venda.cliente_id})
    plano = await db.planos_internet.find_one({"id": venda.plano_id})
    
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    if not plano:
        raise HTTPException(status_code=404, detail="Plano não encontrado")
    
    venda_dict = venda.dict()
    venda_dict["empresa_id"] = empresa_id
    venda_dict["id"] = str(uuid.uuid4())
    venda_dict["status"] = "ativo"
    venda_dict["valor_mensalidade"] = plano["preco_mensal"]
    venda_dict["created_at"] = datetime.now(timezone.utc)
    venda_dict["updated_at"] = datetime.now(timezone.utc)
    
    await db.vendas.insert_one(venda_dict)
    
    return {
        **venda_dict,
        "cliente": cliente,
        "plano": plano
    }

@api_router.patch("/vendas/{venda_id}/status")
async def update_status_venda(
    venda_id: str,
    status: str = Body(..., embed=True),
    current_user: dict = Depends(get_current_user)
):
    """Update sale status (ativo, suspenso, cancelado)"""
    if status not in ["ativo", "suspenso", "cancelado"]:
        raise HTTPException(status_code=400, detail="Status inválido")
    
    result = await db.vendas.update_one(
        {"id": venda_id},
        {"$set": {"status": status, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Venda não encontrada")
    
    return {"message": f"Status da venda atualizado para {status}"}

# FATURAS
@api_router.get("/empresas/{empresa_id}/faturas")
async def get_faturas(
    empresa_id: str,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """List all invoices"""
    query = {"empresa_id": empresa_id}
    if status:
        query["status"] = status
    
    faturas = await db.faturas.find(query, {"_id": 0}).sort("data_vencimento", -1).to_list(1000)
    
    # Enrich with client data
    for fatura in faturas:
        cliente = await db.clientes_venda.find_one({"id": fatura["cliente_id"]}, {"_id": 0})
        fatura["cliente"] = cliente
    
    return faturas

@api_router.post("/empresas/{empresa_id}/faturas")
async def create_fatura(
    empresa_id: str,
    fatura: FaturaCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create new invoice"""
    # Verify sale exists
    venda = await db.vendas.find_one({"id": fatura.venda_id})
    
    if not venda:
        raise HTTPException(status_code=404, detail="Venda não encontrada")
    
    fatura_dict = fatura.dict()
    fatura_dict["empresa_id"] = empresa_id
    fatura_dict["id"] = str(uuid.uuid4())
    fatura_dict["cliente_id"] = venda["cliente_id"]
    fatura_dict["status"] = "pendente"
    fatura_dict["multa"] = 0.0
    fatura_dict["juros"] = 0.0
    fatura_dict["valor_total"] = fatura.valor
    fatura_dict["emails_enviados"] = []
    fatura_dict["created_at"] = datetime.now(timezone.utc)
    fatura_dict["updated_at"] = datetime.now(timezone.utc)
    
    await db.faturas.insert_one(fatura_dict)
    return fatura_dict

@api_router.patch("/faturas/{fatura_id}/status")
async def update_status_fatura(
    fatura_id: str,
    status: str = Body(..., embed=True),
    current_user: dict = Depends(get_current_user)
):
    """Update invoice status"""
    if status not in ["pendente", "pago", "vencido", "cancelado"]:
        raise HTTPException(status_code=400, detail="Status inválido")
    
    update_data = {"status": status, "updated_at": datetime.now(timezone.utc)}
    
    if status == "pago":
        update_data["data_pagamento"] = datetime.now(timezone.utc).date().isoformat()
    
    result = await db.faturas.update_one(
        {"id": fatura_id},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Fatura não encontrada")
    
    return {"message": f"Status da fatura atualizado para {status}"}

# CONFIGURAÇÃO DE COBRANÇA
@api_router.get("/empresas/{empresa_id}/configuracao-cobranca")
async def get_configuracao_cobranca(empresa_id: str, current_user: dict = Depends(get_current_user)):
    """Get billing configuration"""
    config = await db.configuracao_cobranca.find_one({"empresa_id": empresa_id}, {"_id": 0})
    
    if not config:
        # Return default config
        return {
            "empresa_id": empresa_id,
            "ambiente": "sandbox",
            "percentual_multa": 2.0,
            "percentual_juros_dia": 0.033,
            "dias_envio_antecipado": [15, 10, 5, 0]
        }
    
    return config

@api_router.post("/empresas/{empresa_id}/configuracao-cobranca")
async def create_or_update_configuracao_cobranca(
    empresa_id: str,
    config: ConfiguracaoCobrancaCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create or update billing configuration"""
    config_dict = config.dict()
    config_dict["empresa_id"] = empresa_id
    config_dict["updated_at"] = datetime.now(timezone.utc)
    
    # Check if exists
    existing = await db.configuracao_cobranca.find_one({"empresa_id": empresa_id})
    
    if existing:
        # Update
        await db.configuracao_cobranca.update_one(
            {"empresa_id": empresa_id},
            {"$set": config_dict}
        )
        return {"message": "Configuração atualizada com sucesso"}
    else:
        # Create
        config_dict["id"] = str(uuid.uuid4())
        config_dict["created_at"] = datetime.now(timezone.utc)
        await db.configuracao_cobranca.insert_one(config_dict)
        return {"message": "Configuração criada com sucesso"}

# ==================== END VENDAS ENDPOINTS ====================

# Include router
app.include_router(api_router)

# Health check endpoints
@app.get("/")
async def root():
    """Root endpoint - redirects to health check"""
    return {"status": "ok", "message": "ECHO SHOP FinAI Backend", "version": "1.0"}

@app.get("/health")
async def health_check_root():
    """Health check endpoint at root level for K8s probes"""
    return {"status": "healthy", "service": "finai-backend", "timestamp": datetime.now(timezone.utc).isoformat()}

@app.get("/api/health")
async def health_check():
    """Health check endpoint for monitoring"""
    return {"status": "healthy", "service": "finai-backend", "timestamp": datetime.now(timezone.utc).isoformat()}

@app.get("/readiness")
async def readiness_check():
    """Readiness probe endpoint - checks if app is ready to serve traffic"""
    try:
        # Quick DB connectivity check
        await db.command('ping')
        return {"status": "ready", "database": "connected"}
    except Exception as e:
        return {"status": "not_ready", "database": "disconnected", "error": str(e)}


# Security headers middleware
@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    response = await call_next(request)
    
    # Security headers
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    response.headers["Content-Security-Policy"] = "default-src 'self'; script-src 'self' 'unsafe-inline'; style-src 'self' 'unsafe-inline'; img-src 'self' data: https:;"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
    
    return response

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type"],
    max_age=3600,
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    try:
        if scheduler and hasattr(scheduler, '_eventloop') and scheduler._eventloop:
            scheduler.shutdown(wait=False)
    except:
        pass
    client.close()

@app.on_event("startup")
async def startup_event():
    """Initialize scheduled backup on startup and validate critical environment variables"""
    # Validate critical environment variables at runtime
    if JWT_SECRET == 'temp-build-secret':
        raise RuntimeError(
            "JWT_SECRET environment variable is required for production. "
            "Please set it in your deployment configuration."
        )
    
    if WHATSAPP_SERVICE_KEY == 'temp-build-key':
        raise RuntimeError(
            "WHATSAPP_SERVICE_KEY environment variable is required for production. "
            "Please set it in your deployment configuration."
        )
    
    logging.info("✓ Environment variables validated successfully")
    
    # Initialize scheduler in background to avoid blocking startup
    service_account_path = os.environ.get("GOOGLE_SERVICE_ACCOUNT_PATH", "/app/backend/service_account.json")
    
    # Only schedule if Google Drive is configured
    if os.path.exists(service_account_path):
        try:
            # Schedule daily backup at 3 AM
            scheduler.add_job(
                run_scheduled_backup,
                CronTrigger(hour=3, minute=0),
                id='daily_backup',
                name='Daily automated backup to Google Drive',
                replace_existing=True
            )
            # Start scheduler in non-blocking way
            if not scheduler.running:
                scheduler.start()
            logging.info("✓ Automated backup scheduler started - Daily backups at 3:00 AM")
        except Exception as e:
            logging.warning(f"⚠ Failed to start backup scheduler: {e}")
            logging.info("Application will continue without automated backups")
    else:
        logging.warning("⚠ Google Drive not configured - Automated backups disabled")
        logging.info("See /app/BACKUP_SETUP_INSTRUCTIONS.md for setup instructions")

async def run_scheduled_backup():
    """Function to run scheduled backup"""
    try:
        logging.info("=" * 60)
        logging.info("SCHEDULED BACKUP STARTING")
        logging.info("=" * 60)
        
        # Export all data
        backup_data = await export_all_data()
        
        # Generate filename
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"backup_{timestamp}.json"
        
        # Convert to JSON bytes
        json_content = json.dumps(backup_data, indent=2, ensure_ascii=False).encode('utf-8')
        file_size_mb = len(json_content) / (1024 * 1024)
        logging.info(f"Backup file size: {file_size_mb:.2f} MB")
        
        # Upload to Drive
        file_info = await upload_to_drive(json_content, filename)
        
        # Cleanup old backups
        deleted_count = await cleanup_old_backups(keep_days=30)
        
        logging.info("=" * 60)
        logging.info("✓ SCHEDULED BACKUP COMPLETED")
        logging.info(f"  File: {file_info.get('name')}")
        logging.info(f"  Drive ID: {file_info.get('id')}")
        logging.info(f"  Old backups deleted: {deleted_count}")
        logging.info("=" * 60)
        
    except Exception as e:
        logging.error("=" * 60)
        logging.error("✗ SCHEDULED BACKUP FAILED")
        logging.error(f"  Error: {str(e)}")
        logging.error("=" * 60)